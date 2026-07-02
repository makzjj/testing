"""IPQC Excel workbook adapter for template-based production reporting."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from utils.deployment_paths import get_runtime_exports_dir

try:
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.workbook.workbook import Workbook
except ImportError:  # pragma: no cover - guarded at runtime.
    Workbook = Any  # type: ignore[assignment]
    Worksheet = Any  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]


@dataclass(frozen=True)
class IpqcExpectedSummary:
    operator: str = ""
    assembler: str = ""
    serial_number: str = ""
    pwm: str = ""
    other_parameters: str = ""


@dataclass(frozen=True)
class ProductionWorkbookMetadata:
    operator_name: str = ""
    assembler_name: str = ""


@dataclass(frozen=True)
class SamplingWorkbookLayout:
    sheet_name: str
    section_headers: dict[str, int]
    row_lookup: dict[tuple[str, int, str], int]
    raw_labels: list[str]

    def resolve_row(self, section_name: str, pwm_value: int, direction: str) -> int:
        section_key = " ".join(str(section_name).strip().casefold().split())
        direction_key = str(direction).strip()
        if direction_key not in {"+", "-"}:
            raise ValueError(f"Unsupported sampling direction '{direction}'.")
        row = self.row_lookup.get((section_key, int(pwm_value), direction_key))
        if row is None:
            raise ValueError(
                f"Unsupported sampling row for section '{section_name}', PWM {pwm_value}, direction '{direction}'."
            )
        return row


class IpqcExcelAdapter:
    """Loads an IPQC workbook template and reads/writes summary-sheet values."""

    _PROGRAMMING_ROW_LOOKUP: dict[str, int] = {
        "operator": 3,
        "uuid": 5,
        "pwm": 6,
        "proportionate (p)": 7,
        "pid_p": 7,
        "integral (i)": 8,
        "pid_i": 8,
        "derivative (d)": 9,
        "pid_d": 9,
        "pid_slewrate": 10,
        "rampdown_slope": 11,
        "rampdown_step": 12,
        "rampdown_minvel": 13,
        "rampdown_targetoffset": 14,
        "rampdown_region": 15,
        "acceptable_error": 16,
    }

    _SUMMARY_PARAMETER_ALIASES: dict[str, int] = {
        "s/n": 4,
        "sn": 4,
        "serial": 4,
        "serial number": 4,
        "uuid": 4,
        "pwm": 5,
        "other parameters": 6,
        "other parameter": 6,
        "other": 6,
    }

    def __init__(self) -> None:
        self._template_path: Path | None = None
        self._workbook: Workbook | None = None
        self._base_groups: list[str] = []
        self._active_group: str | None = None
        self._last_output_path: Path | None = None
        self._workbook_session_id = 0

    @property
    def template_path(self) -> Path | None:
        return self._template_path

    @property
    def last_output_path(self) -> Path | None:
        return self._last_output_path

    @property
    def workbook_session_id(self) -> int:
        return int(self._workbook_session_id)

    @property
    def available_base_sheet_groups(self) -> list[str]:
        return list(self._base_groups)

    @property
    def active_sheet_group(self) -> str | None:
        return self._active_group

    def has_loaded_workbook(self) -> bool:
        return self._workbook is not None and self._template_path is not None

    def load_template(self, path: str | Path) -> list[str]:
        self._ensure_openpyxl_available()
        template_path = Path(path).expanduser().resolve()
        workbook = load_workbook(filename=template_path)  # type: ignore[misc]
        base_groups = self._detect_base_sheet_groups(workbook.sheetnames)
        if not base_groups:
            raise ValueError("No base IPQC sheet groups were found in workbook.")

        self._template_path = template_path
        self._workbook = workbook
        self._base_groups = base_groups
        self._active_group = base_groups[0]
        self._last_output_path = None
        self._workbook_session_id += 1
        return list(base_groups)

    def select_sheet_group(self, base_group: str) -> None:
        workbook = self._require_workbook()
        if base_group not in self._base_groups:
            raise ValueError(f"Sheet group '{base_group}' is not available in the loaded workbook.")
        if base_group not in workbook.sheetnames:
            raise ValueError(f"Base sheet '{base_group}' is missing from workbook.")
        self._active_group = base_group

    def read_expected_summary(self, *, strict: bool = True) -> IpqcExpectedSummary:
        sheet = self._require_base_sheet()

        operator_row = self._find_programming_label_row(sheet, "Operator")
        assembler_row = self._find_programming_label_row(sheet, "Assembler")
        serial_row = self._find_programming_label_row(sheet, "UUID")
        pwm_row = self._find_programming_label_row(sheet, "PWM")
        operator = self._read_cell_text(sheet, f"B{operator_row}") if operator_row is not None else ""
        assembler = self._read_cell_text(sheet, f"B{assembler_row}") if assembler_row is not None else ""
        serial_number = self._read_cell_text(sheet, f"B{serial_row}") if serial_row is not None else ""
        pwm = self._read_cell_text(sheet, f"B{pwm_row}") if pwm_row is not None else ""
        other_parameters = self._read_cell_text(sheet, f"B{(pwm_row + 1) if pwm_row is not None else 6}")

        if strict:
            if not operator_row:
                raise ValueError(f"Expected operator is missing in sheet '{sheet.title}' column A.")
            if not serial_number:
                raise ValueError(f"Expected serial number/UUID is missing in sheet '{sheet.title}'.")
            if not pwm:
                raise ValueError(f"Expected PWM is missing in sheet '{sheet.title}'.")

        return IpqcExpectedSummary(
            operator=operator,
            assembler=assembler,
            serial_number=serial_number,
            pwm=pwm,
            other_parameters=other_parameters,
        )

    def read_expected_uuid_serial(self) -> str:
        """Read expected UUID/S/N from the label-based Programming sheet."""
        sheet = self._require_base_sheet()
        return self._read_programming_label_value(sheet, "UUID")

    def read_expected_pwm_value(self) -> str:
        """Read expected PWM from the label-based Programming sheet."""
        sheet = self._require_base_sheet()
        return self._read_programming_label_value(sheet, "PWM")

    def read_production_metadata(self) -> ProductionWorkbookMetadata:
        """Read operator and assembler metadata from the active Programming sheet."""
        sheet = self._require_base_sheet()
        operator = self._read_programming_label_value(sheet, "Operator")
        assembler = self._read_programming_label_value(sheet, "Assembler")
        return ProductionWorkbookMetadata(operator_name=operator, assembler_name=assembler)

    def write_production_metadata(self, operator_name: str, assembler_name: str) -> None:
        """Persist operator and assembler names in the active Programming sheet."""
        sheet = self._require_base_sheet()
        operator_row = self._find_programming_label_row(sheet, "Operator")
        if operator_row is None:
            raise ValueError(f"Programming sheet '{sheet.title}' is missing an Operator row.")
        assembler_row = self._find_programming_label_row(sheet, "Assembler")
        if assembler_row is None:
            sheet.insert_rows(operator_row + 1, 1)
            assembler_row = operator_row + 1
            sheet[f"A{assembler_row}"] = "Assembler"
        sheet[f"B{operator_row}"] = str(operator_name).strip()
        sheet[f"B{assembler_row}"] = str(assembler_name).strip()
        sheet[f"C{operator_row}"] = ""
        sheet[f"D{operator_row}"] = ""
        sheet[f"C{assembler_row}"] = ""
        sheet[f"D{assembler_row}"] = ""

    def read_cell_text(self, cell_ref: str) -> str:
        """Read one text cell from the active base sheet."""
        sheet = self._require_base_sheet()
        return self._read_cell_text(sheet, cell_ref)

    def resolve_programming_row(self, parameter_name: str) -> int:
        normalized = self._normalize_programming_label(parameter_name)
        discovered_rows, _raw_labels = self.discover_programming_parameter_rows()
        row = discovered_rows.get(normalized)
        if row is None:
            row = self._PROGRAMMING_ROW_LOOKUP.get(normalized)
        if row is not None:
            return row
        raise ValueError(f"Unsupported programming parameter '{parameter_name}' (normalized: '{normalized}').")

    def read_programming_parameter_source(self, parameter_name: str) -> str:
        sheet = self._require_base_sheet()
        row = self.resolve_programming_row(parameter_name)
        return self._read_cell_text(sheet, f"B{row}")

    def discover_programming_parameter_rows(self) -> tuple[dict[str, int], list[str]]:
        """Find supported programming rows by scanning Column A labels."""
        sheet = self._require_base_sheet()
        discovered_rows: dict[str, int] = {}
        raw_labels: list[str] = []
        for row in range(3, sheet.max_row + 1):
            label = self._read_cell_text(sheet, f"A{row}")
            if not label:
                continue
            raw_labels.append(label)
            normalized = self._normalize_programming_label(label)
            if normalized in self._PROGRAMMING_ROW_LOOKUP:
                discovered_rows[normalized] = row
        return discovered_rows, raw_labels

    def resolve_sampling_sheet_name(self, base_group: str | None = None) -> str:
        group = self._active_group if base_group is None else base_group
        if not group:
            raise RuntimeError("No IPQC sheet group is selected.")
        return f"{group}_D"

    def sample_index_to_column(self, sample_index: int) -> str:
        if get_column_letter is None:
            raise RuntimeError("openpyxl is required for sampling workbook support but is not installed.")
        index = int(sample_index)
        if index < 1 or index > 32:
            raise ValueError(f"Sampling index must be between 1 and 32, got {sample_index}.")
        return get_column_letter(index + 1)

    def discover_sampling_layout(self, base_group: str | None = None) -> SamplingWorkbookLayout:
        workbook = self._require_workbook()
        sheet_name = self.resolve_sampling_sheet_name(base_group)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sampling sheet '{sheet_name}' is missing from workbook.")
        sheet = workbook[sheet_name]
        section_headers: dict[str, int] = {}
        row_lookup: dict[tuple[str, int, str], int] = {}
        raw_labels: list[str] = []
        current_section: str | None = None
        section_pattern = re.compile(r"^(range|speed|time)$", re.IGNORECASE)
        pwm_pattern = re.compile(r"^pwm\s*(\d+)$", re.IGNORECASE)
        sample_pattern = re.compile(r"^([+-])\s*(\d+)$")

        for row in range(1, sheet.max_row + 1):
            label = self._read_cell_text(sheet, f"A{row}")
            if not label:
                continue
            raw_labels.append(label)
            normalized = self._normalize_programming_label(label)
            if section_pattern.match(normalized):
                current_section = normalized
                section_headers[current_section] = row
                continue
            if current_section is None:
                continue
            if pwm_pattern.match(normalized):
                continue
            sample_match = sample_pattern.match(label.strip())
            if not sample_match:
                continue
            direction = sample_match.group(1)
            pwm_value = int(sample_match.group(2))
            row_lookup[(current_section, pwm_value, direction)] = row

        if not row_lookup:
            raise ValueError(
                f"No supported sampling rows were found in sheet '{sheet_name}'. Labels found: {raw_labels}"
            )
        return SamplingWorkbookLayout(
            sheet_name=sheet_name,
            section_headers=section_headers,
            row_lookup=row_lookup,
            raw_labels=raw_labels,
        )

    def write_sampling_result(
        self,
        section_name: str,
        pwm_value: int,
        direction: str,
        sample_index: int,
        actual_value: object,
        *,
        base_group: str | None = None,
    ) -> str:
        workbook = self._require_workbook()
        layout = self.discover_sampling_layout(base_group)
        sheet = workbook[layout.sheet_name]
        row = layout.resolve_row(section_name, pwm_value, direction)
        column = self.sample_index_to_column(sample_index)
        cell_ref = f"{column}{row}"
        sheet[cell_ref] = actual_value
        return cell_ref

    def write_programming_parameter_result(self, parameter_name: str, actual_value: object, check_result: str) -> None:
        sheet = self._require_base_sheet()
        row = self.resolve_programming_row(parameter_name)
        sheet[f"C{row}"] = "" if actual_value is None else str(actual_value)
        sheet[f"D{row}"] = str(check_result)

    def write_summary_result(self, parameter_name: str, actual_value: object, check_result: str) -> None:
        sheet = self._require_base_sheet()
        row = self._resolve_summary_row(parameter_name)
        sheet[f"C{row}"] = "" if actual_value is None else str(actual_value)
        sheet[f"D{row}"] = str(check_result)

    def write_parameter_result(
        self,
        actual_cell: str,
        result_cell: str,
        actual_value: object,
        check_result: str,
    ) -> None:
        """Write one parameter's actual/result cells on the active base sheet."""
        sheet = self._require_base_sheet()
        sheet[actual_cell] = "" if actual_value is None else str(actual_value)
        sheet[result_cell] = str(check_result)

    def _resolve_summary_row(self, parameter_name: str) -> int:
        normalized = self._normalize_programming_label(parameter_name)
        row = self._SUMMARY_PARAMETER_ALIASES.get(normalized)
        if row is None:
            raise ValueError(
                f"Unsupported summary parameter '{parameter_name}' (normalized: '{normalized}')."
            )
        return row

    @staticmethod
    def _normalize_programming_label(value: str) -> str:
        return " ".join(str(value).strip().casefold().split())

    def _read_programming_label_value(self, sheet: Worksheet, label: str) -> str:
        row = self._find_programming_label_row(sheet, label)
        if row is None:
            return ""
        return self._read_cell_text(sheet, f"B{row}")

    def _find_programming_label_row(self, sheet: Worksheet, label: str) -> int | None:
        normalized_label = self._normalize_programming_label(label)
        for row in range(3, sheet.max_row + 1):
            cell_label = self._read_cell_text(sheet, f"A{row}")
            if self._normalize_programming_label(cell_label) == normalized_label:
                return row
        return None

    def suggest_completed_output_path(self) -> Path:
        template_path = self._require_template_path()
        active_group = self._require_active_group()
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S") + "Z"
        base_name = f"{template_path.stem}_{active_group}_completed_{stamp}"
        candidate = get_runtime_exports_dir() / f"{base_name}{template_path.suffix}"
        index = 1
        while candidate.exists():
            candidate = get_runtime_exports_dir() / f"{base_name}_{index}{template_path.suffix}"
            index += 1
        return candidate

    def save_completed_workbook(self, output_path: str | Path) -> Path:
        workbook = self._require_workbook()
        template_path = self._require_template_path()
        target = Path(output_path).expanduser().resolve()
        if target == template_path:
            raise ValueError("Refusing to overwrite the original workbook template.")
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(target)
        self._last_output_path = target
        return target

    def write_raw_samples(self, *_args: object, **_kwargs: object) -> None:
        raise NotImplementedError("Raw sample writing is reserved for a later phase.")

    def write_analysis_values(self, *_args: object, **_kwargs: object) -> None:
        raise NotImplementedError("Analysis/statistics writing is reserved for a later phase.")

    def _require_template_path(self) -> Path:
        if self._template_path is None:
            raise RuntimeError("No IPQC workbook template is loaded.")
        return self._template_path

    def _require_workbook(self) -> Workbook:
        if self._workbook is None:
            raise RuntimeError("No IPQC workbook template is loaded.")
        return self._workbook

    def _require_active_group(self) -> str:
        if not self._active_group:
            raise RuntimeError("No IPQC sheet group is selected.")
        return self._active_group

    def _require_base_sheet(self) -> Worksheet:
        workbook = self._require_workbook()
        base_group = self._require_active_group()
        if base_group not in workbook.sheetnames:
            raise ValueError(f"Base sheet '{base_group}' is missing from workbook.")
        return workbook[base_group]

    @staticmethod
    def _read_cell_text(sheet: Worksheet, cell_ref: str) -> str:
        value = sheet[cell_ref].value
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _detect_base_sheet_groups(sheet_names: list[str]) -> list[str]:
        return sorted(name for name in sheet_names if not (name.endswith("_D") or name.endswith("_A")))

    @staticmethod
    def _ensure_openpyxl_available() -> None:
        if load_workbook is None:
            raise RuntimeError("openpyxl is required for IPQC workbook support but is not installed.")
