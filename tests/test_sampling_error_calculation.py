from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

try:
    from openpyxl import Workbook

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - environment dependent.
    _HAS_OPENPYXL = False

from data.binary_cmd_parser import decode_nodeconfig_motion_polarity
from gui.workspace.controllers.sampling_test_controller import SamplingTestConfig, SamplingTestController
from gui.workspace.models.node_motion_calibration import NodeMotionCalibration
from services.ipqc_excel_adapter import IpqcExcelAdapter
from services.node_motion_calibration_store import NodeMotionCalibrationStore
from services.node_sensor_profile import NodeSensorProfile


def _create_ipqc_workbook(path: Path) -> None:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to create IPQC workbook fixtures.")
    wb = Workbook()
    ws = wb.active
    ws.title = "3X"
    sampling_3x = wb.create_sheet("3X_D")
    wb.create_sheet("3X_A")
    ws["A1"] = "Programming"
    ws["B2"] = "Source"
    ws["C2"] = "Programmed"
    ws["D2"] = "Check"
    ws["A3"] = "Operator"
    ws["A4"] = "UUID"
    ws["A5"] = "PWM"
    ws["B4"] = "1223303010"
    ws["B5"] = "100"
    _populate_sampling_sheet(sampling_3x)
    wb.save(path)


def _populate_sampling_sheet(sheet) -> None:
    pwm_values = [100, 90, 80, 70, 60]
    section_starts = {"Range": 1, "Speed": 18, "Time": 35}
    for section_name, start_row in section_starts.items():
        sheet[f"A{start_row}"] = section_name
        row = start_row + 1
        for pwm in pwm_values:
            sheet[f"A{row}"] = f"PWM {pwm}"
            sheet[f"A{row + 1}"] = f"+{pwm}"
            sheet[f"A{row + 2}"] = f"-{pwm}"
            row += 3


def _build_calibration_store() -> NodeMotionCalibrationStore:
    return NodeMotionCalibrationStore(
        [
            NodeMotionCalibration(3, "X", "Linear", "mm", 32.0, 88064.0),
            NodeMotionCalibration(5, "V", "Rotational", "deg", 49.0, -34117.16),
            NodeMotionCalibration(7, "NZ", "Linear", "mm", 69.5, -80058.18),
            NodeMotionCalibration(8, "RZ", "Rotational", "deg", 119.0, 19569.78),
            NodeMotionCalibration(12, "Z", "Linear", "mm", 50.0, 88064.0),
        ],
        Path("config/node_motion_calibration.xml"),
    )


class _SamplingManualClock:
    def __init__(self, values: list[float]) -> None:
        self._values = list(values)
        self._index = 0

    def __call__(self) -> float:
        if self._index >= len(self._values):
            raise AssertionError("Sampling manual clock was exhausted.")
        value = self._values[self._index]
        self._index += 1
        return value


class _RecordingSamplingController(SamplingTestController):
    def __init__(
        self,
        adapter: IpqcExcelAdapter,
        config: SamplingTestConfig,
        clock,
        *,
        node_id: int,
        calibration_store: NodeMotionCalibrationStore,
    ) -> None:
        super().__init__(
            adapter,
            config,
            clock=clock,
            node_motion_calibration_store=calibration_store,
        )
        self.commands: list[list[int]] = []
        self.measurements: list[object] = []
        self.failures: list[str] = []
        raw_nodeconfig = 0x00
        polarity = decode_nodeconfig_motion_polarity(raw_nodeconfig, allow_unvalidated=True)
        self.set_motion_polarity(polarity)
        self.set_sensor_profile(NodeSensorProfile.from_node_context(node_id, polarity))

    def command_requested(self, payload: list[int]) -> None:
        self.commands.append(list(payload))

    def measurement_completed(self, result) -> None:
        self.measurements.append(result)

    def sampling_failed(self, reason: str) -> None:
        self.failures.append(reason)

    def log_message(self, text: str) -> None:
        return None

    def packet_message(self, text: str) -> None:
        return None

    def state_changed(self, text: str) -> None:
        return None

    def status_changed(self, text: str) -> None:
        return None

    def current_pwm_changed(self, pwm: int) -> None:
        return None

    def current_direction_changed(self, direction: str) -> None:
        return None

    def current_sample_changed(self, sample_index: int) -> None:
        return None

    def samples_completed_changed(self, completed: int, total: int) -> None:
        return None

    def latest_measurement_changed(self, range_value: int, elapsed_seconds: float, speed: float) -> None:
        return None

    def latest_workbook_cell_written(self, cell_ref: str) -> None:
        return None

    def sampling_completed(self) -> None:
        return None

    def sampling_aborted(self, reason: str) -> None:
        return None


@unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for Sampling error tests.")
class SamplingErrorCalculationTests(unittest.TestCase):
    def _build_adapter(self, tmpdir: str) -> IpqcExcelAdapter:
        workbook_path = Path(tmpdir) / "sampling_ipqc.xlsx"
        _create_ipqc_workbook(workbook_path)
        adapter = IpqcExcelAdapter()
        adapter.load_template(workbook_path)
        return adapter

    def _build_controller(
        self,
        tmpdir: str,
        *,
        node_id: int,
        clock_values: list[float],
        calibration_store: NodeMotionCalibrationStore | None = None,
    ) -> _RecordingSamplingController:
        return _RecordingSamplingController(
            self._build_adapter(tmpdir),
            SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
            _SamplingManualClock(clock_values),
            node_id=node_id,
            calibration_store=calibration_store or _build_calibration_store(),
        )

    @staticmethod
    def _getpos_packet(position: int) -> list[int]:
        return [0x82, *list(int(position).to_bytes(4, "big", signed=True))]

    def _drive_home_sequence(self, controller: _RecordingSamplingController, *, start_pos: int) -> None:
        self.assertEqual(controller.commands[0], [0x84, 0x00, 0x50])
        controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
        self.assertEqual(controller.commands[1], [0x81, 0x00, 0x00, 0x00, 0x00])
        controller.handle_runtime_packet([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x00])
        controller.handle_runtime_packet(self._getpos_packet(start_pos))

    def test_positive_error_counts_are_overshoot_for_linear_positive_counts_per_unit(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            controller = self._build_controller(tmpdir, node_id=3, clock_values=[1.0, 1.5])
            self.assertTrue(controller.start(3, "X"))
            self._drive_home_sequence(controller, start_pos=10)

            measured_range_counts = int(32.0 * 88064.0) + 33
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            self.assertIsNone(controller.last_result)
            controller.handle_runtime_packet(self._getpos_packet(10 + measured_range_counts))

            result = controller.measurements[0]
            self.assertEqual(result.range_value, measured_range_counts)
            self.assertAlmostEqual(result.expected_range_counts, 32.0 * 88064.0)
            self.assertAlmostEqual(result.error_counts, 33.0)
            self.assertAlmostEqual(result.error_units, 33.0 / 88064.0)
            self.assertEqual(result.error_unit, "mm")
            self.assertEqual(result.error_classification, "OVERSHOOT")
            self.assertAlmostEqual(result.elapsed_seconds, 0.5)
            self.assertAlmostEqual(result.speed, measured_range_counts / 0.5)

    def test_fractional_expected_counts_are_not_rounded_for_negative_counts_per_unit(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            controller = self._build_controller(tmpdir, node_id=5, clock_values=[2.0, 2.4])
            self.assertTrue(controller.start(5, "V"))
            self._drive_home_sequence(controller, start_pos=10)

            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x4C])
            self.assertIsNone(controller.last_result)
            controller.handle_runtime_packet(self._getpos_packet(1671750))

            result = controller.measurements[0]
            self.assertEqual(result.range_value, 1671740)
            self.assertAlmostEqual(result.expected_range_counts, 49.0 * abs(-34117.16))
            self.assertAlmostEqual(result.error_counts, 1671740.0 - (49.0 * abs(-34117.16)))
            self.assertAlmostEqual(
                result.error_units,
                (1671740.0 - (49.0 * abs(-34117.16))) / abs(-34117.16),
            )
            self.assertLess(result.error_counts, 0.0)
            self.assertLess(result.error_units, 0.0)
            self.assertEqual(result.error_unit, "deg")
            self.assertEqual(result.error_classification, "UNDERSHOOT")
            self.assertAlmostEqual(result.elapsed_seconds, 0.4)
            self.assertAlmostEqual(result.speed, 1671740.0 / 0.4)

    def test_zero_error_is_on_reference_and_uses_existing_measured_range(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            controller = self._build_controller(tmpdir, node_id=12, clock_values=[3.0, 3.25])
            self.assertTrue(controller.start(12, "Z"))
            self._drive_home_sequence(controller, start_pos=10)

            measured_range_counts = int(50.0 * 88064.0)
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            self.assertIsNone(controller.last_result)
            controller.handle_runtime_packet(self._getpos_packet(10 + measured_range_counts))

            result = controller.last_result
            assert result is not None
            self.assertEqual(result.range_value, measured_range_counts)
            self.assertAlmostEqual(result.expected_range_counts, float(measured_range_counts))
            self.assertAlmostEqual(result.error_counts, 0.0)
            self.assertAlmostEqual(result.error_units, 0.0)
            self.assertEqual(result.error_unit, "mm")
            self.assertEqual(result.error_classification, "ON_REFERENCE")
            self.assertEqual(len(controller.measurements), 1)

    def test_missing_calibration_fails_before_any_motion_command(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            calibration_store = NodeMotionCalibrationStore(
                [NodeMotionCalibration(3, "X", "Linear", "mm", 32.0, 88064.0)],
                Path("config/node_motion_calibration.xml"),
            )
            controller = self._build_controller(
                tmpdir,
                node_id=8,
                clock_values=[0.0, 0.5],
                calibration_store=calibration_store,
            )

            self.assertFalse(controller.start(8, "RZ"))
            self.assertEqual(controller.commands, [])
            self.assertEqual(controller.measurements, [])
            self.assertEqual(controller.failures, ["No motion calibration is configured for Node 8 - RZ."])
            self.assertIsNone(controller.last_result)

    def test_sampling_uses_store_injection_without_xml_parsing_or_popup_changes(self) -> None:
        controller_source = Path("gui/workspace/controllers/sampling_test_controller.py").read_text(encoding="utf-8")
        popup_source = Path("gui/workspace/dialogs/sampling_test_popup.py").read_text(encoding="utf-8")
        production_source = Path("gui/workspace/pages/production_page.py").read_text(encoding="utf-8")

        self.assertIn("node_motion_calibration_store=self._bridge.node_motion_calibration_store", production_source)
        self.assertNotIn("ElementTree", controller_source)
        self.assertNotIn("node_motion_calibration.xml", controller_source)
        self.assertNotIn("get_bundle_resource_path", controller_source)
        self.assertNotIn("NodeMotionCalibrationStore.load_default()", controller_source)
        self.assertNotIn("ElementTree", popup_source)
        self.assertNotIn("NodeMotionCalibrationStore", popup_source)
