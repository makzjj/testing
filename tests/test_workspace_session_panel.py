"""Focused tests for the global workspace Session panel and Production metadata flow."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QApplication, QDialog, QLabel, QPushButton

from gui.workspace.models import SessionState
from gui.workspace.pages.production_page import ProductionPage
from gui.workspace.shell.project_workspace_window import ProjectWorkspaceWindow
from gui.workspace.constants import ROUTE_FIRMWARE, ROUTE_MECHANICAL, ROUTE_PRODUCTION, ROUTE_PROJECT_CONFIG
from gui.workspace.widgets import LiveSessionPanel
from gui.main_window import MainWindow
from myconfig.project_models import ProjectDefinition, ProjectFeatures, ProjectUiConfig

try:
    from openpyxl import Workbook

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - environment dependent.
    _HAS_OPENPYXL = False

class _FakeBridge:
    def __init__(self) -> None:
        self._project_definition = ProjectDefinition(
            name="demo",
            display_name="Demo",
            config_path=Path("demo.yaml"),
        )

    @property
    def project_definition(self) -> ProjectDefinition:
        return self._project_definition

    @property
    def project_config_path(self) -> Path:
        return self._project_definition.config_path

    def get_runtime_communication_model(self, *, create_if_missing: bool = False) -> dict:
        return {"connected": False, "ports": [], "selected_port": "", "baud_rates": ["115200"], "selected_baud": "115200"}

    def get_runtime_robot_nodes(self, *, create_if_missing: bool = False) -> dict:
        return {"connected_nodes": [], "rows": []}

    def get_runtime_window(self, *, create_if_missing: bool = False):
        return None

    def get_runtime_robot_power_state(self, *, create_if_missing: bool = False) -> bool | None:
        return None

    def get_session_state(self, active_page: str) -> SessionState:
        return SessionState(
            project_name="Demo",
            connection_text="Offline",
            session_text="Preview mode",
            active_page=active_page,
            metadata_edit_enabled=False,
        )

    def get_boot_messages(self) -> list[str]:
        return []


class WorkspaceSessionPanelTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls._app = QApplication.instance() or QApplication([])

    def test_live_session_panel_shows_operator_assembler_page_and_edit_button(self) -> None:
        panel = LiveSessionPanel()
        state = SessionState(
            project_name="Demo",
            connection_text="Workbook loaded",
            session_text="Metadata ready",
            active_page="Production",
            operator_name="Alice",
            assembler_name="Bob",
            metadata_edit_enabled=True,
        )

        panel.update_state(state)

        self.assertEqual(panel.operator_value.text(), "Alice")
        self.assertEqual(panel.assembler_value.text(), "Bob")
        self.assertEqual(panel.page_value.text(), "Production")
        self.assertTrue(panel.edit_button.isEnabled())
        self.assertEqual(panel.findChild(QLabel, "LiveSessionTitle").text(), "Session")

    @staticmethod
    def _build_project_definition(config_path: Path) -> ProjectDefinition:
        return ProjectDefinition(
            name="demo",
            display_name="Demo",
            config_path=config_path,
            features=ProjectFeatures(firmware_tools=True),
            ui=ProjectUiConfig(workspace="phase2_shell"),
        )

    def _build_workspace_window(self) -> ProjectWorkspaceWindow:
        tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(tmpdir.cleanup)
        config_path = Path(tmpdir.name) / "demo.yaml"
        config_path.write_text(
            """
project:
  name: demo
  display_name: Demo
features:
  firmware_tools: true
ui:
  workspace: phase2_shell
""".strip(),
            encoding="utf-8",
        )
        return ProjectWorkspaceWindow(self._build_project_definition(config_path))

    def test_production_communication_card_no_longer_shows_firmware_version_labels(self) -> None:
        page = ProductionPage(_FakeBridge())

        label_texts = {label.text() for label in page.communication_section.findChildren(QLabel)}
        self.assertNotIn("MCU Firmware Version: -", label_texts)
        self.assertNotIn("Nodes Firmware Version: -", label_texts)
        self.assertNotIn("MCU Firmware Version", " ".join(label_texts))
        self.assertNotIn("Nodes Firmware Version", " ".join(label_texts))
        self.assertIsNone(page.findChild(QLabel, "PageTitle"))
        self.assertIsNone(page.findChild(QLabel, "PageSubtitle"))

    def test_workspace_pages_no_longer_render_dynamic_page_titles_or_subtitles(self) -> None:
        window = self._build_workspace_window()
        window.show()
        self._app.processEvents()

        for route, page in window._pages.items():
            if page is None:
                continue
            self.assertIsNone(page.findChild(QLabel, "PageTitle"), route)
            self.assertIsNone(page.findChild(QLabel, "PageSubtitle"), route)

    def test_robot_node_status_widget_is_persistent_and_reused_across_tabs(self) -> None:
        window = self._build_workspace_window()
        window.show()
        self._app.processEvents()

        shared_section = window.node_status_section
        production_page = window._pages[ROUTE_PRODUCTION]
        self.assertIs(production_page.node_status_section, shared_section)
        self.assertTrue(shared_section.isVisible())

        window.set_active_page(ROUTE_MECHANICAL, log_route_change=False)
        self._app.processEvents()
        self.assertTrue(shared_section.isVisible())
        self.assertIs(window.node_status_section, shared_section)

        window.set_active_page(ROUTE_FIRMWARE, log_route_change=False)
        self._app.processEvents()
        self.assertTrue(shared_section.isVisible())
        self.assertIs(window._pages[ROUTE_PRODUCTION].node_status_section, shared_section)

    def test_shared_node_status_display_state_requires_update_and_survives_tab_switches(self) -> None:
        window = self._build_workspace_window()
        window.show()

        state = {"connected": True, "power_on": True, "connected_nodes": [], "detected_nodes": []}
        scan_requests: list[str] = []

        window._bridge.get_runtime_communication_model = lambda create_if_missing=False: {
            "connected": state["connected"],
            "ports": [],
            "selected_port": "COM11",
            "baud_rates": ["115200"],
            "selected_baud": "115200",
        }
        window._bridge.get_runtime_robot_power_state = lambda create_if_missing=False: state["power_on"]
        window._bridge.get_runtime_robot_nodes = lambda create_if_missing=False: {
            "connected_nodes": list(state["connected_nodes"]),
            "detected_nodes": list(state["detected_nodes"]),
            "rows": [{"node_id": node_id} for node_id in state["connected_nodes"]],
        }
        def _request_scan() -> bool:
            scan_requests.append("scan")
            state["detected_nodes"] = []
            return True
        window._bridge.request_runtime_node_scan = _request_scan

        section = window.node_status_section

        def _is_green(node_id: int) -> bool:
            return "#7ED957" in section._led_by_node_id[node_id].styleSheet()

        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertFalse(_is_green(3))
        self.assertFalse(_is_green(6))

        state["connected_nodes"] = [3, 6]
        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertFalse(_is_green(3))
        self.assertFalse(_is_green(6))

        section.findChild(QPushButton, "UpdateNodesButton").click()
        self._app.processEvents()
        self.assertEqual(scan_requests, ["scan"])
        self.assertFalse(_is_green(3))
        self.assertFalse(_is_green(6))

        state["detected_nodes"] = [3]
        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertTrue(_is_green(3))
        self.assertFalse(_is_green(6))

        window.set_active_page(ROUTE_MECHANICAL, log_route_change=False)
        self._app.processEvents()
        self.assertTrue(_is_green(3))
        self.assertFalse(_is_green(6))

        state["detected_nodes"] = [3, 6]
        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertTrue(_is_green(3))
        self.assertTrue(_is_green(6))

        section.findChild(QPushButton, "ClearNodesButton").click()
        self._app.processEvents()
        self.assertFalse(_is_green(3))
        self.assertFalse(_is_green(6))

        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertFalse(_is_green(3))
        self.assertFalse(_is_green(6))

        section.findChild(QPushButton, "UpdateNodesButton").click()
        self._app.processEvents()
        state["detected_nodes"] = [3]
        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertTrue(_is_green(3))

        state["power_on"] = False
        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertFalse(_is_green(3))
        self.assertFalse(_is_green(6))

        state["power_on"] = True
        state["detected_nodes"] = [3, 6]
        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertFalse(_is_green(3))
        self.assertFalse(_is_green(6))

    def test_shared_node_status_marks_only_detected_nodes_green_during_active_cycle(self) -> None:
        window = self._build_workspace_window()
        window.show()

        state = {
            "connected": True,
            "power_on": True,
            "connected_nodes": [5, 8, 9, 12],
            "detected_nodes": [],
        }

        window._bridge.get_runtime_communication_model = lambda create_if_missing=False: {
            "connected": state["connected"],
            "ports": [],
            "selected_port": "COM11",
            "baud_rates": ["115200"],
            "selected_baud": "115200",
        }
        window._bridge.get_runtime_robot_power_state = lambda create_if_missing=False: state["power_on"]
        window._bridge.get_runtime_robot_nodes = lambda create_if_missing=False: {
            "connected_nodes": list(state["connected_nodes"]),
            "detected_nodes": list(state["detected_nodes"]),
            "rows": [{"node_id": node_id} for node_id in state["connected_nodes"]],
        }
        window._bridge.request_runtime_node_scan = lambda: state.__setitem__("detected_nodes", []) or True

        section = window.node_status_section

        def _is_green(node_id: int) -> bool:
            return "#7ED957" in section._led_by_node_id[node_id].styleSheet()

        window._refresh_shared_node_status()
        self._app.processEvents()
        for node_id in (5, 8, 9, 12):
            self.assertFalse(_is_green(node_id))

        section.findChild(QPushButton, "UpdateNodesButton").click()
        self._app.processEvents()

        for reply_node in (5, 8, 9, 12):
            state["detected_nodes"] = [*state["detected_nodes"], reply_node]
            window._refresh_shared_node_status()
            self._app.processEvents()
            for node_id in (5, 8, 9, 12):
                self.assertEqual(_is_green(node_id), node_id in state["detected_nodes"])

    def test_emergency_stop_badge_is_global_and_does_not_change_node_leds(self) -> None:
        window = self._build_workspace_window()
        window.show()

        state = {
            "connected": True,
            "power_on": True,
            "connected_nodes": [5, 8],
            "detected_nodes": [5],
            "emergency": None,
        }

        window._bridge.get_runtime_communication_model = lambda create_if_missing=False: {
            "connected": state["connected"],
            "ports": [],
            "selected_port": "COM11",
            "baud_rates": ["115200"],
            "selected_baud": "115200",
        }
        window._bridge.get_runtime_robot_power_state = lambda create_if_missing=False: state["power_on"]
        window._bridge.get_runtime_emergency_stop_state = lambda create_if_missing=False: state["emergency"]
        window._bridge.get_runtime_robot_nodes = lambda create_if_missing=False: {
            "connected_nodes": list(state["connected_nodes"]),
            "detected_nodes": list(state["detected_nodes"]),
            "rows": [{"node_id": node_id} for node_id in state["connected_nodes"]],
        }

        section = window.node_status_section
        badge = section.findChild(QLabel, "EmergencyStopBadge")
        robot_power = section.findChild(QPushButton, "RobotPowerButton")

        def _is_green(node_id: int) -> bool:
            return "#7ED957" in section._led_by_node_id[node_id].styleSheet()

        section.begin_visual_update(window._bridge.get_runtime_robot_nodes())
        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertEqual(badge.text(), "ESTOP")
        self.assertIn("#2E9F58", badge.styleSheet())
        self.assertEqual(badge.focusPolicy(), Qt.FocusPolicy.NoFocus)
        self.assertTrue(_is_green(5))
        self.assertFalse(_is_green(8))
        self.assertLess(badge.x(), robot_power.x())

        state["emergency"] = True
        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertEqual(badge.text(), "EMERGENCY BUTTON")
        self.assertIn("#D92D20", badge.styleSheet())
        self.assertTrue(_is_green(5))
        self.assertFalse(_is_green(8))

        window.set_active_page(ROUTE_MECHANICAL, log_route_change=False)
        self._app.processEvents()
        self.assertEqual(badge.text(), "EMERGENCY BUTTON")

        state["emergency"] = False
        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertEqual(badge.text(), "Emergency Stop")
        self.assertIn("#2E9F58", badge.styleSheet())
        self.assertFalse(badge.hasFocus())

    def test_update_nodes_clears_fresh_scan_before_replies_arrive(self) -> None:
        window = self._build_workspace_window()
        window.show()

        state = {
            "connected": True,
            "power_on": True,
            "connected_nodes": [5, 8, 9, 12],
            "detected_nodes": [5, 8, 9, 12],
        }

        window._bridge.get_runtime_communication_model = lambda create_if_missing=False: {
            "connected": state["connected"],
            "ports": [],
            "selected_port": "COM11",
            "baud_rates": ["115200"],
            "selected_baud": "115200",
        }
        window._bridge.get_runtime_robot_power_state = lambda create_if_missing=False: state["power_on"]
        window._bridge.get_runtime_robot_nodes = lambda create_if_missing=False: {
            "connected_nodes": list(state["connected_nodes"]),
            "detected_nodes": list(state["detected_nodes"]),
            "rows": [{"node_id": node_id} for node_id in state["connected_nodes"]],
        }
        window._bridge.request_runtime_node_scan = lambda: state.__setitem__("detected_nodes", []) or True

        section = window.node_status_section

        def _is_green(node_id: int) -> bool:
            return "#7ED957" in section._led_by_node_id[node_id].styleSheet()

        section.begin_visual_update(window._bridge.get_runtime_robot_nodes())
        window._refresh_shared_node_status()
        self._app.processEvents()
        self.assertTrue(_is_green(5))

        section.findChild(QPushButton, "UpdateNodesButton").click()
        self._app.processEvents()
        self.assertFalse(_is_green(5))
        self.assertFalse(_is_green(8))
        self.assertFalse(_is_green(9))
        self.assertFalse(_is_green(12))


class _FakeScanTimer:
    def __init__(self) -> None:
        self.started_with: list[int] = []
        self.stopped = False

    def start(self, interval: int) -> None:
        self.started_with.append(interval)

    def stop(self) -> None:
        self.stopped = True


class _FakeScanBackend:
    def __init__(self) -> None:
        self.sent_nodes: list[int] = []

    def send_node_id_request(self, node_id: int):
        self.sent_nodes.append(node_id)
        return bytearray([0x25, 0xA5, 0x01, node_id, 0x31, 0x02, 0x86, 0x3F])


class MainWindowNodeScanTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls._app = QApplication.instance() or QApplication([])

    def test_dispatch_node_scan_batch_sends_all_node_queries_immediately(self) -> None:
        logs: list[str] = []
        updates: list[str] = []
        window = type("ScanHarness", (), {})()
        window.backend_client = _FakeScanBackend()
        window.node_scan_timeout_timer = _FakeScanTimer()
        window.scan_active = False
        window.cancel_scanning = False
        window.current_scan_node = 99
        window.detected_nodes = {5, 8}
        window.validate_connection_state = lambda: True
        window.log = logs.append
        window.update_node_status_display = lambda: updates.append("refresh")

        result = MainWindow.dispatch_node_scan_batch(window)

        self.assertTrue(result)
        self.assertTrue(window.scan_active)
        self.assertTrue(window._batch_node_scan_active)
        self.assertEqual(window.current_scan_node, 2)
        self.assertEqual(window.detected_nodes, set())
        self.assertEqual(window.backend_client.sent_nodes, list(range(2, 18)))
        self.assertEqual(window.node_scan_timeout_timer.started_with, [500])
        self.assertEqual(updates, ["refresh"])

    def test_workspace_session_edit_button_routes_to_active_production_page(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            config_path = Path(tmpdir) / "demo.yaml"
            config_path.write_text(
                """
project:
  name: demo
  display_name: Demo
features:
  firmware_tools: true
ui:
  workspace: phase2_shell
""".strip(),
                encoding="utf-8",
            )
            project = ProjectDefinition(
                name="demo",
                display_name="Demo",
                config_path=config_path,
                features=ProjectFeatures(firmware_tools=True),
                ui=ProjectUiConfig(workspace="phase2_shell"),
            )

            production_state = SessionState(
                project_name="Demo",
                connection_text="Workbook loaded",
                session_text="Metadata ready",
                active_page="Production",
                operator_name="Alice",
                assembler_name="Bob",
                metadata_edit_enabled=True,
            )
            edit_calls: list[ProductionPage] = []

            def _build_session_state(self: ProductionPage) -> SessionState:
                return production_state

            def _handle_session_metadata_edit_requested(self: ProductionPage) -> bool:
                edit_calls.append(self)
                return True

            with patch.object(ProductionPage, "build_session_state", _build_session_state), patch.object(
                ProductionPage,
                "handle_session_metadata_edit_requested",
                _handle_session_metadata_edit_requested,
            ):
                window = ProjectWorkspaceWindow(project)
                window.set_active_page(ROUTE_PRODUCTION, log_route_change=False)
                self.assertTrue(window.live_session_panel.edit_button.isEnabled())
                window.live_session_panel.edit_button.click()
                self._app.processEvents()

                self.assertEqual(len(edit_calls), 1)
                self.assertIsInstance(edit_calls[0], ProductionPage)

                window.set_active_page(ROUTE_PROJECT_CONFIG, log_route_change=False)
                self._app.processEvents()
                self.assertFalse(window.live_session_panel.edit_button.isEnabled())

                window.set_active_page(ROUTE_PRODUCTION, log_route_change=False)
                self._app.processEvents()
                self.assertTrue(window.live_session_panel.edit_button.isEnabled())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for workbook metadata tests.")
    def test_production_page_load_prompts_and_saves_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            config_path = Path(tmpdir) / "demo.yaml"
            config_path.write_text(
                """
project:
  name: demo
  display_name: Demo
features:
  firmware_tools: true
ui:
  workspace: phase2_shell
""".strip(),
                encoding="utf-8",
            )
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            workbook = Workbook()
            summary = workbook.active
            summary.title = "3X"
            summary["A1"] = "Programming"
            summary["A3"] = "Operator"
            summary["A4"] = "UUID"
            summary["A5"] = "PWM"
            summary["B4"] = "1223303010"
            summary["B5"] = "100"
            workbook.create_sheet("3X_D")
            workbook.create_sheet("3X_A")
            workbook.save(workbook_path)

            page = ProductionPage(_FakeBridge())

            with patch("gui.workspace.pages.production_page.QFileDialog.getOpenFileName", return_value=(str(workbook_path), "Excel Files (*.xlsx)")), patch(
                "gui.workspace.pages.production_page.ProductionMetadataDialog.exec",
                return_value=QDialog.DialogCode.Accepted,
            ), patch(
                "gui.workspace.pages.production_page.ProductionMetadataDialog.metadata_values",
                return_value=("Operator A", "Assembler A"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            metadata = page._ipqc_excel_adapter.read_production_metadata()
            self.assertEqual(metadata.operator_name, "Operator A")
            self.assertEqual(metadata.assembler_name, "Assembler A")


if __name__ == "__main__":
    unittest.main()
