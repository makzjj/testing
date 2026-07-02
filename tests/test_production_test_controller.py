"""Focused tests for the Production runtime-backed ML 2.0 node flow."""

from __future__ import annotations

import os
import re
import tempfile
import time
import unittest
import types
from pathlib import Path
from unittest.mock import patch

import pytest
from PyQt6.QtCore import QObject, pyqtSignal
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QApplication, QLabel, QMessageBox, QPushButton

from gui.workspace.pages.production_page import ProductionPage, _SamplingSession
from gui.workspace.dialogs.sampling_test_popup import SamplingTestPopup
from gui.workspace.pages.single_axis_functional_popup import SingleAxisFunctionalPopup
from gui.workspace.controllers.single_axis_functional_test_controller import (
    FunctionalTestConfig,
    SingleAxisFunctionalTestController,
)
from gui.workspace.controllers.sampling_test_controller import SamplingResumeContext, SamplingTestController
from gui.workspace.widgets import ResponsiveRow
from gui.workspace.pages.production_parameter_controller import (
    EEPROM_SAVE_COMMAND,
    SET_COMMAND_SUFFIX,
    ParameterDefinition,
    ParameterVerificationResult,
    ProductionParameterController,
    build_eeprom_save_payload,
    build_pwm_read_payload,
    build_pwm_write_payload,
    build_run,
    build_uuid_read_payload,
    build_uuid_write_payload,
    decode_eeprom_save_response,
    decode_pwm_response,
    decode_uuid_response,
    default_workbook_parameter_definitions,
    format_uuid_like_source,
    parse_pwm_value,
    parse_uuid_value,
    validate_uuid_format,
)
from services.communication_log_store import CommunicationLogStore
from data.binary_cmd_parser import decode_nodeconfig_motion_polarity
from services.node_sensor_profile import NodeSensorProfile
from services.node_sensor_profile import NodeSensorProfile
from myconfig.constants import COMMANDS

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

try:
    from openpyxl import Workbook, load_workbook

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - environment dependent.
    _HAS_OPENPYXL = False

class _FakeBackendClient:
    def __init__(self, *, connected: bool = True) -> None:
        self._connected = connected
        self.sent_commands: list[tuple[int, list[int]]] = []
        self.stop_commands: list[int] = []

    def is_connected(self) -> bool:
        return self._connected

    def get_command_bytes(self, _command_name: str, fallback: list[int] | None = None) -> list[int]:
        return list(COMMANDS.get(_command_name, fallback or []))

    def send_command_bytes(self, node_id: int, command_bytes: list[int]) -> bytearray:
        self.sent_commands.append((node_id, list(command_bytes)))
        return bytearray([0x25, 0xA5, 0x01, node_id, 0x31, len(command_bytes), *command_bytes])

    def send_stop_motor(self, node_id: int) -> bytearray:
        self.stop_commands.append(node_id)
        return bytearray([0x25, 0xA5, 0x01, node_id, 0x31, 0x01, 0xDD])


class _FakeRuntimeWindow(QObject):
    packet_received = pyqtSignal(object)

    def __init__(self, *, connected: bool = True, mcu_version: str | None = "v1.0.0") -> None:
        super().__init__()
        self.backend_client = _FakeBackendClient(connected=connected)
        self.communication_log_store = CommunicationLogStore(max_entries=32)
        self.mcu_version = mcu_version
        self._selected_port = "COM11"
        self._selected_baud = "115200"
        self.sys_mode = {"text": "Ready" if connected else "System Off", "node_id": 0x01, "state_value": 1 if connected else 0}
        self.node_status = {
            2: {"connected": False},
            3: {"connected": False},
            8: {"connected": False},
        }
        self.scan_requests = 0


class _FakeBridge:
    def __init__(self, runtime_window: _FakeRuntimeWindow | None) -> None:
        self.runtime_window = runtime_window
        self.create_requests = 0

    def get_runtime_window(self, *, create_if_missing: bool = False):
        if create_if_missing:
            self.create_requests += 1
        return self.runtime_window

    def get_runtime_communication_log_store(self, *, create_if_missing: bool = False):
        runtime_window = self.get_runtime_window(create_if_missing=create_if_missing)
        if runtime_window is None:
            return None
        return runtime_window.communication_log_store

    def get_runtime_connection_state(self, *, create_if_missing: bool = False) -> tuple[bool, bool]:
        runtime_window = self.get_runtime_window(create_if_missing=create_if_missing)
        if runtime_window is None:
            return False, False
        serial_connected = runtime_window.backend_client.is_connected()
        return serial_connected, serial_connected

    def get_runtime_communication_model(self, *, create_if_missing: bool = False) -> dict:
        runtime_window = self.get_runtime_window(create_if_missing=create_if_missing)
        if runtime_window is None:
            return {
                "ports": [{"label": "COM11 ✅ (Valid)", "value": "COM11"}],
                "selected_port": "COM11",
                "baud_rates": ["115200", "230400", "345600"],
                "selected_baud": "115200",
                "connected": False,
            }
        return {
            "ports": [{"label": "COM11 ✅ (Valid)", "value": "COM11"}],
            "selected_port": runtime_window._selected_port,
            "baud_rates": ["115200", "230400", "345600"],
            "selected_baud": runtime_window._selected_baud,
            "connected": runtime_window.backend_client.is_connected(),
        }

    def connect_runtime_serial(self, *, port: str, baud_rate: int) -> bool:
        runtime_window = self.get_runtime_window(create_if_missing=True)
        if runtime_window is None:
            return False
        runtime_window._selected_port = port
        runtime_window._selected_baud = str(baud_rate)
        runtime_window.backend_client._connected = True
        return True

    def disconnect_runtime_serial(self) -> None:
        runtime_window = self.get_runtime_window(create_if_missing=True)
        if runtime_window is None:
            return
        runtime_window.backend_client._connected = False

    def get_runtime_robot_power_state(self, *, create_if_missing: bool = False) -> bool | None:
        runtime_window = self.get_runtime_window(create_if_missing=create_if_missing)
        if runtime_window is None:
            return None
        sys_mode = getattr(runtime_window, "sys_mode", None)
        if not isinstance(sys_mode, dict):
            return None
        if sys_mode.get("node_id") == 0x01 and sys_mode.get("state_value") == 0:
            return False
        if str(sys_mode.get("text", "")).strip().lower() == "system off":
            return False
        return True

    def send_runtime_robot_power(self, power_on: bool) -> bytearray:
        runtime_window = self.get_runtime_window(create_if_missing=True)
        if runtime_window is None:
            raise RuntimeError("Runtime backend is unavailable for Production operations.")
        backend_client = runtime_window.backend_client
        payload = backend_client.get_command_bytes("ROBOT On" if power_on else "ROBOT Off")
        backend_client.send_command_bytes(0x01, payload)
        runtime_window.sys_mode = {"text": "Ready" if power_on else "System Off", "node_id": 0x01, "state_value": 1 if power_on else 0}
        return bytearray(payload)

    def get_runtime_robot_nodes(self, *, create_if_missing: bool = False) -> dict:
        runtime_window = self.get_runtime_window(create_if_missing=create_if_missing)
        if runtime_window is None:
            return {"connected_nodes": [], "rows": []}
        rows = []
        connected_nodes = []
        for node_id, status in runtime_window.node_status.items():
            if not status.get("connected", False):
                continue
            connected_nodes.append(node_id)
            rows.append(
                {
                    "node_id": node_id,
                    "node": f"Node {node_id:02d} ✅ Connected",
                    "firmware": str(status.get("firmware", "")),
                    "uuid": str(status.get("uuid", "")),
                    "node_type": str(status.get("type", "")),
                    "status": str(status.get("interrupt", "")),
                }
            )
        return {"connected_nodes": sorted(connected_nodes), "rows": rows}

    def request_runtime_node_scan(self) -> bool:
        runtime_window = self.get_runtime_window(create_if_missing=True)
        if runtime_window is None:
            return False
        runtime_window.scan_requests += 1
        return True


class _FakeRobotPowerButton:
    def __init__(self, text: str) -> None:
        self._text = text

    def text(self) -> str:
        return self._text


class _FakeRobotPowerMessageBox:
    class Icon:
        Question = object()

    class ButtonRole:
        AcceptRole = object()
        DestructiveRole = object()
        RejectRole = object()

    next_choice_text: str | None = None
    instances: list["_FakeRobotPowerMessageBox"] = []

    def __init__(self, parent=None) -> None:
        self.parent = parent
        self.icon = None
        self.window_title = ""
        self.text = ""
        self.buttons: list[_FakeRobotPowerButton] = []
        self._clicked_button: _FakeRobotPowerButton | None = None
        _FakeRobotPowerMessageBox.instances.append(self)

    def setIcon(self, icon) -> None:
        self.icon = icon

    def setWindowTitle(self, title: str) -> None:
        self.window_title = title

    def setText(self, text: str) -> None:
        self.text = text

    def addButton(self, text: str, _role) -> _FakeRobotPowerButton:
        button = _FakeRobotPowerButton(text)
        self.buttons.append(button)
        return button

    def exec(self) -> int:
        choice = type(self).next_choice_text
        self._clicked_button = None
        if choice is not None:
            for button in self.buttons:
                if button.text() == choice:
                    self._clicked_button = button
                    break
        return 0

    def clickedButton(self):
        return self._clicked_button


class ProductionPageWorkflowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls._app = QApplication.instance() or QApplication([])

    def test_parse_pwm_value_validation(self) -> None:
        self.assertEqual(ProductionPage._parse_pwm_value("100"), 100)
        self.assertEqual(parse_pwm_value("-100"), -100)
        self.assertEqual(parse_pwm_value("32767"), 32767)
        with self.assertRaisesRegex(ValueError, "required"):
            ProductionPage._parse_pwm_value(" ")
        with self.assertRaisesRegex(ValueError, "digits only"):
            ProductionPage._parse_pwm_value("10x")
        with self.assertRaisesRegex(ValueError, "16-bit"):
            ProductionPage._parse_pwm_value("70000")

    @staticmethod
    def _create_ipqc_workbook(path: Path, *, with_optional_fields: bool = True) -> None:
        if not _HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required to create IPQC workbook fixtures.")
        wb = Workbook()
        ws = wb.active
        ws.title = "3X"
        sampling_3x = wb.create_sheet("3X_D")
        wb.create_sheet("3X_A")
        ws_4y = wb.create_sheet("4Y")
        sampling_4y = wb.create_sheet("4Y_D")
        wb.create_sheet("4Y_A")
        ws["A1"] = "Programming"
        ws["B2"] = "Source"
        ws["C2"] = "Programmed"
        ws["D2"] = "Check"
        ws["A3"] = "Operator"
        ws["A4"] = "Assembler"
        ws["A5"] = "UUID"
        ws["A6"] = "PWM"
        ws["A7"] = "Proportionate (P)"
        ws["A8"] = "Integral (I)"
        ws["A9"] = "Derivative (D)"
        ws["A10"] = "PID_SlewRate"
        ws["A11"] = "RampDown_Slope"
        ws["A12"] = "RampDown_Step"
        ws["A13"] = "RampDown_MinVel"
        ws["A14"] = "RampDown_TargetOffset"
        ws["A15"] = "RampDown_Region"
        ws["A16"] = "Acceptable_Error"
        ws["B3"] = "operator-a"
        ws["B4"] = "assembler-a"
        ws["B5"] = "1223303010"
        ws["B6"] = "100"
        if with_optional_fields:
            ws["B7"] = "0.125"
            ws["B8"] = "0.025"
            ws["B9"] = "0.010"
            ws["B10"] = "1500"
            ws["B11"] = "-25"
            ws["B12"] = "4"
            ws["B13"] = "8"
            ws["B14"] = "-12"
            ws["B15"] = "75"
            ws["B16"] = "30"
            ws["C3"] = "N/A"
            ws["D3"] = "N/A"
        ProductionPageWorkflowTests._populate_sampling_sheet(sampling_3x)
        ws_4y["A1"] = "Programming"
        ws_4y["B2"] = "Source"
        ws_4y["C2"] = "Programmed"
        ws_4y["D2"] = "Check"
        ws_4y["A3"] = "Operator"
        ws_4y["A4"] = "Assembler"
        ws_4y["A5"] = "UUID"
        ws_4y["A6"] = "PWM"
        ws_4y["A7"] = "Proportionate (P)"
        ws_4y["A8"] = "Integral (I)"
        ws_4y["A9"] = "Derivative (D)"
        ws_4y["A10"] = "PID_SlewRate"
        ws_4y["A11"] = "RampDown_Slope"
        ws_4y["A12"] = "RampDown_Step"
        ws_4y["A13"] = "RampDown_MinVel"
        ws_4y["A14"] = "RampDown_TargetOffset"
        ws_4y["A15"] = "RampDown_Region"
        ws_4y["A16"] = "Acceptable_Error"
        ws_4y["B3"] = "operator-a"
        ws_4y["B4"] = "assembler-a"
        ws_4y["B5"] = "1223303010"
        ws_4y["B6"] = "100"
        if with_optional_fields:
            ws_4y["B7"] = "0.125"
            ws_4y["B8"] = "0.025"
            ws_4y["B9"] = "0.010"
            ws_4y["B10"] = "1500"
            ws_4y["B11"] = "-25"
            ws_4y["B12"] = "4"
            ws_4y["B13"] = "8"
            ws_4y["B14"] = "-12"
            ws_4y["B15"] = "75"
            ws_4y["B16"] = "30"
            ws_4y["C3"] = "N/A"
            ws_4y["D3"] = "N/A"
        ProductionPageWorkflowTests._populate_sampling_sheet(sampling_4y)
        wb.save(path)

    @staticmethod
    def _populate_sampling_sheet(sheet) -> None:
        pwm_values = [100, 90, 80, 70, 60]
        section_starts = {"Range": 1, "Speed": 18, "Time": 35}
        for section_name, start_row in section_starts.items():
            sheet[f"A{start_row}"] = section_name
            row = start_row + 1
            for pwm in pwm_values:
                sheet[f"A{row}"] = f"PWM {pwm}"
                sheet[f"A{row + 1}"] = f"+{pwm}"
                sheet[f"A{row + 2}"] = f"-{pwm}"
                row += 3

    @staticmethod
    def _build_parameter_verify_packet(definition, actual_value: int | str) -> dict:
        if definition.name == "UUID":
            payload = [0x3A, *build_uuid_write_payload(int(actual_value))[2:]]
            return {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0xE0, "params": payload}
        if definition.name == "PWM":
            payload = [0x00, *list(int(actual_value).to_bytes(2, "big", signed=True))]
            return {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x85, "params": payload}

        payload = [0x3A]
        if definition.sub_id is not None:
            payload.append(definition.sub_id)
        payload.extend(list(int(actual_value).to_bytes(definition.value_size, "big", signed=definition.signed)))
        return {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": definition.command_id, "params": payload}

    @staticmethod
    def _build_parameter_write_ack_packet(definition, actual_value: int | str) -> dict:
        if definition.name == "PWM":
            payload = [0x53, *list(int(actual_value).to_bytes(2, "big", signed=True))]
            return {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x84, "params": payload}
        return ProductionPageWorkflowTests._build_parameter_verify_packet(definition, actual_value)

    @staticmethod
    def _select_node(page: ProductionPage, node_text: str = "Node 6 - H") -> None:
        combo = page.test_control_section._combo
        for index in range(combo.count()):
            if combo.itemText(index) == node_text:
                combo.setCurrentIndex(index)
                if page._single_axis_passed:
                    match = re.search(r"Node\s+(\d+)", node_text)
                    if match is not None:
                        ProductionPageWorkflowTests._seed_sampling_context(
                            page,
                            node_id=int(match.group(1)),
                            nodeconfig=0x00,
                        )
                break

    @staticmethod
    def _seed_sampling_context(page: ProductionPage, *, node_id: int, nodeconfig: int) -> None:
        polarity = decode_nodeconfig_motion_polarity(nodeconfig)
        profile = NodeSensorProfile.from_node_context(node_id, polarity)
        page._sampling_motion_polarity = polarity
        page._sampling_sensor_profile = profile
        page._sampling_controller.set_motion_polarity(polarity)
        page._sampling_controller.set_sensor_profile(profile)
        page._refresh_sampling_action_states()

    @staticmethod
    def _populate_updated_programming_values(sheet) -> None:
        sheet["B5"] = "1243203029"
        sheet["B6"] = "0"
        sheet["B7"] = "2000"
        sheet["B8"] = "1"
        sheet["B9"] = "35000"
        sheet["B10"] = "0"
        sheet["B11"] = "6"
        sheet["B12"] = "3"
        sheet["B13"] = "90"
        sheet["B14"] = "512"
        sheet["B15"] = "5"
        sheet["B16"] = "256"

    def test_production_page_updates_ui_for_runtime_backed_selected_node_pass(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        combo = page.test_control_section._combo
        for index in range(combo.count()):
            if combo.itemText(index) == "Node 8 - RZ":
                combo.setCurrentIndex(index)
                break
        page._handle_run_test()
        self._app.processEvents()

        self.assertEqual(page.result_summary_section._status_label.text(), "TESTING")
        self.assertIn("[INFO] TESTING: Running Production test for Node 8 RZ.", page.progress_section.to_plain_text())

        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 8, "cmd": 0xCB, "params": [0xA5, 0x5A]}
        )
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 8, "cmd": 0xC8, "params": [0x3A, 1, 2, 3]}
        )
        runtime_window.packet_received.emit(
            {
                "status": "ok",
                "type": "can_over_uart",
                "sender": 8,
                "cmd": 0x82,
                "params": [0x00, 0x00, 0x01, 0xC8],
            }
        )
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 8, "cmd": 0xD8, "params": [1, 0]}
        )
        self._app.processEvents()

        self.assertEqual(page.result_summary_section._status_label.text(), "PASS")
        self.assertIn("All profile steps passed", page.result_summary_section._reason_label.text())

    def test_selected_node_changes_invalidate_verification_cache_and_stage_results(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
            self._app.processEvents()
            with patch(
                "gui.workspace.pages.single_axis_functional_popup.SingleAxisFunctionalPopup.ask_start_sampling",
                return_value=False,
            ):
                page._handle_single_axis_test_requested()
                assert page._single_axis_popup is not None
                page._single_axis_popup.node_combo.setCurrentIndex(1)
                page._single_axis_popup.mark_passed()
            self._app.processEvents()
            self._select_node(page, "Node 6 - H")

            definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}

            def seed_verified_context(node_id: int, node_name: str, base_group: str) -> None:
                sheet_name = page._ipqc_excel_adapter.resolve_sampling_sheet_name(base_group)
                self._seed_sampling_context(page, node_id=node_id, nodeconfig=0x00)
                page._last_parameter_verification_results_by_name = {
                    "UUID": ParameterVerificationResult(definitions["UUID"], "1223306010", "1223306010", True, ""),
                    "PWM": ParameterVerificationResult(definitions["PWM"], "100", "100", True, ""),
                }
                page._parameter_verification_context_key = page._current_production_context_key()
                page._workbook_verification_passed = True
                page._single_axis_passed = True
                page.stage_section.set_stage_status("single_axis", "pass")
                page.stage_section.set_stage_status("sampling", "pass")
                page._sampling_controller._resume_context = SamplingResumeContext(
                    node_id=node_id,
                    node_name=node_name,
                    base_group=base_group,
                    sheet_name=sheet_name,
                    pwm_values=(100,),
                    samples_per_direction=1,
                    current_pwm_index=0,
                    current_pwm=100,
                    current_sample_index=1,
                    current_direction="+",
                    completed_measurements=1,
                    total_measurements=2,
                    terminal_state=SamplingTestController.S_SAMPLE_WAIT_SENSOR,
                    reason="cached",
                    resumable=True,
                    sample_incomplete=True,
                    home_sensor="L",
                    middle_target=None,
                )

            current_group = page._ipqc_excel_adapter.active_sheet_group or "3X"
            seed_verified_context(6, "H", current_group)
            requests = page._build_workbook_parameter_requests(6, "H")
            filtered_requests, skipped_labels = page._filter_verify_requests_by_previous_verification(requests or [])
            self.assertTrue(skipped_labels)
            self.assertLess(len(filtered_requests), len(requests or []))

            self._select_node(page, "Node 8 - RZ")
            self._app.processEvents()
            page._handle_test_control_node_selected()
            self.assertFalse(page._single_axis_passed)
            self.assertFalse(page.stage_section.stage_enabled("sampling"))
            self.assertTrue(page.stage_section.stage_enabled("single_axis"))
            self.assertFalse(page._workbook_verification_passed)
            self.assertEqual(page._last_parameter_verification_results_by_name, {})
            self.assertIsNone(page._sampling_controller.resume_context)
            self.assertFalse(page._sampling_controller.can_resume)
            new_requests = page._build_workbook_parameter_requests(8, "RZ") or []
            filtered_new_requests, skipped_new_labels = page._filter_verify_requests_by_previous_verification(new_requests)
            self.assertEqual(skipped_new_labels, [])
            self.assertEqual(len(filtered_new_requests), len(new_requests))

            seed_verified_context(8, "RZ", current_group)
            available_groups = page._ipqc_excel_adapter.available_base_sheet_groups
            alternate_group = next((group for group in available_groups if group != current_group), None)
            self.assertIsNotNone(alternate_group)
            assert alternate_group is not None
            page._handle_ipqc_sheet_group_changed(alternate_group)
            self._app.processEvents()
            self.assertEqual(page._last_parameter_verification_results_by_name, {})
            self.assertIsNone(page._sampling_controller.resume_context)
            self.assertFalse(page._sampling_controller.can_resume)

            current_group = page._ipqc_excel_adapter.active_sheet_group or alternate_group
            seed_verified_context(8, "RZ", current_group)
            workbook_path_2 = Path(tmpdir) / "ipqc_reload.xlsx"
            self._create_ipqc_workbook(workbook_path_2)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path_2), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
            self._app.processEvents()
            page._handle_test_control_node_selected()
            self.assertFalse(page._single_axis_passed)
            self.assertFalse(page.stage_section.stage_enabled("sampling"))
            self.assertEqual(page._last_parameter_verification_results_by_name, {})
            self.assertIsNone(page._sampling_controller.resume_context)
            self.assertFalse(page._sampling_controller.can_resume)
        self.assertIn("#b0b7c3", page.stage_section._rows["single_axis"][0].styleSheet().lower())

    def test_profile_step_results_keep_workbook_flow_only(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        page._handle_run_test()
        runtime_window.packet_received.emit({"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0xCB, "params": [0xA5, 0x5A]})
        runtime_window.packet_received.emit({"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0xC8, "params": [0x3A, 1, 2, 3]})
        runtime_window.packet_received.emit({"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0x82, "params": [0x00, 0x00, 0x00, 0x2A]})
        runtime_window.packet_received.emit({"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0xD8, "params": [0, 1]})
        self._app.processEvents()

        self.assertFalse(hasattr(page, "_result_logger"))

    def test_production_page_hides_test_profile_selector(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        self.assertFalse(hasattr(page.test_control_section, "_profile_combo"))

    def test_production_page_run_test_still_works_without_profile_selector(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        page._handle_run_test()
        self._app.processEvents()
        self.assertEqual(page.result_summary_section._status_label.text(), "TESTING")

    def test_production_page_writes_persistent_rows_before_pwm_and_triggers_one_eeprom_save(self) -> None:
        if not _HAS_OPENPYXL:
            self.skipTest("openpyxl is required for workbook flow tests.")

        runtime_window = _FakeRuntimeWindow(connected=True)
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(template_path)
            workbook = load_workbook(template_path)
            summary = workbook["3X"]
            for row in range(7, 17):
                summary[f"B{row}"] = None
            workbook.save(template_path)

            page._ipqc_excel_adapter.load_template(template_path)
            page._workbook_loaded = True
            page.uuid_section.set_workbook_path(str(template_path))
            page.uuid_section.set_sheet_groups(["3X"], "3X")
            page._refresh_workbook_action_states()

            combo = page.test_control_section._combo
            for index in range(combo.count()):
                if combo.itemText(index) == "Node 6 - H":
                    combo.setCurrentIndex(index)
                    break

            page._handle_write_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[0][1][0], 0xE0)
            self.assertTrue(page._workbook_parameter_write_pending)

            uuid_write_payload = runtime_window.backend_client.sent_commands[0][1]
            runtime_window.packet_received.emit(
                {
                    "status": "ok",
                    "type": "can_over_uart",
                    "sender": 6,
                    "cmd": 0xE0,
                    "params": [0x3A, *uuid_write_payload[2:]],
                }
            )
            self._app.processEvents()

            self.assertEqual(runtime_window.backend_client.sent_commands[1], (6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]))
            self.assertTrue(page._workbook_eeprom_save_pending)

            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": EEPROM_SAVE_COMMAND, "params": [0x0A, 0x00]}
            )
            self._app.processEvents()

            self.assertEqual(runtime_window.backend_client.sent_commands[2], (6, [0x84, 0x00, 0x64]))
            self.assertTrue(page._workbook_runtime_write_pending)

            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x84, "params": [0x53, 0x00, 0x64]}
            )
            self._app.processEvents()

            self.assertFalse(page._workbook_runtime_write_pending)
            self.assertIn("Runtime parameters written", page.uuid_section.last_workbook_action_text)

    def test_production_page_full_parameter_write_save_runtime_and_verify_sequence(self) -> None:
        if not _HAS_OPENPYXL:
            self.skipTest("openpyxl is required for workbook flow tests.")

        runtime_window = _FakeRuntimeWindow(connected=True)
        runtime_window.node_status[6] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "H",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(template_path)
            workbook = load_workbook(template_path)
            summary = workbook["3X"]
            summary["B3"] = "operator-a"
            summary["B4"] = "assembler-a"
            summary["B5"] = "123456789"
            summary["B6"] = "80"
            summary["B7"] = "2000"
            summary["B8"] = "1"
            summary["B9"] = "35000"
            summary["B10"] = "0"
            summary["B11"] = "40"
            summary["B12"] = "3"
            summary["B13"] = "50"
            summary["B14"] = "512"
            summary["B15"] = "10"
            summary["B16"] = "256"
            workbook.save(template_path)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(template_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            combo = page.test_control_section._combo
            for index in range(combo.count()):
                if combo.itemText(index) == "Node 6 - H":
                    combo.setCurrentIndex(index)
                    break

            page._handle_write_uuid()
            self._app.processEvents()

            expected_write_sequence: list[tuple[int, list[int]]] = [
                (6, build_uuid_write_payload(123456789)),
                (6, definitions["PID_P"].build_write_command(2000)),
                (6, definitions["PID_I"].build_write_command(1)),
                (6, definitions["PID_D"].build_write_command(35000)),
                (6, definitions["PID_SlewRate"].build_write_command(0)),
                (6, definitions["RampDown_Slope"].build_write_command(40)),
                (6, definitions["RampDown_Step"].build_write_command(3)),
                (6, definitions["RampDown_MinVel"].build_write_command(50)),
                (6, definitions["RampDown_TargetOffset"].build_write_command(512)),
                (6, definitions["RampDown_Region"].build_write_command(10)),
                (6, definitions["Acceptable_Error"].build_write_command(256)),
            ]

            self.assertEqual(runtime_window.backend_client.sent_commands[0], expected_write_sequence[0])
            self.assertEqual(len(runtime_window.backend_client.sent_commands), 1)
            self.assertTrue(page._workbook_parameter_write_pending)

            write_response_payloads = {
                "UUID": [0x3A, *build_uuid_write_payload(123456789)[2:]],
                "PID_P": [0x3A, 0x70, *list((2000).to_bytes(4, "big", signed=True))],
                "PID_I": [0x3A, 0x69, *list((1).to_bytes(4, "big", signed=True))],
                "PID_D": [0x3A, 0x64, *list((35000).to_bytes(4, "big", signed=True))],
                "PID_SlewRate": [0x3A, *list((0).to_bytes(2, "big", signed=False))],
                "RampDown_Slope": [0x3A, *list((40).to_bytes(2, "big", signed=True))],
                "RampDown_Step": [0x3A, 0x03],
                "RampDown_MinVel": [0x3A, 0x32],
                "RampDown_TargetOffset": [0x3A, *list((512).to_bytes(2, "big", signed=True))],
                "RampDown_Region": [0x3A, 0x0A],
                "Acceptable_Error": [0x3A, *list((256).to_bytes(2, "big", signed=False))],
            }

            for index, definition_name in enumerate(
                [
                    "UUID",
                    "PID_P",
                    "PID_I",
                    "PID_D",
                    "PID_SlewRate",
                    "RampDown_Slope",
                    "RampDown_Step",
                    "RampDown_MinVel",
                    "RampDown_TargetOffset",
                    "RampDown_Region",
                    "Acceptable_Error",
                ],
                start=0,
            ):
                definition = definitions[definition_name]
                payload = runtime_window.backend_client.sent_commands[index][1]
                self.assertEqual(payload[0], definition.command_id)
                if definition.sub_id is not None:
                    self.assertEqual(payload[2], definition.sub_id)
                self.assertEqual(len(runtime_window.backend_client.sent_commands), index + 1)
                runtime_window.packet_received.emit(
                    {
                        "status": "ok",
                        "type": "can_over_uart",
                        "sender": 6,
                        "cmd": definition.command_id,
                        "params": write_response_payloads[definition_name],
                    }
                )
                self._app.processEvents()
                if definition_name != "Acceptable_Error":
                    self.assertEqual(len(runtime_window.backend_client.sent_commands), index + 2)

            self.assertEqual(runtime_window.backend_client.sent_commands[len(expected_write_sequence)], (6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]))
            self.assertEqual(runtime_window.backend_client.sent_commands.count((6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX])), 1)
            self.assertTrue(page._workbook_eeprom_save_pending)

            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": EEPROM_SAVE_COMMAND, "params": [0x0A, 0x00]}
            )
            self._app.processEvents()

            self.assertEqual(runtime_window.backend_client.sent_commands[len(expected_write_sequence) + 1], (6, [0x84, 0x00, 0x50]))
            self.assertTrue(page._workbook_runtime_write_pending)
            self.assertTrue(page._workbook_eeprom_settle_active)

            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x84, "params": [0x53, 0x00, 0x50]}
            )
            self._app.processEvents()

            self.assertFalse(page._workbook_runtime_write_pending)
            self.assertFalse(page._workbook_eeprom_save_pending)
            self.assertTrue(page._workbook_eeprom_settle_active)

            page._handle_eeprom_save_settle_finished()
            self._app.processEvents()
            self.assertFalse(page._workbook_eeprom_settle_active)
            self.assertTrue(page.uuid_section.verify_button.isEnabled())

            page._handle_verify_uuid()
            self._app.processEvents()

            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (6, [0xE0, 0x3F]))
            self.assertEqual(runtime_window.backend_client.sent_commands.count((6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX])), 1)
            save_command_index = runtime_window.backend_client.sent_commands.index((6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]))
            pwm_write_index = runtime_window.backend_client.sent_commands.index((6, build_pwm_write_payload(80)))
            self.assertGreater(pwm_write_index, save_command_index)

            verify_order = [
                "UUID",
                "PWM",
                "PID_P",
                "PID_I",
                "PID_D",
                "PID_SlewRate",
                "RampDown_Slope",
                "RampDown_Step",
                "RampDown_MinVel",
                "RampDown_TargetOffset",
                "RampDown_Region",
                "Acceptable_Error",
            ]
            verify_response_payloads = {
                "UUID": [0x3A, *build_uuid_write_payload(123456789)[2:]],
                "PWM": [0x00, 0x50],
                "PID_P": [0x3A, 0x70, *list((2000).to_bytes(4, "big", signed=True))],
                "PID_I": [0x3A, 0x69, *list((1).to_bytes(4, "big", signed=True))],
                "PID_D": [0x3A, 0x64, *list((35000).to_bytes(4, "big", signed=True))],
                "PID_SlewRate": [0x3A, *list((0).to_bytes(2, "big", signed=False))],
                "RampDown_Slope": [0x3A, *list((40).to_bytes(2, "big", signed=True))],
                "RampDown_Step": [0x3A, 0x03],
                "RampDown_MinVel": [0x3A, 0x32],
                "RampDown_TargetOffset": [0x3A, *list((512).to_bytes(2, "big", signed=True))],
                "RampDown_Region": [0x3A, 0x0A],
                "Acceptable_Error": [0x3A, *list((256).to_bytes(2, "big", signed=False))],
            }

            expected_verify_commands = [
                (6, [0xE0, 0x3F]),
                (6, [0x85]),
                (6, [0xE7, 0x3F, 0x70]),
                (6, [0xE7, 0x3F, 0x69]),
                (6, [0xE7, 0x3F, 0x64]),
                (6, [0xED, 0x3F]),
                (6, [0x89, 0x3F]),
                (6, [0x8B, 0x3F]),
                (6, [0x8C, 0x3F]),
                (6, [0xE1, 0x3F]),
                (6, [0xE2, 0x3F]),
                (6, [0xEC, 0x3F]),
            ]

            self.assertTrue(page.uuid_section.verify_button.isEnabled())
            page._handle_verify_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], expected_verify_commands[0])

            for index, definition_name in enumerate(verify_order):
                definition = definitions[definition_name]
                expected_command = expected_verify_commands[index]
                self.assertEqual(runtime_window.backend_client.sent_commands[-1], expected_command)
                payload = verify_response_payloads[definition_name]
                runtime_window.packet_received.emit(
                    {
                        "status": "ok",
                        "type": "can_over_uart",
                        "sender": 6,
                        "cmd": definition.command_id if definition_name != "PWM" else 0x85,
                        "params": payload,
                }
                )
                self._app.processEvents()
                if index < len(verify_order) - 1:
                    self.assertEqual(runtime_window.backend_client.sent_commands[-1], expected_verify_commands[index + 1])

            output_sheet = page._ipqc_excel_adapter._workbook["3X"]
            expected_rows = {
                5: ("123456789", "PASS"),
                6: ("80", "PASS"),
                7: ("2000", "PASS"),
                8: ("1", "PASS"),
                9: ("35000", "PASS"),
                10: ("0", "PASS"),
                11: ("40", "PASS"),
                12: ("3", "PASS"),
                13: ("50", "PASS"),
                14: ("512", "PASS"),
                15: ("10", "PASS"),
                16: ("256", "PASS"),
            }
            for row, (actual_text, check_text) in expected_rows.items():
                with self.subTest(row=row):
                    self.assertEqual(output_sheet[f"C{row}"].value, actual_text)
                    self.assertEqual(output_sheet[f"D{row}"].value, check_text)

            self.assertEqual(page.uuid_section.workbook_validation_text, "Workbook Validation: PASSED")
            self.assertTrue(page.uuid_section.save_button.isEnabled())
            self.assertEqual(runtime_window.backend_client.sent_commands.count((6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX])), 1)
            self.assertFalse(page._workbook_parameter_write_pending)
            self.assertFalse(page._workbook_runtime_write_pending)
            self.assertFalse(page._workbook_eeprom_save_pending)
            self.assertFalse(page._workbook_eeprom_settle_active)

    def test_production_page_uses_two_column_top_layout_and_section_order(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        first_row = page.content_layout.itemAt(0).widget()
        second_widget = page.content_layout.itemAt(1).widget()
        third_widget = page.content_layout.itemAt(2).widget()
        fourth_widget = page.content_layout.itemAt(3).widget()

        self.assertIsInstance(first_row, ResponsiveRow)
        self.assertIs(second_widget, page.node_status_section)
        self.assertIs(third_widget, page.uuid_section)
        self.assertIs(fourth_widget, page.progress_section)

        first_layout = first_row.layout()
        self.assertIs(first_layout.itemAt(0).widget(), page.communication_section)
        self.assertIs(first_layout.itemAt(1).widget(), page.stage_section)
        self.assertEqual(first_row.stretch_factors(), (1, 1))
        self.assertIsNone(page.result_summary_section.parent())

    def test_production_page_shows_compact_status_with_refresh_and_clear_buttons(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        page.resize(1280, 800)
        self._app.processEvents()

        button_texts = [button.text() for button in page.progress_section.findChildren(QPushButton)]
        self.assertIn("Refresh", button_texts)
        self.assertIn("Clear", button_texts)
        self.assertEqual(page.progress_section.windowTitle(), "")
        self.assertTrue(hasattr(page.progress_section, "_log_output"))
        self.assertGreaterEqual(page.progress_section._log_output.minimumHeight(), 220)
        self.assertEqual(
            page.progress_section._log_output.horizontalScrollBarPolicy(),
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff,
        )
        self.assertEqual(page.horizontalScrollBarPolicy(), Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

    def test_production_page_node_status_clear_resets_led_to_dark_green(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        runtime_window.node_status[8]["connected"] = True
        page.node_status_section.set_connected(True)
        page.node_status_section.set_robot_power_state(True)
        page.node_status_section.begin_visual_update({"connected_nodes": []})
        page._refresh_robot_nodes()
        self.assertIn("#7ed957", page.node_status_section._led_by_node_id[8].styleSheet().lower())
        page._handle_clear_nodes_requested()
        self.assertIn("#1e5e20", page.node_status_section._led_by_node_id[8].styleSheet().lower())
        node_labels = [label.text() for label in page.node_status_section.findChildren(QLabel)]
        self.assertEqual(node_labels.count("Node"), 1)
        self.assertEqual(len(page.node_status_section._led_by_node_id), 15)

    def test_production_page_shows_communication_card_controls(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        self.assertGreaterEqual(page.communication_section._port_combo.count(), 1)
        self.assertGreaterEqual(page.communication_section._baud_combo.count(), 1)
        self.assertEqual(page.communication_section._connect_button.text(), "Disconnect")
        self.assertFalse(hasattr(page.communication_section, "_firmware_label"))
        self.assertFalse(hasattr(page.communication_section, "_nodes_firmware_label"))

    def test_production_page_populates_com_ports_on_startup_without_runtime_page(self) -> None:
        bridge = _FakeBridge(None)
        page = ProductionPage(bridge)
        page.show()
        self._app.processEvents()

        self.assertEqual(bridge.create_requests, 0)
        self.assertGreaterEqual(page.communication_section._port_combo.count(), 1)
        self.assertEqual(page.communication_section._port_combo.currentData(), "COM11")

    def test_production_page_robot_power_button_order_and_connection_state(self) -> None:
        disconnected_page = ProductionPage(_FakeBridge(None))
        button_row = disconnected_page.node_status_section.body_layout.itemAt(0).layout()
        self.assertIsNotNone(button_row)
        assert button_row is not None
        self.assertIsNotNone(button_row.itemAt(0).spacerItem())
        button_texts = [button_row.itemAt(index).widget().text() for index in range(button_row.count()) if button_row.itemAt(index).widget() is not None]
        self.assertEqual(button_texts, ["Emergency Button", "Robot Power ON/OFF", "Update Nodes", "Clear"])
        self.assertFalse(disconnected_page.node_status_section._robot_power_button.isEnabled())

        connected_page = ProductionPage(_FakeBridge(_FakeRuntimeWindow(connected=True)))
        self.assertTrue(connected_page.node_status_section._robot_power_button.isEnabled())

    def test_robot_power_button_cancels_without_sending(self) -> None:
        runtime_window = _FakeRuntimeWindow(connected=True)
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        _FakeRobotPowerMessageBox.next_choice_text = "Cancel"
        _FakeRobotPowerMessageBox.instances.clear()
        with patch("gui.workspace.pages.production_page.QMessageBox", _FakeRobotPowerMessageBox):
            page._handle_robot_power_requested()

        self.assertEqual(len(_FakeRobotPowerMessageBox.instances), 1)
        dialog = _FakeRobotPowerMessageBox.instances[0]
        self.assertEqual(dialog.window_title, "Robot Power")
        self.assertEqual(dialog.text, "Choose robot power command to send.")
        self.assertEqual([button.text() for button in dialog.buttons], ["Power ON", "Power OFF", "Cancel"])
        self.assertEqual(runtime_window.backend_client.sent_commands, [])
        self.assertIn("Robot power command cancelled.", page.progress_section.to_plain_text())

    def test_robot_power_button_sends_on_then_off_payloads_via_existing_backend_path(self) -> None:
        runtime_window = _FakeRuntimeWindow(connected=True)
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        _FakeRobotPowerMessageBox.instances.clear()
        with patch("gui.workspace.pages.production_page.QMessageBox", _FakeRobotPowerMessageBox):
            _FakeRobotPowerMessageBox.next_choice_text = "Power ON"
            page._handle_robot_power_requested()
            _FakeRobotPowerMessageBox.next_choice_text = "Power OFF"
            page._handle_robot_power_requested()

        self.assertEqual(
            runtime_window.backend_client.sent_commands,
            [
                (1, COMMANDS["ROBOT On"]),
                (1, COMMANDS["ROBOT Off"]),
            ],
        )
        log_text = page.progress_section.to_plain_text()
        self.assertIn("Robot power ON command sent.", log_text)
        self.assertIn("Robot power OFF command sent.", log_text)

    def test_robot_power_button_succeeds_when_power_state_is_unavailable(self) -> None:
        runtime_window = _FakeRuntimeWindow(connected=True)
        runtime_window.sys_mode = None
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        _FakeRobotPowerMessageBox.next_choice_text = "Power ON"
        _FakeRobotPowerMessageBox.instances.clear()
        with patch("gui.workspace.pages.production_page.QMessageBox", _FakeRobotPowerMessageBox):
            page._handle_robot_power_requested()

        self.assertEqual(runtime_window.backend_client.sent_commands, [(1, COMMANDS["ROBOT On"])])
        self.assertNotIn("robot power state is unavailable", page.progress_section.to_plain_text().lower())

    def test_production_page_shows_runtime_robot_nodes_status_and_supports_dropdown_sync(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[8] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "123456",
            "type": "RZ",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        page.robot_nodes_section.set_connected(True)
        page.robot_nodes_section.set_robot_power_state(True)
        page.robot_nodes_section.begin_visual_update({"connected_nodes": []})
        page._refresh_robot_nodes()
        self.assertIn("#7ed957", page.robot_nodes_section._led_by_node_id[8].styleSheet().lower())
        page._handle_runtime_node_selected(8)
        selected_node_id, _selected_name = page.test_control_section.selected_node()
        self.assertEqual(selected_node_id, 8)

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook UI tests.")
    def test_production_page_loads_ipqc_workbook_and_shows_expected_preview(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=True)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            self.assertEqual(page.uuid_section._selected_group, "3X")
            self.assertIn("Loaded Workbook", page.uuid_section._loaded_workbook_label.text())
            self.assertTrue(page.uuid_section._loaded_workbook_label.text().endswith("ipqc.xlsx"))
            self.assertEqual(page.uuid_section._loaded_workbook_label.toolTip(), str(workbook_path))
            self.assertEqual(page.uuid_section._expected_serial_value, "1223303010")
            self.assertEqual(page.uuid_section._expected_pwm_value, "100")
            self.assertEqual(page.uuid_section._expected_other_value, "-")
            self.assertEqual(page.uuid_section.workbook_validation_text, "Workbook Validation: READY")
            self.assertTrue(page.uuid_section.verify_button.isEnabled())
            self.assertTrue(page.uuid_section.write_button.isEnabled())
            self.assertFalse(page.uuid_section.save_button.isEnabled())
            self.assertEqual(page.uuid_section.last_workbook_action_text, "Workbook loaded; no write performed yet")
            self.assertEqual(runtime_window.backend_client.sent_commands, [])
            log_text = page.progress_section.to_plain_text()
            self.assertIn("Loaded Programming table with 12 supported parameter(s).", log_text)
            self.assertNotIn("Expected S/N / UUID:", log_text)
            self.assertNotIn("Expected PWM:", log_text)
            self.assertIn("loaded ipqc workbook", page.progress_section.to_html().lower())
            self.assertIn("#2e7d32", page.progress_section.to_html().lower())

    def test_production_page_reload_clears_stale_parameter_workflow_state(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=True)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            page._handle_write_uuid()
            self._app.processEvents()
            self.assertTrue(page._workbook_parameter_write_pending or page._workbook_runtime_write_pending)
            self.assertEqual(page._parameter_controller._parameter_operation_mode, "write")

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            self.assertFalse(page._workbook_parameter_write_pending)
            self.assertFalse(page._workbook_runtime_write_pending)
            self.assertFalse(page._workbook_eeprom_save_pending)
            self.assertFalse(page._workbook_eeprom_save_failed)
            self.assertFalse(page._workbook_eeprom_settle_active)
            self.assertIsNone(page._parameter_controller._pending_parameter_request)
            self.assertIsNone(page._parameter_controller._parameter_operation_mode)
            self.assertEqual(page._parameter_controller._parameter_requests, [])
            self.assertEqual(page._parameter_controller._parameter_results, [])
            self.assertEqual(page.uuid_section.workbook_validation_text, "Workbook Validation: READY")
            self.assertEqual(page.uuid_section.last_workbook_action_text, "Workbook loaded; no write performed yet")
            self.assertTrue(page.uuid_section.write_button.isEnabled())
            self.assertTrue(page.uuid_section.verify_button.isEnabled())
            self.assertFalse(page.uuid_section.save_button.isEnabled())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook write wiring tests.")
    def test_production_page_verify_all_matching_reads_without_writing_or_saving(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[6] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "H",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}
        workbook_values = {
            "UUID": 1243203029,
            "PWM": 0,
            "PID_P": 2000,
            "PID_I": 1,
            "PID_D": 35000,
            "PID_SlewRate": 0,
            "RampDown_Slope": 6,
            "RampDown_Step": 3,
            "RampDown_MinVel": 90,
            "RampDown_TargetOffset": 512,
            "RampDown_Region": 5,
            "Acceptable_Error": 256,
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path)
            workbook = load_workbook(workbook_path)
            self._populate_updated_programming_values(workbook["3X"])
            workbook.save(workbook_path)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            self._select_node(page)
            page._handle_verify_uuid()
            self._app.processEvents()

            verify_order = [
                "UUID",
                "PWM",
                "PID_P",
                "PID_I",
                "PID_D",
                "PID_SlewRate",
                "RampDown_Slope",
                "RampDown_Step",
                "RampDown_MinVel",
                "RampDown_TargetOffset",
                "RampDown_Region",
                "Acceptable_Error",
            ]
            expected_read_sequence = [
                (6, definitions[name].build_read_command())
                for name in verify_order
            ]
            for index, name in enumerate(verify_order):
                definition = definitions[name]
                runtime_window.packet_received.emit(
                    self._build_parameter_verify_packet(definition, workbook_values[name])
                )
                self._app.processEvents()
                self.assertEqual(runtime_window.backend_client.sent_commands[index], expected_read_sequence[index])

            self.assertEqual(runtime_window.backend_client.sent_commands, expected_read_sequence)
            self.assertNotIn((6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]), runtime_window.backend_client.sent_commands)
            self.assertEqual(page.result_summary_section._status_label.text(), "PASS")
            self.assertTrue(page.uuid_section.save_button.isEnabled())
            self.assertIn("Workbook parameter read-back verification", page.progress_section.to_plain_text())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook write wiring tests.")
    def test_production_page_verify_mismatches_reads_without_writing_or_saving(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[6] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "H",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}
        workbook_values = {
            "UUID": 1243203029,
            "PWM": 0,
            "PID_P": 2000,
            "PID_I": 1,
            "PID_D": 35000,
            "PID_SlewRate": 0,
            "RampDown_Slope": 6,
            "RampDown_Step": 3,
            "RampDown_MinVel": 90,
            "RampDown_TargetOffset": 512,
            "RampDown_Region": 5,
            "Acceptable_Error": 256,
        }
        actual_values = {
            "UUID": 1243203030,
            "PWM": 0,
            "PID_P": 3000,
            "PID_I": 2,
            "PID_D": 40000,
            "PID_SlewRate": 0,
            "RampDown_Slope": 6,
            "RampDown_Step": 3,
            "RampDown_MinVel": 90,
            "RampDown_TargetOffset": 512,
            "RampDown_Region": 5,
            "Acceptable_Error": 256,
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path)
            workbook = load_workbook(workbook_path)
            self._populate_updated_programming_values(workbook["3X"])
            workbook.save(workbook_path)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            self._select_node(page)
            page._handle_verify_uuid()
            self._app.processEvents()

            verify_order = [
                "UUID",
                "PWM",
                "PID_P",
                "PID_I",
                "PID_D",
                "PID_SlewRate",
                "RampDown_Slope",
                "RampDown_Step",
                "RampDown_MinVel",
                "RampDown_TargetOffset",
                "RampDown_Region",
                "Acceptable_Error",
            ]
            expected_read_sequence = [
                (6, definitions[name].build_read_command())
                for name in verify_order
            ]
            for index, name in enumerate(verify_order):
                definition = definitions[name]
                runtime_window.packet_received.emit(
                    self._build_parameter_verify_packet(definition, actual_values[name])
                )
                self._app.processEvents()
                self.assertEqual(runtime_window.backend_client.sent_commands[index], expected_read_sequence[index])

            self.assertEqual(runtime_window.backend_client.sent_commands, expected_read_sequence)
            self.assertNotIn((6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]), runtime_window.backend_client.sent_commands)
            self.assertEqual(page.result_summary_section._status_label.text(), "FAIL")
            self.assertFalse(page.uuid_section.save_button.isEnabled())
            self.assertIn("Workbook Validation: FAILED", page.uuid_section.workbook_validation_text)

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook write wiring tests.")
    def test_production_page_write_filters_to_failed_previous_verification_results_and_sends_one_eeprom_save(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[6] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "H",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path)
            workbook = load_workbook(workbook_path)
            self._populate_updated_programming_values(workbook["3X"])
            workbook.save(workbook_path)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            self._select_node(page)

            results: list[ParameterVerificationResult] = []
            for name, actual_value, passed in [
                ("UUID", 1243203030, False),
                ("PWM", 0, True),
                ("PID_P", 3000, False),
                ("PID_I", 2, False),
                ("PID_D", 40000, False),
                ("PID_SlewRate", 0, True),
                ("RampDown_Slope", 6, True),
                ("RampDown_Step", 3, True),
                ("RampDown_MinVel", 90, True),
                ("RampDown_TargetOffset", 512, True),
                ("RampDown_Region", 5, True),
                ("Acceptable_Error", 256, True),
            ]:
                definition = definitions[name]
                expected_text = workbook["3X"][definition.expected_cell].value
                actual_text = str(actual_value)
                reason = (
                    f"{definition.label} read-back verification"
                    if passed
                    else f"{definition.label} read-back verification - expected {expected_text}, actual {actual_text}"
                )
                results.append(
                    ParameterVerificationResult(
                        definition=definition,
                        expected_text=str(expected_text),
                        actual_text=actual_text,
                        passed=passed,
                        reason=reason,
                    )
                )

            page._handle_parameter_verification_finished(
                False,
                "Workbook parameter read-back verification",
                results,
            )
            self._app.processEvents()

            page._handle_write_uuid()
            self._app.processEvents()
            self.assertEqual(
                runtime_window.backend_client.sent_commands[0],
                (6, build_uuid_write_payload(1243203029)),
            )
            self.assertEqual(page.progress_section.to_plain_text().count("WRITE PLAN: 4 mismatched parameter(s): S/N, PID_P, PID_I, PID_D"), 1)
            self.assertIn(
                "SKIPPED: 8 parameter(s) already matched read-back values.",
                page.progress_section.to_plain_text(),
            )

            write_response_payloads = {
                "UUID": [0x3A, *build_uuid_write_payload(1243203029)[2:]],
                "PID_P": [0x3A, 0x70, *list((2000).to_bytes(4, "big", signed=True))],
                "PID_I": [0x3A, 0x69, *list((1).to_bytes(4, "big", signed=True))],
                "PID_D": [0x3A, 0x64, *list((35000).to_bytes(4, "big", signed=True))],
            }
            write_order = ["UUID", "PID_P", "PID_I", "PID_D"]
            for index, name in enumerate(write_order):
                definition = definitions[name]
                runtime_window.packet_received.emit(
                    {
                        "status": "ok",
                        "type": "can_over_uart",
                        "sender": 6,
                        "cmd": definition.command_id,
                        "params": write_response_payloads[name],
                    }
                )
                self._app.processEvents()

            self.assertEqual(
                runtime_window.backend_client.sent_commands[1:4],
                [
                    (6, definitions["PID_P"].build_write_command(2000)),
                    (6, definitions["PID_I"].build_write_command(1)),
                    (6, definitions["PID_D"].build_write_command(35000)),
                ],
            )
            self.assertEqual(
                runtime_window.backend_client.sent_commands[4],
                (6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]),
            )
            self.assertEqual(runtime_window.backend_client.sent_commands.count((6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX])), 1)

            runtime_window.packet_received.emit(
                {
                    "status": "ok",
                    "type": "can_over_uart",
                    "sender": 6,
                    "cmd": EEPROM_SAVE_COMMAND,
                    "params": [0x0A, 0x00],
                }
            )
            self._app.processEvents()
            page._handle_eeprom_save_settle_finished()
            self._app.processEvents()

            self.assertFalse(page._workbook_runtime_write_pending)
            self.assertFalse(page._workbook_parameter_write_pending)
            self.assertFalse(page._workbook_eeprom_save_pending)
            self.assertFalse(page._workbook_eeprom_settle_active)
            self.assertFalse(page.uuid_section.save_button.isEnabled())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook write wiring tests.")
    def test_production_page_write_skips_matching_pwm_and_does_not_send_eeprom_save(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[6] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "H",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path)
            workbook = load_workbook(workbook_path)
            self._populate_updated_programming_values(workbook["3X"])
            workbook.save(workbook_path)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            self._select_node(page)

            results: list[ParameterVerificationResult] = []
            for name, actual_value, passed in [
                ("UUID", 1243203029, True),
                ("PWM", 10, False),
                ("PID_P", 2000, True),
                ("PID_I", 1, True),
                ("PID_D", 35000, True),
                ("PID_SlewRate", 0, True),
                ("RampDown_Slope", 6, True),
                ("RampDown_Step", 3, True),
                ("RampDown_MinVel", 90, True),
                ("RampDown_TargetOffset", 512, True),
                ("RampDown_Region", 5, True),
                ("Acceptable_Error", 256, True),
            ]:
                definition = definitions[name]
                expected_text = workbook["3X"][definition.expected_cell].value
                actual_text = str(actual_value)
                reason = (
                    f"{definition.label} read-back verification"
                    if passed
                    else f"{definition.label} read-back verification - expected {expected_text}, actual {actual_text}"
                )
                results.append(
                    ParameterVerificationResult(
                        definition=definition,
                        expected_text=str(expected_text),
                        actual_text=actual_text,
                        passed=passed,
                        reason=reason,
                    )
                )

            page._handle_parameter_verification_finished(
                False,
                "Workbook parameter read-back verification",
                results,
            )
            self._app.processEvents()

            page._handle_write_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[0], (6, build_pwm_write_payload(0)))
            self.assertNotIn((6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]), runtime_window.backend_client.sent_commands)
            self.assertEqual(len(runtime_window.backend_client.sent_commands), 1)

            runtime_window.packet_received.emit(
                {
                    "status": "ok",
                    "type": "can_over_uart",
                    "sender": 6,
                    "cmd": 0x84,
                    "params": [0x53, 0x00, 0x00],
                }
            )
            self._app.processEvents()
            self.assertNotIn((6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]), runtime_window.backend_client.sent_commands)
            self.assertFalse(page._workbook_eeprom_save_pending)
            self.assertFalse(page._workbook_runtime_write_pending)

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook write wiring tests.")
    def test_production_page_write_with_all_previous_results_matching_sends_no_write_and_no_eeprom_save(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[6] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "H",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path)
            workbook = load_workbook(workbook_path)
            self._populate_updated_programming_values(workbook["3X"])
            workbook.save(workbook_path)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            self._select_node(page)

            results: list[ParameterVerificationResult] = []
            for name in [
                "UUID",
                "PWM",
                "PID_P",
                "PID_I",
                "PID_D",
                "PID_SlewRate",
                "RampDown_Slope",
                "RampDown_Step",
                "RampDown_MinVel",
                "RampDown_TargetOffset",
                "RampDown_Region",
                "Acceptable_Error",
            ]:
                definition = definitions[name]
                expected_text = workbook["3X"][definition.expected_cell].value
                actual_text = str(expected_text)
                results.append(
                    ParameterVerificationResult(
                        definition=definition,
                        expected_text=str(expected_text),
                        actual_text=actual_text,
                        passed=True,
                        reason=f"{definition.label} read-back verification",
                    )
                )

            page._handle_parameter_verification_finished(
                True,
                "Workbook parameter read-back verification",
                results,
            )
            self._app.processEvents()

            page._handle_write_uuid()
            self._app.processEvents()

            self.assertEqual(runtime_window.backend_client.sent_commands, [])
            self.assertEqual(page.result_summary_section._status_label.text(), "READY")
            self.assertIn(
                "No parameter writes required; all workbook values already match MCU read-back.",
                page.result_summary_section._reason_label.text(),
            )
            self.assertNotIn((6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]), runtime_window.backend_client.sent_commands)

    def test_production_page_logs_workbook_load_failure_in_red(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with patch("gui.workspace.pages.production_page.QFileDialog.getOpenFileName", return_value=("bad.xlsx", "Excel Files (*.xlsx)")):
            with patch.object(page._ipqc_excel_adapter, "load_template", side_effect=RuntimeError("broken workbook")):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

        self.assertIn("IPQC workbook load failed", page.progress_section.to_plain_text())
        self.assertIn("#c62828", page.progress_section.to_html().lower())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook write wiring tests.")
    def test_production_page_write_uuid_sends_write_command_using_workbook_expected_value(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            page._handle_write_uuid()
            self._app.processEvents()

            expected_uuid = 1223303010
            self.assertTrue(runtime_window.backend_client.sent_commands)
            self.assertIn((3, build_uuid_write_payload(expected_uuid)), runtime_window.backend_client.sent_commands)
            self.assertEqual(len(runtime_window.backend_client.sent_commands), 1)
            self.assertTrue(page._workbook_parameter_write_pending)

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook write wiring tests.")
    def test_production_page_write_uuid_timeout_disables_verify_and_reports_quiet_mode_issue(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        page._parameter_controller._timeout_ms = 20

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            page._handle_write_uuid()
            self._app.processEvents()
            runtime_window.packet_received.emit(
                {
                    "status": "ok",
                    "type": "can_over_uart",
                    "sender": 3,
                    "cmd": 0xE0,
                    "params": [0x3A, *build_uuid_write_payload(1223303010)[2:]],
                }
            )
            self._app.processEvents()

            deadline = time.monotonic() + 1.0
            while time.monotonic() < deadline and page.uuid_section.workbook_validation_text != "Workbook Validation: FAILED (EEPROM save ACK not received; check command payload and quiet mode.)":
                self._app.processEvents()
                time.sleep(0.01)

            self.assertEqual(
                page.uuid_section.workbook_validation_text,
                "Workbook Validation: FAILED (EEPROM save ACK not received; check command payload and quiet mode.)",
            )
            self.assertFalse(page.uuid_section.verify_button.isEnabled())
            self.assertIn("EEPROM save ACK not received; check command payload and quiet mode.", page.result_summary_section._reason_label.text())
            timeout_line = next(
                line for line in page.progress_section.to_plain_text().splitlines() if "Timed out waiting for EEPROM save ACK." in line
            )
            self.assertIn("[FAIL] Timed out waiting for EEPROM save ACK.", timeout_line)
            self.assertNotIn("Node", timeout_line)
            self.assertNotIn("Axis", timeout_line)

    def test_progress_log_coloring_uses_black_green_and_red(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        page.progress_section.clear_log()
        page.progress_section.append_step("Info entry")
        page.progress_section.append_step("Pass entry", level="pass")
        page.progress_section.append_step("Fail entry", level="fail")

        html = page.progress_section.to_html().lower()
        self.assertIn("#000000", html)
        self.assertIn("#2e7d32", html)
        self.assertIn("#c62828", html)

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook UUID verify tests.")
    def test_production_page_verify_after_load_without_prior_write_sets_passed_and_enables_save(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            self.assertTrue(page.uuid_section.verify_button.isEnabled())
            self.assertFalse(page.uuid_section.save_button.isEnabled())

            expected_uuid = 1223303010
            page._handle_verify_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0xE0, 0x3F]))
            self.assertNotIn((3, build_uuid_write_payload(expected_uuid)), runtime_window.backend_client.sent_commands)

            runtime_window.packet_received.emit(
                {
                    "status": "ok",
                    "type": "can_over_uart",
                    "sender": 3,
                    "cmd": 0xE0,
                    "params": [0x3A, *build_uuid_write_payload(expected_uuid)[2:]],
                }
            )
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0x85]))
            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0x85, "params": [0x00, 0x64]}
            )
            self._app.processEvents()

            output_sheet = page._ipqc_excel_adapter._workbook["3X"]
            self.assertEqual(output_sheet["C5"].value, str(expected_uuid))
            self.assertEqual(output_sheet["D5"].value, "PASS")
            self.assertEqual(output_sheet["C6"].value, "100")
            self.assertEqual(output_sheet["D6"].value, "PASS")
            self.assertEqual(page.uuid_section.workbook_validation_text, "Workbook Validation: PASSED")
            self.assertTrue(page.uuid_section.save_button.isEnabled())
            self.assertFalse(hasattr(page, "_result_logger"))

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook UUID verify tests.")
    def test_production_page_verify_after_load_without_prior_write_sets_failed_on_mismatch(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            page._handle_verify_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0xE0, 0x3F]))
            self.assertNotIn((3, build_uuid_write_payload(1223303011)), runtime_window.backend_client.sent_commands)

            runtime_window.packet_received.emit(
                {
                    "status": "ok",
                    "type": "can_over_uart",
                    "sender": 3,
                    "cmd": 0xE0,
                    "params": [0x3A, *build_uuid_write_payload(1223303011)[2:]],
                }
            )
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0x85]))
            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0x85, "params": [0x00, 0x32]}
            )
            self._app.processEvents()

            output_sheet = page._ipqc_excel_adapter._workbook["3X"]
            self.assertEqual(output_sheet["C5"].value, "1223303011")
            self.assertEqual(output_sheet["D5"].value, "FAIL")
            self.assertEqual(output_sheet["C6"].value, "50")
            self.assertEqual(output_sheet["D6"].value, "FAIL")
            self.assertIn("Workbook Validation: FAILED", page.uuid_section.workbook_validation_text)
            self.assertFalse(page.uuid_section.save_button.isEnabled())
            self.assertIn("expected 1223303010, actual 1223303011", page.progress_section.to_plain_text())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook write wiring tests.")
    def test_production_page_write_uuid_logs_pwm_blocked_when_b6_invalid(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)
            wb = load_workbook(workbook_path)
            wb["3X"]["B6"] = "bad-pwm"
            wb.save(workbook_path)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            page._handle_write_uuid()
            self._app.processEvents()

            self.assertEqual(runtime_window.backend_client.sent_commands, [])
        self.assertIn("Expected PWM in workbook B6 is invalid", page.result_summary_section._reason_label.text())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook write wiring tests.")
    def test_production_page_workbook_write_failure_is_reporting_error(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            with patch.object(page._ipqc_excel_adapter, "write_programming_parameter_result", side_effect=OSError("disk full")):
                page._update_uuid_cells_in_workbook_memory("1223303011", True)
                self._app.processEvents()

            self.assertEqual(page.result_summary_section._status_label.text(), "REPORTING ERROR")
            self.assertIn("writing IPQC workbook report failed", page.result_summary_section._reason_label.text())
            self.assertIn("failed", page.uuid_section.last_workbook_action_text.lower())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook UUID verify tests.")
    def test_production_page_verify_uses_workbook_expected_sn_and_writes_result_cells(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            expected_uuid = 1223303010
            page._handle_verify_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0xE0, 0x3F]))

            response_params = [0x3A, *build_uuid_write_payload(expected_uuid)[2:]]
            runtime_window.packet_received.emit(
                {
                    "status": "ok",
                    "type": "can_over_uart",
                    "sender": 3,
                    "cmd": 0xE0,
                    "params": response_params,
                }
            )
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0x85]))
            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0x85, "params": [0x00, 0x64]}
            )
            self._app.processEvents()

            output_sheet = page._ipqc_excel_adapter._workbook["3X"]
            self.assertEqual(output_sheet["C5"].value, str(expected_uuid))
            self.assertEqual(output_sheet["D5"].value, "PASS")
            self.assertEqual(output_sheet["C6"].value, "100")
            self.assertEqual(output_sheet["D6"].value, "PASS")
            self.assertEqual(page.uuid_section.workbook_validation_text, "Workbook Validation: PASSED")
            self.assertTrue(page.uuid_section.save_button.isEnabled())
            self.assertEqual(page.progress_section.to_plain_text().count("[PASS] Workbook parameter read-back verification"), 1)

            self.assertFalse(hasattr(page, "_result_logger"))

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook UUID verify tests.")
    def test_production_page_verify_mismatch_writes_fail_result_cells(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            page._handle_verify_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0xE0, 0x3F]))
            runtime_window.packet_received.emit(
                {
                    "status": "ok",
                    "type": "can_over_uart",
                    "sender": 3,
                    "cmd": 0xE0,
                    "params": [0x3A, *build_uuid_write_payload(1223303011)[2:]],
                }
            )
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0x85]))
            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0x85, "params": [0x00, 0x32]}
            )
            self._app.processEvents()

            output_sheet = page._ipqc_excel_adapter._workbook["3X"]
            self.assertEqual(output_sheet["C5"].value, "1223303011")
            self.assertEqual(output_sheet["D5"].value, "FAIL")
            self.assertEqual(output_sheet["C6"].value, "50")
            self.assertEqual(output_sheet["D6"].value, "FAIL")
            self.assertIn("Workbook Validation: FAILED", page.uuid_section.workbook_validation_text)
            self.assertFalse(page.uuid_section.save_button.isEnabled())
            self.assertIn("[FAIL] UUID read-back verification", page.progress_section.to_plain_text())
            self.assertIn("expected 1223303010, actual 1223303011", page.progress_section.to_plain_text())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook UUID verify tests.")
    def test_production_page_verify_pwm_mismatch_writes_fail_for_pwm_cells(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            page._handle_verify_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0xE0, 0x3F]))
            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0xE0, "params": [0x3A, *build_uuid_write_payload(1223303010)[2:]]}
            )
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0x85]))
            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0x85, "params": [0x00, 0x32]}
            )
            self._app.processEvents()

            output_sheet = page._ipqc_excel_adapter._workbook["3X"]
            self.assertEqual(output_sheet["C5"].value, "1223303010")
            self.assertEqual(output_sheet["D5"].value, "PASS")
            self.assertEqual(output_sheet["C6"].value, "50")
            self.assertEqual(output_sheet["D6"].value, "FAIL")

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook UUID verify tests.")
    def test_production_page_verify_pwm_timeout_writes_fail_after_read_back_only(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            output_sheet = page._ipqc_excel_adapter._workbook["3X"]
            self.assertIn(output_sheet["C6"].value, (None, ""))
            self.assertIn(output_sheet["D6"].value, (None, ""))
            page._handle_verify_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [0xE0, 0x3F]))
            runtime_window.packet_received.emit(
                {"status": "ok", "type": "can_over_uart", "sender": 3, "cmd": 0xE0, "params": [0x3A, *build_uuid_write_payload(1223303010)[2:]]}
            )
            self._app.processEvents()
            self.assertIn(output_sheet["C6"].value, (None, ""))
            self.assertIn(output_sheet["D6"].value, (None, ""))
            page._parameter_controller._handle_parameter_verify_timeout()
            self._app.processEvents()

            self.assertEqual(output_sheet["C5"].value, "1223303010")
            self.assertEqual(output_sheet["D5"].value, "PASS")
            self.assertIn(output_sheet["C6"].value, (None, ""))
            self.assertIn(output_sheet["D6"].value, (None, ""))
            self.assertIn("actual timeout", page.progress_section.to_plain_text())
            self.assertIn("Workbook Validation: FAILED", page.uuid_section.workbook_validation_text)

    def test_production_page_eeprom_save_failure_marks_validation_failed(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            page._handle_write_uuid()
            self._app.processEvents()
            runtime_window.packet_received.emit(
                {
                    "status": "ok",
                    "type": "can_over_uart",
                    "sender": 3,
                    "cmd": 0xE0,
                    "params": [0x3A, *build_uuid_write_payload(1223303010)[2:]],
                }
            )
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands[-1], (3, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]))
            page._parameter_controller._handle_eeprom_save_timeout()
            self._app.processEvents()

            self.assertEqual(page.result_summary_section._status_label.text(), "FAIL")
            self.assertIn("EEPROM save", page.result_summary_section._reason_label.text())
            self.assertIn("Workbook Validation: FAILED", page.uuid_section.workbook_validation_text)

    def test_production_page_logs_mixed_parameter_results_with_pass_and_fail_levels(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}
        results = [
            ParameterVerificationResult(
                definition=definitions["UUID"],
                expected_text="1223303010",
                actual_text="",
                passed=False,
                reason="S/N read-back verification - expected 1223303010, actual timeout",
            ),
            ParameterVerificationResult(
                definition=definitions["PWM"],
                expected_text="10",
                actual_text="10",
                passed=True,
                reason="PWM read-back verification",
            ),
        ]

        page._handle_parameter_verification_finished(
            False,
            "S/N read-back verification - expected 1223303010, actual timeout",
            results,
        )

        log_text = page.progress_section.to_plain_text()
        self.assertIn("[FAIL] UUID read-back verification - expected 1223303010, actual timeout", log_text)
        self.assertIn("[PASS] PWM read-back verification - expected 10, actual 10", log_text)

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook save tests.")
    def test_production_page_save_completed_workbook_shows_output_path(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        runtime_window.node_status[3] = {
            "connected": True,
            "firmware": "v1.0.0",
            "uuid": "",
            "type": "X",
            "interrupt": "OK",
        }
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            output_path = Path(tmpdir) / "ipqc_completed.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)

            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            page._update_uuid_cells_in_workbook_memory("1223303010", True)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getSaveFileName",
                return_value=(str(output_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_save_completed_workbook()
                self._app.processEvents()

            self.assertEqual(page.uuid_section.workbook_output_text, str(output_path.resolve()))
            self.assertTrue(output_path.exists())

    def test_production_status_history_and_clear_button(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        for index in range(40):
            page.progress_section.append_step(f"log line {index}")
        self._app.processEvents()

        self.assertIn("[INFO] log line 39", page.progress_section.to_plain_text())

        clear_buttons = [button for button in page.progress_section.findChildren(QPushButton) if button.text() == "Clear"]
        self.assertEqual(len(clear_buttons), 1)
        clear_buttons[0].click()
        self._app.processEvents()
        self.assertEqual(page.progress_section.to_plain_text().strip(), "")


class ProductionParameterControllerTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls._app = QApplication.instance() or QApplication([])

    @staticmethod
    def _create_ipqc_workbook(path: Path, *, with_optional_fields: bool = True) -> None:
        ProductionPageWorkflowTests._create_ipqc_workbook(path, with_optional_fields=with_optional_fields)

    def test_uuid_helper_functions(self) -> None:
        self.assertEqual(parse_uuid_value("1234567890"), 1234567890)
        self.assertEqual(parse_uuid_value("0x499602D2"), 1234567890)
        self.assertEqual(parse_pwm_value("100"), 100)
        self.assertEqual(parse_pwm_value("-12"), -12)
        self.assertEqual(build_pwm_read_payload(), [0x85])
        self.assertEqual(build_uuid_read_payload(), [0xE0, 0x3F])
        self.assertEqual(build_uuid_write_payload(1234567890), [0xE0, 0x3D, 0x00, 0x49, 0x96, 0x02, 0xD2])
        self.assertEqual(build_pwm_write_payload(100), [0x84, 0x00, 0x64])
        self.assertEqual(build_pwm_write_payload(-2), [0x84, 0xFF, 0xFE])
        self.assertEqual(format_uuid_like_source(1234567890, "1234567890"), "1234567890")
        self.assertEqual(format_uuid_like_source(1234567890, "0x499602D2"), "0x499602D2")
        self.assertEqual(validate_uuid_format(1223306010, 6), (True, ""))
        is_valid, invalid_message = validate_uuid_format(1223305010, 6)
        self.assertFalse(is_valid)
        self.assertIn("does not match node_id", invalid_message)
        decoded_ok, decoded_uuid, _ = decode_uuid_response([0xE0, 0x3A, 0x00, 0x49, 0x96, 0x02, 0xD2])
        self.assertTrue(decoded_ok)
        self.assertEqual(decoded_uuid, 1234567890)
        pwm_decoded_ok, pwm_decoded_value, _ = decode_pwm_response([0x85, 0x00, 0x50])
        self.assertTrue(pwm_decoded_ok)
        self.assertEqual(pwm_decoded_value, 80)
        pwm_write_ok, pwm_write_value, _ = decode_pwm_response([0x84, 0x53, 0xFF, 0xFE])
        self.assertTrue(pwm_write_ok)
        self.assertEqual(pwm_write_value, -2)

    def test_new_workbook_parameter_definitions_cover_layout_and_payloads(self) -> None:
        controller = ProductionParameterController(_FakeBridge(_FakeRuntimeWindow()))
        defs = {definition.name: definition for definition in default_workbook_parameter_definitions()}
        expected_cells = {
            "UUID": ("B5", "C5", "D5"),
            "PWM": ("B6", "C6", "D6"),
            "PID_P": ("B7", "C7", "D7"),
            "PID_I": ("B8", "C8", "D8"),
            "PID_D": ("B9", "C9", "D9"),
            "PID_SlewRate": ("B10", "C10", "D10"),
            "RampDown_Slope": ("B11", "C11", "D11"),
            "RampDown_Step": ("B12", "C12", "D12"),
            "RampDown_MinVel": ("B13", "C13", "D13"),
            "RampDown_TargetOffset": ("B14", "C14", "D14"),
            "RampDown_Region": ("B15", "C15", "D15"),
            "Acceptable_Error": ("B16", "C16", "D16"),
        }
        for name, cells in expected_cells.items():
            with self.subTest(name=name):
                definition = defs[name]
                self.assertEqual((definition.expected_cell, definition.actual_cell, definition.result_cell), cells)

        pid_p_request = controller.build_parameter_request(defs["PID_P"], 6, "H", "1.25")
        self.assertEqual(pid_p_request.expected_value, 1_250_000)
        self.assertEqual(
            defs["PID_P"].build_write_command(pid_p_request.expected_value),
            [0xE7, 0x3D, 0x70, *list((1_250_000).to_bytes(4, "big", signed=True))],
        )
        self.assertEqual(defs["PID_P"].build_read_command(), [0xE7, 0x3F, 0x70])
        self.assertEqual(
            defs["PID_P"].decode_response([0xE7, 0x3A, 0x70, *list((1_250_000).to_bytes(4, "big", signed=True))]),
            (True, 1_250_000, ""),
        )

        pid_i_request = controller.build_parameter_request(defs["PID_I"], 6, "H", "-0.25")
        self.assertEqual(pid_i_request.expected_value, -250_000)
        self.assertEqual(
            defs["PID_I"].build_write_command(pid_i_request.expected_value),
            [0xE7, 0x3D, 0x69, *list((-250_000).to_bytes(4, "big", signed=True))],
        )
        self.assertEqual(defs["PID_I"].build_read_command(), [0xE7, 0x3F, 0x69])

        pid_d_request = controller.build_parameter_request(defs["PID_D"], 6, "H", "0.75")
        self.assertEqual(pid_d_request.expected_value, 750_000)
        self.assertEqual(
            defs["PID_D"].build_write_command(pid_d_request.expected_value),
            [0xE7, 0x3D, 0x64, *list((750_000).to_bytes(4, "big", signed=True))],
        )
        self.assertEqual(defs["PID_D"].build_read_command(), [0xE7, 0x3F, 0x64])

        slope_request = controller.build_parameter_request(defs["RampDown_Slope"], 6, "H", "-25")
        target_request = controller.build_parameter_request(defs["RampDown_TargetOffset"], 6, "H", "-12")
        self.assertEqual(slope_request.expected_value, -25)
        self.assertEqual(target_request.expected_value, -12)
        self.assertEqual(defs["RampDown_Slope"].build_write_command(-25), [0x89, 0x3D, 0xFF, 0xE7])
        self.assertEqual(defs["RampDown_TargetOffset"].build_write_command(-12), [0xE1, 0x3D, 0xFF, 0xF4])
        self.assertEqual(defs["RampDown_Slope"].decode_response([0x89, 0x3A, 0xFF, 0xE7]), (True, -25, ""))
        self.assertEqual(defs["RampDown_TargetOffset"].decode_response([0xE1, 0x3A, 0xFF, 0xF4]), (True, -12, ""))

        self.assertEqual(defs["PID_SlewRate"].build_write_command(1500), [0xED, 0x3D, 0x05, 0xDC])
        self.assertEqual(defs["PID_SlewRate"].build_read_command(), [0xED, 0x3F])
        self.assertEqual(defs["PID_SlewRate"].decode_response([0xED, 0x3A, 0x05, 0xDC]), (True, 1500, ""))

        self.assertEqual(defs["RampDown_Step"].build_write_command(4), [0x8B, 0x3D, 0x04])
        self.assertEqual(defs["RampDown_MinVel"].build_write_command(8), [0x8C, 0x3D, 0x08])
        self.assertEqual(defs["RampDown_Region"].build_write_command(75), [0xE2, 0x3D, 0x4B])
        self.assertEqual(defs["Acceptable_Error"].build_write_command(30), [0xEC, 0x3D, 0x00, 0x1E])
        self.assertEqual(defs["RampDown_Region"].decode_response([0xE2, 0x3A, 0x4B]), (True, 75, ""))
        self.assertEqual(defs["Acceptable_Error"].decode_response([0xEC, 0x3A, 0x00, 0x1E]), (True, 30, ""))
        self.assertFalse(defs["PWM"].persistent)
        self.assertTrue(defs["UUID"].persistent)

    def test_new_workbook_parameter_validation_rejects_out_of_range_values(self) -> None:
        controller = ProductionParameterController(_FakeBridge(_FakeRuntimeWindow()))
        defs = {definition.name: definition for definition in default_workbook_parameter_definitions()}

        with self.assertRaisesRegex(ValueError, "RampDown_Step"):
            controller.build_parameter_request(defs["RampDown_Step"], 6, "H", "256")
        with self.assertRaisesRegex(ValueError, "RampDown_MinVel"):
            controller.build_parameter_request(defs["RampDown_MinVel"], 6, "H", "-1")
        with self.assertRaisesRegex(ValueError, "RampDown_Region"):
            controller.build_parameter_request(defs["RampDown_Region"], 6, "H", "101")
        with self.assertRaisesRegex(ValueError, "Acceptable_Error"):
            controller.build_parameter_request(defs["Acceptable_Error"], 6, "H", "70000")


@unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for Sampling page integration tests.")
class SamplingPageIntegrationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls._app = QApplication.instance() or QApplication([])

    @staticmethod
    def _create_ipqc_workbook(path: Path, *, with_optional_fields: bool = True) -> None:
        ProductionPageWorkflowTests._create_ipqc_workbook(path, with_optional_fields=with_optional_fields)

    def _load_workbook(self, page: ProductionPage, workbook_path: Path) -> None:
        with patch(
            "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
            return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
        ):
            page._handle_load_ipqc_workbook()
        self._app.processEvents()

    def _enable_single_axis_pass(self, page: ProductionPage, *, start_sampling_prompt: bool = False) -> None:
        with patch(
            "gui.workspace.pages.single_axis_functional_popup.SingleAxisFunctionalPopup.ask_start_sampling",
            return_value=start_sampling_prompt,
        ):
            page._handle_single_axis_test_requested()
            assert page._single_axis_popup is not None
            page._single_axis_popup.node_combo.setCurrentIndex(1)
            page._single_axis_popup.mark_passed()
        self._app.processEvents()
        try:
            node_id, _node_name = page.test_control_section.selected_node()
        except RuntimeError:
            return
        ProductionPageWorkflowTests._seed_sampling_context(page, node_id=node_id, nodeconfig=0x00)

    @staticmethod
    def _select_node(page: ProductionPage, node_text: str = "Node 6 - H") -> None:
        combo = page.test_control_section._combo
        for index in range(combo.count()):
            if combo.itemText(index) == node_text:
                combo.setCurrentIndex(index)
                break

    def _prepare_sampling_page(self, node_text: str = "Node 6 - H") -> ProductionPage:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            self._select_node(page, node_text)
            ProductionPageWorkflowTests._seed_sampling_context(page, node_id=6, nodeconfig=0x00)
            with patch.object(SamplingTestController, "start", autospec=True, return_value=True):
                page._handle_start_sampling_requested()
                assert page._sampling_popup is not None
                page._sampling_popup.start_button.click()
                self._app.processEvents()
            assert page._sampling_popup is not None
        return page

    def _sampling_stage_button(self, page: ProductionPage) -> QPushButton:
        return page.stage_section._rows["sampling"][1]

    @staticmethod
    def _create_7nz_ipqc_workbook(path: Path) -> None:
        wb = Workbook()
        for index, base_group in enumerate(("7NZ", "7NX")):
            ws = wb.active if index == 0 else wb.create_sheet(base_group)
            if index == 0:
                ws.title = base_group
            sampling = wb.create_sheet(f"{base_group}_D")
            wb.create_sheet(f"{base_group}_A")
            ws["A1"] = "Programming"
            ws["B2"] = "Source"
            ws["C2"] = "Programmed"
            ws["D2"] = "Check"
            ws["A3"] = "Operator"
            ws["A4"] = "Assembler"
            ws["A5"] = "UUID"
            ws["A6"] = "PWM"
            ws["A7"] = "Proportionate (P)"
            ws["A8"] = "Integral (I)"
            ws["A9"] = "Derivative (D)"
            ws["A10"] = "PID_SlewRate"
            ws["A11"] = "RampDown_Slope"
            ws["A12"] = "RampDown_Step"
            ws["A13"] = "RampDown_MinVel"
            ws["A14"] = "RampDown_TargetOffset"
            ws["A15"] = "RampDown_Region"
            ws["A16"] = "Acceptable_Error"
            ws["B3"] = f"operator-{base_group.lower()}"
            ws["B4"] = f"assembler-{base_group.lower()}"
            ws["B5"] = "1223307010"
            ws["B6"] = "100"
            ws["B7"] = "0.125"
            ws["B8"] = "0.025"
            ws["B9"] = "0.010"
            ws["B10"] = "1500"
            ws["B11"] = "-25"
            ws["B12"] = "4"
            ws["B13"] = "8"
            ws["B14"] = "-12"
            ws["B15"] = "75"
            ws["B16"] = "30"
            ProductionPageWorkflowTests._populate_sampling_sheet(sampling)
        wb.save(path)

    @staticmethod
    def _mark_sampling_controller_running(controller, *args, **kwargs) -> bool:
        controller._running = True
        return True

    def test_sampling_cannot_start_before_single_axis_pass(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        self._assert_sampling_button_disabled(page)

        with patch.object(SamplingTestController, "start", autospec=True) as start_mock:
            page._handle_start_sampling_requested()
            self._app.processEvents()

        self.assertFalse(start_mock.called)
        self.assertIsNotNone(page._sampling_popup)
        assert page._sampling_popup is not None
        self.assertTrue(page._sampling_popup.isVisible())
        self.assertFalse(page._sampling_popup.start_button.isEnabled())
        self._app.processEvents()

        self.assertEqual(runtime_window.backend_client.sent_commands, [])
        self.assertIn("Sampling is available after Single Axis passes.", page.progress_section.to_plain_text())

    def test_sampling_does_not_start_when_serial_is_disconnected(self) -> None:
        runtime_window = _FakeRuntimeWindow(connected=False)
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            page._handle_ipqc_sheet_group_changed("3X")
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

            with patch.object(SamplingTestController, "start", autospec=True) as start_mock:
                page._handle_start_sampling_requested()
                self._app.processEvents()

        self.assertFalse(start_mock.called)
        self.assertEqual(runtime_window.backend_client.sent_commands, [])
        self.assertIsNotNone(page._sampling_popup)
        assert page._sampling_popup is not None
        self.assertTrue(page._sampling_popup.isVisible())
        self.assertFalse(page._sampling_popup.start_button.isEnabled())
        self.assertIn("Connect the serial link before starting Sampling.", page.progress_section.to_plain_text())

    def test_sampling_starts_after_single_axis_pass_and_uses_selected_node_and_base_group(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            self.assertTrue(self._sampling_stage_button(page).isEnabled())
            ProductionPageWorkflowTests._select_node(page, "Node 8 - RZ")

            with patch.object(SamplingTestController, "start", autospec=True, return_value=True) as start_mock:
                page._handle_start_sampling_requested()
                self._app.processEvents()

            self.assertFalse(start_mock.called)
            call_args = start_mock.call_args
            self.assertIsNone(call_args)
            self.assertIsNotNone(page._sampling_popup)
            assert page._sampling_popup is not None
            self.assertIsInstance(page._sampling_popup, SamplingTestPopup)
            self.assertEqual(page._sampling_popup.selected_node_value.text(), "Node 8 - RZ")
            self.assertEqual(page._sampling_popup.sampling_sheet_value.text(), "3X_D")
            self.assertTrue(page._sampling_popup.isVisible())
            self.assertTrue(page._sampling_popup.start_button.isEnabled())
            self.assertIn("Sampling ready for Node 8 RZ using 3X_D", page._sampling_popup.log_output.toPlainText())

    def test_sampling_context_survives_refresh_and_reopen_for_stable_base_group(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc_7nz.xlsx"
            self._create_7nz_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            page._handle_ipqc_sheet_group_changed("7NZ")
            self._select_node(page, "Node 7 - NZ")
            self._enable_single_axis_pass(page)

            self.assertTrue(page._single_axis_passed)
            self.assertIsNotNone(page._sampling_motion_polarity)
            self.assertIsNotNone(page._sampling_sensor_profile)
            self.assertTrue(self._sampling_stage_button(page).isEnabled())

            page._handle_start_sampling_requested()
            self._app.processEvents()

            popup = page._sampling_popup
            self.assertIsNotNone(popup)
            assert popup is not None
            self.assertEqual(popup.sampling_sheet_value.text(), "7NZ_D")
            self.assertTrue(popup.start_button.isEnabled())

            page.refresh()
            self._app.processEvents()

            self.assertTrue(page._single_axis_passed)
            self.assertIsNotNone(page._sampling_motion_polarity)
            self.assertIsNotNone(page._sampling_sensor_profile)
            self.assertEqual(popup.sampling_sheet_value.text(), "7NZ_D")
            self.assertTrue(popup.start_button.isEnabled())
            self.assertTrue(self._sampling_stage_button(page).isEnabled())

            popup.close()
            self._app.processEvents()
            page._handle_start_sampling_requested()
            self._app.processEvents()

            self.assertTrue(page._single_axis_passed)
            self.assertIsNotNone(page._sampling_motion_polarity)
            self.assertIsNotNone(page._sampling_sensor_profile)
            self.assertEqual(popup.sampling_sheet_value.text(), "7NZ_D")
            self.assertTrue(popup.start_button.isEnabled())
            self.assertTrue(self._sampling_stage_button(page).isEnabled())

    def test_sampling_context_clears_only_on_real_node_workbook_or_base_group_change(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc_7nz.xlsx"
            workbook_path_2 = Path(tmpdir) / "ipqc_7nz_next.xlsx"
            self._create_7nz_ipqc_workbook(workbook_path)
            self._create_7nz_ipqc_workbook(workbook_path_2)
            self._load_workbook(page, workbook_path)
            page._handle_ipqc_sheet_group_changed("7NZ")
            self._select_node(page, "Node 7 - NZ")
            self._enable_single_axis_pass(page)
            page._handle_start_sampling_requested()
            self._app.processEvents()

            popup = page._sampling_popup
            self.assertIsNotNone(popup)
            assert popup is not None

            self._select_node(page, "Node 8 - RZ")
            page._handle_test_control_node_selected()
            self._app.processEvents()

            self.assertFalse(page._single_axis_passed)
            self.assertIsNone(page._sampling_motion_polarity)
            self.assertIsNone(page._sampling_sensor_profile)
            self.assertFalse(self._sampling_stage_button(page).isEnabled())
            self.assertFalse(popup.start_button.isEnabled())

            page._handle_ipqc_sheet_group_changed("7NX")
            self._app.processEvents()
            self.assertFalse(page._single_axis_passed)
            self.assertIsNone(page._sampling_motion_polarity)
            self.assertIsNone(page._sampling_sensor_profile)
            self.assertFalse(self._sampling_stage_button(page).isEnabled())
            self.assertFalse(popup.start_button.isEnabled())
            self.assertEqual(popup.sampling_sheet_value.text(), "7NX_D")

            self._select_node(page, "Node 7 - NZ")
            self._enable_single_axis_pass(page)
            page._handle_start_sampling_requested()
            self._app.processEvents()

            self.assertTrue(page._single_axis_passed)
            self.assertIsNotNone(page._sampling_motion_polarity)
            self.assertIsNotNone(page._sampling_sensor_profile)
            self.assertTrue(self._sampling_stage_button(page).isEnabled())
            self.assertTrue(popup.start_button.isEnabled())
            self.assertEqual(popup.sampling_sheet_value.text(), "7NX_D")

            self._load_workbook(page, workbook_path_2)
            self._app.processEvents()
            self.assertFalse(page._single_axis_passed)
            self.assertIsNone(page._sampling_motion_polarity)
            self.assertIsNone(page._sampling_sensor_profile)
            self.assertFalse(self._sampling_stage_button(page).isEnabled())
            self.assertFalse(popup.start_button.isEnabled())

    def test_missing_sampling_sheet_fails_before_any_run_command_is_sent(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            wb = load_workbook(workbook_path)
            del wb["3X_D"]
            wb.save(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")
            page._single_axis_passed = True
            ProductionPageWorkflowTests._seed_sampling_context(page, node_id=6, nodeconfig=0x00)

            page._handle_start_sampling_requested()
            self._app.processEvents()

        self.assertEqual(runtime_window.backend_client.sent_commands, [])
        self.assertIn("Sampling workbook is not ready:", page.progress_section.to_plain_text())

    def test_sampling_popup_close_and_reopen_keeps_packet_updates_alive(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

            with patch.object(
                SamplingTestController,
                "start",
                autospec=True,
                side_effect=self._mark_sampling_controller_running,
            ):
                page._handle_start_sampling_requested()
                self._app.processEvents()
                assert page._sampling_popup is not None
                page._sampling_popup.start_button.click()
                self._app.processEvents()
                assert page._sampling_popup is not None
                page._sampling_popup.start_button.click()
                self._app.processEvents()
                assert page._sampling_popup is not None
                page._sampling_popup.start_button.click()
                self._app.processEvents()
                assert page._sampling_popup is not None
                page._sampling_popup.start_button.click()
                self._app.processEvents()

        assert page._sampling_popup is not None
        popup = page._sampling_popup
        popup.close()
        self._app.processEvents()
        self.assertFalse(popup.isVisible())

        page._handle_sampling_current_pwm_changed(70)
        page._handle_sampling_state_changed("SAMPLE_WAIT_SENSOR")
        self._app.processEvents()

        self.assertEqual(popup.current_pwm_value.text(), "70")
        self.assertEqual(popup.state_value.text(), "SAMPLE_WAIT_SENSOR")

        popup.show()
        self._app.processEvents()
        self.assertTrue(popup.isVisible())
        self.assertEqual(popup.current_pwm_value.text(), "70")

    def test_sampling_progress_routes_to_existing_logs(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")
            page._single_axis_passed = True
            ProductionPageWorkflowTests._seed_sampling_context(page, node_id=6, nodeconfig=0x00)

            page._handle_start_sampling_requested()
            self._app.processEvents()

        self.assertIsNotNone(page._sampling_popup)
        self.assertIsNotNone(page._sampling_popup)
        assert page._sampling_popup is not None
        self.assertIn("Sampling ready for Node 6 H using 3X_D", page._sampling_popup.log_output.toPlainText())

        fake_result = types.SimpleNamespace(
            sample_index=1,
            direction="+",
            range_value=2500481,
            elapsed_seconds=10.0176,
            speed=249608.01,
            workbook_cells={"Range": "B4", "Speed": "B21", "Time": "B38"},
        )
        page._handle_sampling_packet_message("[TX] Node 6: 88 FF 42")
        page._handle_sampling_packet_message("[RX] Node 6: 88 53 FF 42")
        page._handle_sampling_measurement_completed(fake_result)
        self._app.processEvents()

        operator_text = page._sampling_popup.log_output.toPlainText()
        packet_text = page._sampling_popup.packet_log_output.toPlainText()
        self.assertIn(
            "Sample 1/32 + complete | range=2500481 | time=10.018s | speed=249608.01",
            operator_text,
        )
        self.assertIn("[TX] Node 6: 88 FF 42", packet_text)
        self.assertIn("[RX] Node 6: 88 53 FF 42", packet_text)

    def test_stop_while_sampling_running_sends_dd(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

            with patch.object(
                SamplingTestController,
                "start",
                autospec=True,
                side_effect=self._mark_sampling_controller_running,
            ):
                page._handle_start_sampling_requested()
                self._app.processEvents()
                assert page._sampling_popup is not None
                page._sampling_popup.start_button.click()
                self._app.processEvents()
                page._sampling_popup.stop_button.click()
                self._app.processEvents()

        sent_commands = [command for _node_id, command in runtime_window.backend_client.sent_commands]
        self.assertIn([0xDD], sent_commands)

    def test_sampling_button_disabled_until_single_axis_passes(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        button = self._sampling_stage_button(page)
        self.assertFalse(button.isEnabled())
        self.assertIn("Single Axis", button.toolTip())

    def test_sampling_button_becomes_enabled_after_single_axis_pass_and_not_now(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

        self.assertTrue(self._sampling_stage_button(page).isEnabled())

    def test_single_axis_prompt_yes_starts_sampling_via_existing_path(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

            with patch.object(SamplingTestController, "start", autospec=True, return_value=True) as start_mock:
                self._enable_single_axis_pass(page, start_sampling_prompt=True)
                self._app.processEvents()

            self.assertFalse(start_mock.called)
            self.assertIsNotNone(page._sampling_popup)
            assert page._sampling_popup is not None
            self.assertTrue(page._sampling_popup.isVisible())
            self.assertTrue(page._sampling_popup.start_button.isEnabled())

    def test_single_axis_prompt_no_keeps_sampling_enabled(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page, start_sampling_prompt=False)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

        self.assertTrue(self._sampling_stage_button(page).isEnabled())

    def _assert_sampling_button_disabled(self, page: ProductionPage) -> None:
        self.assertFalse(self._sampling_stage_button(page).isEnabled())

    def test_sampling_popup_can_open_after_single_axis_pass_and_shows_context(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 8 - RZ")

            page._handle_start_sampling_requested()
            self._app.processEvents()

        self.assertIsNotNone(page._sampling_popup)
        assert page._sampling_popup is not None
        self.assertTrue(page._sampling_popup.isVisible())
        self.assertEqual(page._sampling_popup.selected_node_value.text(), "Node 8 - RZ")
        self.assertEqual(page._sampling_popup.sampling_sheet_value.text(), "3X_D")
        self.assertEqual(page._sampling_popup.final_status_value.text(), "IDLE")
        self.assertFalse(page._sampling_popup.stop_button.isEnabled())
        self.assertTrue(page._sampling_popup.start_button.isEnabled())

    def test_sampling_popup_resume_button_is_between_start_and_stop_and_disabled_initially(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 8 - RZ")

            page._handle_start_sampling_requested()
            self._app.processEvents()

        assert page._sampling_popup is not None
        popup = page._sampling_popup
        self.assertLess(popup.start_button.y(), popup.resume_button.y())
        self.assertLess(popup.resume_button.y(), popup.stop_button.y())
        self.assertLess(popup.stop_button.y(), popup.close_button.y())
        self.assertFalse(popup.resume_button.isEnabled())
        self.assertIn("Resume unavailable: Sampling has not started.", popup.resume_hint_value.text())

    def test_sampling_popup_resume_button_enables_after_abort_and_disables_after_encoder_reset_failure(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")
            selected_node_id, selected_node_name = page.test_control_section.selected_node()
            active_group = page._ipqc_excel_adapter.active_sheet_group

            popup = page._ensure_sampling_popup()
            self._app.processEvents()
            page._sampling_controller._running = False
            page._sampling_controller._state = SamplingTestController.S_ABORTED
            ProductionPageWorkflowTests._seed_sampling_context(page, node_id=6, nodeconfig=0x00)
            page._sampling_controller._resume_context = SamplingResumeContext(
                node_id=selected_node_id,
                node_name=selected_node_name,
                base_group=str(active_group or ""),
                sheet_name=page._ipqc_excel_adapter.resolve_sampling_sheet_name(active_group) if active_group else "-",
                pwm_values=(100,),
                samples_per_direction=1,
                current_pwm_index=0,
                current_pwm=100,
                current_sample_index=1,
                current_direction="HOME",
                completed_measurements=0,
                total_measurements=2,
                terminal_state=SamplingTestController.S_ABORTED,
                reason="Sampling aborted by user.",
                resumable=True,
                sample_incomplete=False,
            )
            page._refresh_sampling_action_states()

            self.assertTrue(popup.resume_button.isEnabled())
            self.assertIn("Resume from PWM 100, sample 1", popup.resume_hint_value.text())

            page._sampling_controller._resume_context = SamplingResumeContext(
                node_id=selected_node_id,
                node_name=selected_node_name,
                base_group=str(active_group or ""),
                sheet_name=page._ipqc_excel_adapter.resolve_sampling_sheet_name(active_group) if active_group else "-",
                pwm_values=(100,),
                samples_per_direction=1,
                current_pwm_index=0,
                current_pwm=100,
                current_sample_index=1,
                current_direction="+",
                completed_measurements=0,
                total_measurements=2,
                terminal_state=SamplingTestController.S_FAILED,
                reason="Unexpected encoder reset during sampling.",
                resumable=False,
                sample_incomplete=True,
            )
            page._sampling_controller._state = SamplingTestController.S_FAILED
            ProductionPageWorkflowTests._seed_sampling_context(page, node_id=6, nodeconfig=0x00)
            page._refresh_sampling_action_states()

            self.assertFalse(popup.resume_button.isEnabled())
            self.assertIn("Resume unavailable:", popup.resume_hint_value.text())
            self.assertIn("encoder reset", popup.resume_hint_value.text().lower())

    def test_sampling_popup_resume_click_reuses_runtime_session_without_crash(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")
            selected_node_id, selected_node_name = page.test_control_section.selected_node()
            active_group = page._ipqc_excel_adapter.active_sheet_group
            assert active_group is not None

            popup = page._ensure_sampling_popup()
            ProductionPageWorkflowTests._seed_sampling_context(page, node_id=6, nodeconfig=0x00)
            page._sampling_session = _SamplingSession(
                node_id=selected_node_id,
                node_name=selected_node_name,
                runtime_window=runtime_window,
            )
            page._sampling_controller._running = False
            page._sampling_controller._state = SamplingTestController.S_ABORTED
            page._sampling_controller._current_pwm = 100
            page._sampling_controller._current_sample_index = 2
            page._sampling_controller._current_direction = "-"
            page._sampling_controller._completed_measurements = 3
            page._sampling_controller._total_measurements = 4
            page._sampling_controller._resume_context = SamplingResumeContext(
                node_id=selected_node_id,
                node_name=selected_node_name,
                base_group=str(active_group),
                sheet_name=page._ipqc_excel_adapter.resolve_sampling_sheet_name(active_group),
                pwm_values=(100,),
                samples_per_direction=1,
                current_pwm_index=0,
                current_pwm=100,
                current_sample_index=2,
                current_direction="-",
                completed_measurements=3,
                total_measurements=4,
                terminal_state=SamplingTestController.S_ABORTED,
                reason="Sampling aborted by user.",
                resumable=True,
                sample_incomplete=True,
            )
            page._refresh_sampling_action_states()
            self.assertTrue(popup.resume_button.isEnabled())

            popup.set_state_text("FAILED")
            popup.set_status_text("Unexpected encoder reset during sampling.")
            popup.set_final_status("FAILED")
            popup.set_reason_text("Unexpected encoder reset during sampling.", tone="red")
            popup.set_failure_context_text("PWM 100 | Direction + | Sample 1")
            popup.set_resume_hint("Resume unavailable: old state")
            with patch.object(page._sampling_controller, "resume", wraps=page._sampling_controller.resume) as resume_mock:
                popup.resume_button.click()
                self._app.processEvents()

            self.assertEqual(resume_mock.call_count, 1)
            self.assertEqual(runtime_window.backend_client.sent_commands[0], (selected_node_id, [0x84, 0x00, 0x50]))
            self.assertTrue(page._sampling_controller.is_active())
            self.assertEqual(popup.final_status_value.text(), "RUNNING")
            self.assertEqual(popup.reason_value.text(), "-")
            self.assertEqual(popup.failure_context_value.text(), "-")
            self.assertEqual(popup.resume_hint_value.text(), "Sampling is running.")

    def test_sampling_popup_uses_compact_three_row_layout(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

            page._handle_start_sampling_requested()
            self._app.processEvents()

        assert page._sampling_popup is not None
        popup = page._sampling_popup
        titles = {
            label.text()
            for label in popup.findChildren(QLabel)
            if label.objectName() == "SectionTitle"
        }
        self.assertIn("Sampling Summary", titles)
        self.assertIn("Sampling Progress", titles)
        self.assertIn("Last Sample", titles)
        self.assertIn("Operator Log", titles)
        self.assertIn("Packet Log", titles)
        self.assertNotIn("Failure Details", titles)
        self.assertNotIn("Controls", titles)
        self.assertNotIn("Latest Workbook Cells", {label.text() for label in popup.findChildren(QLabel)})
        self.assertFalse(popup.range_mode_combo.isEnabled())
        self.assertTrue(popup.samples_per_pwm_combo.isEnabled())
        self.assertTrue(popup.pwm_selection_combo.isEnabled())
        self.assertEqual(popup.samples_per_pwm_combo.currentText(), "32")
        self.assertEqual(popup.pwm_selection_combo.currentText(), "All")
        self.assertEqual(popup.selected_pwm_values(), (100, 90, 80, 70, 60))
        self.assertEqual(popup.selected_samples_per_pwm(), 32)
        self.assertTrue(popup.clear_logs_button.isVisible())
        button_texts = {button.text() for button in popup.findChildren(QPushButton)}
        self.assertIn("Clear Logs", button_texts)
        self.assertNotIn("Save Logs", button_texts)
        self.assertTrue(popup.start_button.isVisible())
        self.assertTrue(popup.stop_button.isVisible())
        self.assertTrue(popup.close_button.isVisible())

    def test_sampling_clear_logs_clears_only_visible_log_text(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

            page._handle_start_sampling_requested()
            self._app.processEvents()

        assert page._sampling_popup is not None
        popup = page._sampling_popup
        popup.append_operator_log("Operator line")
        popup.append_packet_log("[TX] Node 6: 88 00 64")
        popup.set_state_text("RUNNING")
        popup.set_status_text("Sampling started")
        popup.set_reason_text("-")
        popup.set_current_pwm(100)
        popup.set_current_direction("+")
        popup.set_current_sample(3, 32)
        popup.set_completed_counts(8, 320)
        popup.set_latest_measurement_details(123, 4.5678, 27.5)

        popup.clear_logs_button.click()
        self._app.processEvents()

        self.assertEqual(popup.log_output.toPlainText().strip(), "")
        self.assertEqual(popup.packet_log_output.toPlainText().strip(), "")
        self.assertEqual(popup.state_value.text(), "RUNNING")
        self.assertEqual(popup.status_value.text(), "Sampling started")
        self.assertEqual(popup.reason_value.text(), "-")
        self.assertEqual(popup.current_pwm_value.text(), "100")
        self.assertEqual(popup.current_direction_value.text(), "Positive")
        self.assertEqual(popup.current_sample_value.text(), "Sample 3 / 32")
        self.assertEqual(popup.completed_count_value.text(), "8 / 320")
        self.assertEqual(popup.latest_range_value.text(), "123 counts")
        self.assertEqual(popup.latest_time_value.text(), "4.568 s")
        self.assertEqual(popup.latest_speed_value.text(), "27.50 counts/s")

    def test_sampling_failure_and_abort_reason_appears_in_summary_area(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

            page._handle_start_sampling_requested()
            self._app.processEvents()

        assert page._sampling_popup is not None
        popup = page._sampling_popup

        page._sampling_controller._latest_terminal_result = types.SimpleNamespace(
            final_status="FAILED",
            status_text="FAILED",
            reason="Unexpected encoder reset during sampling.",
            failure_context="PWM 100 | Direction + | Sample 1",
            resume_text="Unavailable - encoder reset requires a fresh start.",
        )
        page._handle_sampling_failed("Sensor timeout")
        self._app.processEvents()
        self.assertEqual(popup.status_value.text(), "FAILED")
        self.assertEqual(popup.final_status_value.text(), "FAILED")
        self.assertEqual(popup.reason_value.text(), "Unexpected encoder reset during sampling.")
        self.assertEqual(popup.failure_context_value.text(), "PWM 100 | Direction + | Sample 1")
        self.assertEqual(popup.resume_hint_value.text(), "Unavailable - encoder reset requires a fresh start.")
        self.assertEqual(popup.final_status_value.text(), "FAILED")
        self.assertIn("#dc2626", popup.reason_value.styleSheet())

        page._sampling_controller._latest_terminal_result = types.SimpleNamespace(
            final_status="ABORTED",
            status_text="ABORTED",
            reason="Sampling aborted by user.",
            failure_context="PWM 100 | Direction + | Sample 1",
            resume_text="Unavailable - sampling was aborted.",
        )
        page._handle_sampling_aborted("Operator stop")
        self._app.processEvents()
        self.assertEqual(popup.status_value.text(), "ABORTED")
        self.assertEqual(popup.reason_value.text(), "Sampling aborted by user.")
        self.assertEqual(popup.failure_context_value.text(), "PWM 100 | Direction + | Sample 1")
        self.assertEqual(popup.resume_hint_value.text(), "Unavailable - sampling was aborted.")
        self.assertEqual(popup.final_status_value.text(), "ABORTED")
        self.assertIn("#d97706", popup.reason_value.styleSheet())

    def test_popup_start_button_sends_first_sampling_run_command(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")
            page._single_axis_passed = True
            ProductionPageWorkflowTests._seed_sampling_context(page, node_id=6, nodeconfig=0x00)

            page._handle_start_sampling_requested()
            self._app.processEvents()
            assert page._sampling_popup is not None
            popup = page._sampling_popup
            popup.set_state_text("ABORTED")
            popup.set_status_text("Sensor timeout")
            popup.set_final_status("ABORTED")
            popup.set_reason_text("Sensor timeout", tone="red")
            popup.set_failure_context_text("PWM 100 | Direction + | Sample 1")
            popup.set_resume_hint("Resume unavailable: old state")
            self.assertTrue(popup.start_button.isEnabled())
            self.assertTrue(popup.samples_per_pwm_combo.isEnabled())
            self.assertTrue(popup.pwm_selection_combo.isEnabled())
            popup.start_button.click()
            self._app.processEvents()

        sent_commands = [command for _node_id, command in runtime_window.backend_client.sent_commands]
        self.assertEqual(sent_commands[0], [0x84, 0x00, 0x50])
        self.assertFalse(popup.start_button.isEnabled())
        self.assertTrue(popup.stop_button.isEnabled())
        self.assertEqual(popup.final_status_value.text(), "RUNNING")
        self.assertEqual(popup.state_value.text(), "HOME_WAIT_VEL_ACK")
        self.assertEqual(popup.reason_value.text(), "-")
        self.assertEqual(popup.failure_context_value.text(), "-")
        self.assertEqual(popup.resume_hint_value.text(), "Sampling is running.")
        self.assertEqual(popup.status_value.text(), "Setting home velocity")

    def test_popup_selected_configuration_drives_pwm_90_debug_run(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")
            page._single_axis_passed = True
            ProductionPageWorkflowTests._seed_sampling_context(page, node_id=6, nodeconfig=0x00)

            page._handle_start_sampling_requested()
            self._app.processEvents()
            assert page._sampling_popup is not None
            popup = page._sampling_popup
            popup.pwm_selection_combo.setCurrentText("90")
            popup.samples_per_pwm_combo.setCurrentText("4")
            self.assertEqual(popup.selected_pwm_values(), (90,))
            self.assertEqual(popup.selected_samples_per_pwm(), 4)
            popup.start_button.click()
            self._app.processEvents()

            self.assertFalse(popup.samples_per_pwm_combo.isEnabled())
            self.assertFalse(popup.pwm_selection_combo.isEnabled())

            runtime_window.packet_received.emit([0x84, 0x53, 0x00, 0x50])
            runtime_window.packet_received.emit([0x81, 0x53, 0x82, 0x00, 0x00, 0x00, 0x00])
            self._app.processEvents()
            runtime_window.packet_received.emit([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x00])
            self._app.processEvents()
            runtime_window.packet_received.emit([0x82, 0x00, 0x00, 0x00, 10])
            self._app.processEvents()

            self.assertEqual(runtime_window.backend_client.sent_commands[3][1], build_run(90))
            self.assertEqual(popup.current_pwm_value.text(), "90")
            self.assertEqual(popup.current_sample_value.text(), "Sample 1 / 4")

    def test_sampling_popup_fields_update_from_controller_hooks(self) -> None:
        page = self._prepare_sampling_page()
        popup = page._sampling_popup
        assert popup is not None
        page._handle_sampling_state_changed("SAMPLE_WAIT_SENSOR")
        page._handle_sampling_status_changed("Waiting for R sensor event")
        page._handle_sampling_current_pwm_changed(90)
        page._handle_sampling_current_direction_changed("+")
        page._handle_sampling_current_sample_changed(7)
        page._handle_sampling_completed_count_changed(84, 320)
        page._handle_sampling_latest_measurement_changed(180, 0.25, 720.0)
        page._handle_sampling_latest_cell_written("AG42")
        fake_result = types.SimpleNamespace(
            sample_index=7,
            direction="+",
            range_value=180,
            elapsed_seconds=0.25,
            speed=720.0,
            workbook_cells={"Range": "B4", "Speed": "B21", "Time": "B38"},
        )
        page._handle_sampling_measurement_completed(fake_result)
        self._app.processEvents()

        self.assertEqual(popup.state_value.text(), "SAMPLE_WAIT_SENSOR")
        self.assertEqual(popup.status_value.text(), "Waiting for R sensor event")
        self.assertEqual(popup.current_pwm_value.text(), "90")
        self.assertEqual(popup.current_direction_value.text(), "Positive")
        self.assertEqual(popup.current_sample_value.text(), "Sample 7 / 32")
        self.assertEqual(popup.completed_count_value.text(), "84 / 320")
        self.assertEqual(popup.latest_range_value.text(), "180 counts")
        self.assertEqual(popup.latest_time_value.text(), "0.250 s")
        self.assertEqual(popup.latest_speed_value.text(), "720.00 counts/s")
        self.assertIn(
            "Sample 7/32 + complete | range=180 | time=0.250s | speed=720.00",
            popup.log_output.toPlainText(),
        )

    def test_sampling_page_progress_log_keeps_only_high_level_messages(self) -> None:
        page = self._prepare_sampling_page()
        assert page._sampling_popup is not None
        popup = page._sampling_popup

        page._handle_sampling_state_changed("SAMPLE_WAIT_SENSOR")
        page._handle_sampling_status_changed("Waiting for R sensor event")
        page._handle_sampling_current_pwm_changed(90)
        page._handle_sampling_current_direction_changed("+")
        page._handle_sampling_current_sample_changed(7)
        page._handle_sampling_completed_count_changed(84, 320)
        page._handle_sampling_latest_measurement_changed(180, 0.25, 720.0)
        page._handle_sampling_latest_cell_written("AG42")

        progress_text = page.progress_section.to_plain_text()
        self.assertIn("TESTING: Sampling started for Node 6 H", progress_text)
        self.assertNotIn("Sampling state:", progress_text)
        self.assertNotIn("Sampling direction:", progress_text)
        self.assertNotIn("Sampling PWM:", progress_text)
        self.assertNotIn("Sampling sample index:", progress_text)
        self.assertNotIn("Sampling progress:", progress_text)
        self.assertNotIn("Latest sampling measurement:", progress_text)
        self.assertNotIn("Latest workbook cell written:", progress_text)
        self.assertNotIn("Waiting for R sensor event", progress_text)
        self.assertNotIn("Setting home velocity", progress_text)

    def test_sampling_page_progress_log_routing_keeps_lifecycle_messages_only(self) -> None:
        page = self._prepare_sampling_page()
        assert page._sampling_popup is not None
        popup = page._sampling_popup

        page._sampling_controller._resume_context = SamplingResumeContext(
            node_id=6,
            node_name="H",
            base_group="3X",
            sheet_name="3X_D",
            pwm_values=(100,),
            samples_per_direction=1,
            current_pwm_index=0,
            current_pwm=100,
            current_sample_index=1,
            current_direction="-",
            completed_measurements=1,
            total_measurements=2,
            terminal_state=SamplingTestController.S_ABORTED,
            reason="Sampling aborted by user.",
            resumable=True,
            sample_incomplete=True,
        )
        with patch.object(SamplingTestController, "resume", autospec=True, return_value=True):
            page._handle_sampling_popup_resume_requested()
            self._app.processEvents()

        self.assertIn("Sampling resumed for Node 6 H from PWM 100, sample 1", page.progress_section.to_plain_text())

        page._handle_sampling_completed()
        self._app.processEvents()
        self.assertIn("[PASS] Sampling PASSED for Node 6 H", page.progress_section.to_plain_text())

        page = self._prepare_sampling_page()
        assert page._sampling_popup is not None
        page._handle_sampling_failed("Sensor timeout")
        self._app.processEvents()
        self.assertIn("[FAIL] Sampling FAILED for Node 6 H: Sensor timeout", page.progress_section.to_plain_text())

        page = self._prepare_sampling_page()
        assert page._sampling_popup is not None
        page._handle_sampling_aborted("Operator stop")
        self._app.processEvents()
        self.assertIn("[FAIL] Sampling ABORTED for Node 6 H", page.progress_section.to_plain_text())

    def test_sampling_popup_still_receives_detailed_operator_and_packet_logs(self) -> None:
        page = self._prepare_sampling_page()
        assert page._sampling_popup is not None
        popup = page._sampling_popup
        page.progress_section.clear_log()

        page._handle_sampling_log("Sampling direction: +")
        page._handle_sampling_packet_message("[TX] Node 6: 88 00 64")
        page._handle_sampling_measurement_completed(
            types.SimpleNamespace(
                sample_index=1,
                direction="+",
                range_value=120,
                elapsed_seconds=0.25,
                speed=480.0,
                workbook_cells={"Range": "B4", "Speed": "B21", "Time": "B38"},
            )
        )

        popup_text = popup.log_output.toPlainText()
        packet_text = popup.packet_log_output.toPlainText()
        self.assertIn("Sampling direction: +", popup_text)
        self.assertIn("Sample 1/32 + complete | range=120 | time=0.250s | speed=480.00", popup_text)
        self.assertIn("[TX] Node 6: 88 00 64", packet_text)
        self.assertNotIn("Sampling direction: +", page.progress_section.to_plain_text())
        self.assertNotIn("[TX] Node 6: 88 00 64", page.progress_section.to_plain_text())

    def test_sampling_terminal_states_reenable_start_and_disable_stop(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")
            ProductionPageWorkflowTests._seed_sampling_context(page, node_id=6, nodeconfig=0x00)

            with patch.object(SamplingTestController, "start", autospec=True, return_value=True):
                page._handle_start_sampling_requested()
                self._app.processEvents()

        assert page._sampling_popup is not None
        popup = page._sampling_popup
        self.assertTrue(popup.start_button.isEnabled())
        self.assertFalse(popup.stop_button.isEnabled())

        popup.start_button.click()
        self._app.processEvents()

        self.assertFalse(popup.start_button.isEnabled())
        self.assertTrue(popup.stop_button.isEnabled())

        page._sampling_controller._running = False
        page._sampling_controller._latest_terminal_result = types.SimpleNamespace(
            final_status="COMPLETED",
            status_text="Sampling completed",
            reason="-",
            failure_context="-",
            resume_text="-",
        )
        page._handle_sampling_completed()
        self._app.processEvents()
        self.assertTrue(self._sampling_stage_button(page).isEnabled())
        self.assertTrue(popup.start_button.isEnabled())
        self.assertFalse(popup.stop_button.isEnabled())
        self.assertEqual(popup.status_value.text(), "Sampling completed")
        self.assertEqual(popup.reason_value.text(), "-")
        self.assertEqual(popup.failure_context_value.text(), "-")
        self.assertEqual(popup.resume_hint_value.text(), "-")

        with patch.object(
            SamplingTestController,
            "start",
            autospec=True,
            side_effect=self._mark_sampling_controller_running,
        ):
            page._handle_start_sampling_requested()
            self._app.processEvents()
            assert page._sampling_popup is not None
            page._sampling_popup.start_button.click()
            self._app.processEvents()

        assert page._sampling_popup is not None
        popup = page._sampling_popup
        page._sampling_controller._running = False
        page._sampling_controller._latest_terminal_result = types.SimpleNamespace(
            final_status="FAILED",
            status_text="FAILED",
            reason="Sensor timeout.",
            failure_context="PWM 100 | Direction + | Sample 1",
            resume_text="Unavailable - sampling requires a fresh start.",
        )
        page._handle_sampling_failed("boom")
        self._app.processEvents()
        self.assertTrue(self._sampling_stage_button(page).isEnabled())
        self.assertTrue(popup.start_button.isEnabled())
        self.assertFalse(popup.stop_button.isEnabled())
        self.assertEqual(popup.status_value.text(), "FAILED")
        self.assertEqual(popup.reason_value.text(), "Sensor timeout.")
        self.assertEqual(popup.failure_context_value.text(), "PWM 100 | Direction + | Sample 1")
        self.assertEqual(popup.resume_hint_value.text(), "Unavailable - sampling requires a fresh start.")

        with patch.object(
            SamplingTestController,
            "start",
            autospec=True,
            side_effect=self._mark_sampling_controller_running,
        ):
            page._handle_start_sampling_requested()
            self._app.processEvents()
            assert page._sampling_popup is not None
            page._sampling_popup.start_button.click()
            self._app.processEvents()

        assert page._sampling_popup is not None
        popup = page._sampling_popup
        page._sampling_controller._running = False
        page._sampling_controller._latest_terminal_result = types.SimpleNamespace(
            final_status="ABORTED",
            status_text="ABORTED",
            reason="Sampling aborted by user.",
            failure_context="PWM 100 | Direction + | Sample 1",
            resume_text="Unavailable - sampling was aborted.",
        )
        page._handle_sampling_aborted("stop")
        self._app.processEvents()
        self.assertTrue(self._sampling_stage_button(page).isEnabled())
        self.assertTrue(popup.start_button.isEnabled())
        self.assertFalse(popup.stop_button.isEnabled())
        self.assertEqual(popup.status_value.text(), "ABORTED")
        self.assertEqual(popup.reason_value.text(), "Sampling aborted by user.")
        self.assertEqual(popup.failure_context_value.text(), "PWM 100 | Direction + | Sample 1")
        self.assertEqual(popup.resume_hint_value.text(), "Unavailable - sampling was aborted.")

    def test_sampling_popup_stop_requests_abort_once(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

            with patch.object(
                SamplingTestController,
                "start",
                autospec=True,
                side_effect=self._mark_sampling_controller_running,
            ), patch.object(page._sampling_controller, "abort_by_user", return_value=True) as abort_mock:
                page._handle_start_sampling_requested()
                self._app.processEvents()
                assert page._sampling_popup is not None
                popup = page._sampling_popup
                popup.start_button.click()
                self._app.processEvents()
                page._handle_sampling_latest_cell_written("B2")
                popup.stop_requested.emit()
                self._app.processEvents()

            self.assertEqual(abort_mock.call_count, 1)
        self.assertTrue(page._sampling_popup is not None)

    def test_sampling_popup_close_does_not_stop_sampling_and_reopen_keeps_state(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

            with patch.object(
                SamplingTestController,
                "start",
                autospec=True,
                side_effect=self._mark_sampling_controller_running,
            ), patch.object(page._sampling_controller, "abort_by_user", return_value=True) as abort_mock:
                page._handle_start_sampling_requested()
                self._app.processEvents()
                assert page._sampling_popup is not None
                popup = page._sampling_popup
                popup.start_button.click()
                self._app.processEvents()
                page._handle_sampling_current_pwm_changed(100)
                page._handle_sampling_current_direction_changed("+")
                popup.close()
                self._app.processEvents()
                self.assertFalse(popup.isVisible())
                self.assertEqual(popup.current_pwm_value.text(), "100")
                popup.show()
                self._app.processEvents()
                self.assertTrue(popup.isVisible())
                self.assertEqual(popup.current_pwm_value.text(), "100")
                self.assertFalse(abort_mock.called)

    def test_start_while_sampling_running_reuses_popup_without_second_start(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            ProductionPageWorkflowTests._create_ipqc_workbook(workbook_path)
            self._load_workbook(page, workbook_path)
            self._enable_single_axis_pass(page)
            ProductionPageWorkflowTests._select_node(page, "Node 6 - H")

            with patch.object(
                SamplingTestController,
                "start",
                autospec=True,
                side_effect=self._mark_sampling_controller_running,
            ) as start_mock:
                page._handle_start_sampling_requested()
                self._app.processEvents()
                self.assertTrue(page._sampling_popup is not None and page._sampling_popup.isVisible())
                assert page._sampling_popup is not None
                page._sampling_popup.start_button.click()
                self._app.processEvents()
                page._handle_start_sampling_requested()
                self._app.processEvents()

            self.assertEqual(start_mock.call_count, 1)
            self.assertIsNotNone(page._sampling_popup)
            assert page._sampling_popup is not None
            self.assertTrue(page._sampling_popup.isVisible())

    def test_reset_workbook_parameter_workflow_clears_inflight_state(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        defs = {definition.name: definition for definition in default_workbook_parameter_definitions()}
        request = controller.build_parameter_request(defs["PWM"], 3, "X", "100")

        self.assertTrue(controller.write_parameters([request]))
        self.assertEqual(controller._parameter_operation_mode, "write")
        self.assertIsNotNone(controller._pending_parameter_request)
        self.assertEqual(controller._parameter_requests, [request])

        controller.reset_workbook_parameter_workflow()

        self.assertIsNone(controller._parameter_operation_mode)
        self.assertIsNone(controller._pending_parameter_request)
        self.assertEqual(controller._parameter_requests, [])
        self.assertEqual(controller._parameter_results, [])
        self.assertIsNone(controller._pending_eeprom_save)
        self.assertFalse(controller._eeprom_settle_active)
        self.assertIsNone(controller.last_verify_actual_uuid)
        self.assertEqual(controller.last_verify_actual_uuid_text, "")
        self.assertEqual(controller.last_verify_raw_response_hex, "")
        self.assertIsNone(controller.last_verify_actual_pwm)
        self.assertEqual(controller.last_verify_actual_pwm_text, "")
        self.assertEqual(controller.last_verify_pwm_raw_response_hex, "")

    def test_write_pwm_sends_set_pwm_payload_to_selected_node(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge)

        ok, _message = controller.write_pwm(6, "H", 100, expected_pwm_text="100")
        self.assertTrue(ok)
        self.assertEqual(runtime_window.backend_client.sent_commands[0], (6, [0x84, 0x00, 0x64]))

    def test_verify_pwm_uses_getvel_and_decodes_response(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        events: list[tuple[bool, str]] = []
        controller.pwm_verification_finished.connect(lambda passed, reason: events.append((passed, reason)))

        self.assertTrue(controller.verify_pwm(6, "H", 80, expected_pwm_text="80"))
        self.assertEqual(runtime_window.backend_client.sent_commands[0], (6, [0x85]))
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x85, "params": [0x00, 0x50]}
        )
        self._app.processEvents()
        self.assertTrue(events)
        self.assertTrue(events[-1][0])
        self.assertEqual(controller.last_verify_actual_pwm, 80)
        self.assertEqual(controller.last_verify_actual_pwm_text, "80")

    def test_verify_pwm_accepts_expected_values_as_string_and_int(self) -> None:
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}
        for expected_input in ("10", 10):
            with self.subTest(expected_input=expected_input):
                runtime_window = _FakeRuntimeWindow()
                bridge = _FakeBridge(runtime_window)
                controller = ProductionParameterController(bridge, timeout_ms=100)
                events: list[tuple[bool, str, object]] = []
                controller.parameter_verification_finished.connect(lambda passed, reason, results: events.append((passed, reason, results)))

                request = controller.build_parameter_request(definitions["PWM"], 6, "H", expected_input)
                self.assertTrue(controller.verify_parameters([request]))
                self.assertEqual(runtime_window.backend_client.sent_commands[0], (6, [0x85]))
                runtime_window.packet_received.emit(
                    {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x85, "params": [0x00, 0x0A]}
                )
                self._app.processEvents()

                self.assertTrue(events)
                self.assertTrue(events[-1][0])
                result = events[-1][2][0]
                self.assertEqual(result.definition.name, "PWM")
                self.assertEqual(result.actual_text, "10")

    def test_verify_pwm_fails_when_actual_value_differs(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}
        events: list[tuple[bool, str, object]] = []
        controller.parameter_verification_finished.connect(lambda passed, reason, results: events.append((passed, reason, results)))

        request = controller.build_parameter_request(definitions["PWM"], 6, "H", "10")
        self.assertTrue(controller.verify_parameters([request]))
        self.assertEqual(runtime_window.backend_client.sent_commands[0], (6, [0x85]))
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x85, "params": [0x00, 0x50]}
        )
        self._app.processEvents()

        self.assertTrue(events)
        self.assertFalse(events[-1][0])
        self.assertIn("expected 10, actual 80", events[-1][1])

    def test_verify_loaded_uuid_times_out_when_response_is_missing(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=20)
        events: list[tuple[bool, str]] = []
        controller.verification_finished.connect(lambda passed, reason: events.append((passed, reason)))

        csv_text = "node_id,node_name,uuid\n6,H,1223306010\n"
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "uuid_valid.csv"
            csv_path.write_text(csv_text, encoding="utf-8")
            self.assertTrue(controller.load_uuid_csv(str(csv_path)))

        self.assertTrue(controller.verify_loaded_uuid(6, "H"))
        self.assertEqual(runtime_window.backend_client.sent_commands[0], (6, [0xE0, 0x3F]))

        deadline = time.monotonic() + 0.5
        while time.monotonic() < deadline and not events:
            self._app.processEvents()
            time.sleep(0.01)

        self.assertTrue(events)
        self.assertFalse(events[-1][0])
        self.assertIn("Timed out", events[-1][1])

    def test_parameter_pipeline_verifies_uuid_and_pwm_with_shared_definitions(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}
        requests = [
            controller.build_parameter_request(definitions["UUID"], 6, "H", "1223306010"),
            controller.build_parameter_request(definitions["PWM"], 6, "H", "80"),
        ]
        events: list[tuple[bool, str, object]] = []
        controller.parameter_verification_finished.connect(lambda passed, reason, results: events.append((passed, reason, results)))

        self.assertTrue(controller.verify_parameters(requests))
        self.assertEqual(runtime_window.backend_client.sent_commands[0], (6, [0xE0, 0x3F]))
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0xE0, "params": [0x3A, *build_uuid_write_payload(1223306010)[2:]]}
        )
        self._app.processEvents()
        self.assertEqual(runtime_window.backend_client.sent_commands[-1], (6, [0x85]))
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x85, "params": [0x00, 0x50]}
        )
        self._app.processEvents()

        self.assertTrue(events)
        self.assertTrue(events[-1][0])
        result_names = [result.definition.name for result in events[-1][2]]
        self.assertEqual(result_names, ["UUID", "PWM"])

    def test_parameter_pipeline_dummy_definition_uses_shared_flow(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        dummy = ParameterDefinition(
            name="DUMMY",
            expected_cell="B9",
            actual_cell="C9",
            result_cell="D9",
            command_id=0x92,
            write_operator=0x3D,
            read_operator=0x3F,
            write_response_operator=0x3A,
            read_response_operator=0x3A,
            value_size=1,
            signed=False,
            persistent=True,
            sub_id=None,
            parse_expected=lambda value: int(str(value).strip()),
            build_write_command=lambda value: [0x92, int(value) & 0xFF],
            build_read_command=lambda: [0x92],
            decode_response=lambda payload: (True, payload[1], "") if len(payload) > 1 else (False, None, "short"),
            format_actual=lambda actual, _expected: str(actual),
            compare=lambda expected, _expected_text, actual, _actual_text: int(expected) == int(actual),
        )
        request = controller.build_parameter_request(dummy, 6, "H", "7")
        events: list[tuple[bool, str, object]] = []
        controller.parameter_verification_finished.connect(lambda passed, reason, results: events.append((passed, reason, results)))

        ok, message = controller.write_parameters([request])
        self.assertTrue(ok, message)
        self.assertEqual(runtime_window.backend_client.sent_commands[-1], (6, [0x92, 0x07]))
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x92, "params": [0x07]}
        )
        self._app.processEvents()
        self.assertTrue(controller.verify_parameters([request]))
        self.assertEqual(runtime_window.backend_client.sent_commands[-1], (6, [0x92]))
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x92, "params": [0x07]}
        )
        self._app.processEvents()

        self.assertTrue(events[-1][0])
        self.assertEqual(events[-1][2][0].definition.name, "DUMMY")

    def test_parameter_pipeline_dummy_definition_mismatch_fails(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        dummy = ParameterDefinition(
            name="DUMMY",
            expected_cell="B9",
            actual_cell="C9",
            result_cell="D9",
            command_id=0x92,
            write_operator=0x3D,
            read_operator=0x3F,
            write_response_operator=0x3A,
            read_response_operator=0x3A,
            value_size=1,
            signed=False,
            persistent=True,
            sub_id=None,
            parse_expected=lambda value: int(str(value).strip()),
            build_write_command=lambda value: [0x92, int(value) & 0xFF],
            build_read_command=lambda: [0x92],
            decode_response=lambda payload: (True, payload[1], "") if len(payload) > 1 else (False, None, "short"),
            format_actual=lambda actual, _expected: str(actual),
            compare=lambda expected, _expected_text, actual, _actual_text: int(expected) == int(actual),
        )
        request = controller.build_parameter_request(dummy, 6, "H", "7")
        events: list[tuple[bool, str, object]] = []
        controller.parameter_verification_finished.connect(lambda passed, reason, results: events.append((passed, reason, results)))

        self.assertTrue(controller.verify_parameters([request]))
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x92, "params": [0x09]}
        )
        self._app.processEvents()

        self.assertTrue(events)
        self.assertFalse(events[-1][0])
        self.assertIn("expected 7, actual 9", events[-1][1])

    def test_parameter_pipeline_dummy_string_definition_uses_shared_flow(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        dummy = ParameterDefinition(
            name="DUMMY-ID",
            expected_cell="B10",
            actual_cell="C10",
            result_cell="D10",
            command_id=0x93,
            write_operator=None,
            read_operator=0x3F,
            write_response_operator=None,
            read_response_operator=0x3A,
            value_size=1,
            signed=False,
            persistent=True,
            sub_id=None,
            parse_expected=lambda value: str(value).strip().upper(),
            build_write_command=None,
            build_read_command=lambda: [0x93],
            decode_response=lambda payload: (True, chr(payload[1]), "") if len(payload) > 1 else (False, None, "short"),
            format_actual=lambda actual, _expected: str(actual),
            compare=lambda expected, _expected_text, actual, _actual_text: str(expected) == str(actual),
        )
        request = controller.build_parameter_request(dummy, 6, "H", " a ")
        events: list[tuple[bool, str, object]] = []
        controller.parameter_verification_finished.connect(lambda passed, reason, results: events.append((passed, reason, results)))

        self.assertTrue(controller.verify_parameters([request]))
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": 0x93, "params": [ord("A")]}
        )
        self._app.processEvents()

        self.assertTrue(events)
        self.assertTrue(events[-1][0])
        self.assertEqual(events[-1][2][0].actual_text, "A")

    def test_parameter_pipeline_uuid_timeout_still_fails_with_timeout(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=20)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}
        request = controller.build_parameter_request(definitions["UUID"], 6, "H", "1223306010")
        events: list[tuple[bool, str, object]] = []
        controller.parameter_verification_finished.connect(lambda passed, reason, results: events.append((passed, reason, results)))

        self.assertTrue(controller.verify_parameters([request]))
        deadline = time.monotonic() + 0.5
        while time.monotonic() < deadline and not events:
            self._app.processEvents()
            time.sleep(0.01)

        self.assertTrue(events)
        self.assertFalse(events[-1][0])
        self.assertIn("actual timeout", events[-1][1])

    def test_verify_parameters_blocks_during_eeprom_settle_and_allows_after(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}
        request = controller.build_parameter_request(definitions["UUID"], 6, "H", "1223306010")
        events: list[tuple[bool, str, object]] = []
        controller.parameter_verification_finished.connect(lambda passed, reason, results: events.append((passed, reason, results)))

        self.assertTrue(controller.save_parameters_to_eeprom(6, "H"))
        self.assertEqual(runtime_window.backend_client.sent_commands[-1], (6, [EEPROM_SAVE_COMMAND, SET_COMMAND_SUFFIX]))
        runtime_window.packet_received.emit(
            {"status": "ok", "type": "can_over_uart", "sender": 6, "cmd": EEPROM_SAVE_COMMAND, "params": [0x0A, 0x00]}
        )
        self._app.processEvents()

        sent_before_verify = len(runtime_window.backend_client.sent_commands)
        self.assertFalse(controller.verify_parameters([request]))
        self.assertEqual(len(runtime_window.backend_client.sent_commands), sent_before_verify)
        self.assertTrue(events)
        self.assertFalse(events[-1][0])
        self.assertIn("settle is still active", events[-1][1])

        controller._handle_eeprom_settle_timeout()
        self.assertTrue(controller.verify_parameters([request]))
        self.assertEqual(runtime_window.backend_client.sent_commands[-1], (6, [0xE0, 0x3F]))

    def test_load_uuid_csv_validation_blocks_invalid_rows(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge)

        csv_text = "node_id,node_name,uuid\n6,H,1223305010\n2,Y,1223302010\n"
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "uuid_invalid.csv"
            csv_path.write_text(csv_text, encoding="utf-8")
            is_valid = controller.load_uuid_csv(str(csv_path))

        self.assertFalse(is_valid)
        self.assertEqual(controller.rows, [])
        self.assertGreaterEqual(len(controller.errors), 1)

    def test_write_loaded_uuid_sends_write_payload_only(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge)

        csv_text = "node_id,node_name,uuid\n6,H,1223306010\n"
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "uuid_valid.csv"
            csv_path.write_text(csv_text, encoding="utf-8")
            self.assertTrue(controller.load_uuid_csv(str(csv_path)))

        ok, _message = controller.write_loaded_uuid(6, "H")
        self.assertTrue(ok)
        self.assertEqual(
            runtime_window.backend_client.sent_commands[0],
            (6, [0xE0, 0x3D, 0x00, 0x48, 0xEA, 0x2B, 0x1A]),
        )
        self.assertEqual(len(runtime_window.backend_client.sent_commands), 1)

    def test_verify_loaded_uuid_passes_for_matching_response(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        events: list[tuple[bool, str]] = []
        controller.verification_finished.connect(lambda passed, reason: events.append((passed, reason)))

        csv_text = "node_id,node_name,uuid\n6,H,1223306010\n"
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "uuid_valid.csv"
            csv_path.write_text(csv_text, encoding="utf-8")
            self.assertTrue(controller.load_uuid_csv(str(csv_path)))

        self.assertTrue(controller.verify_loaded_uuid(6, "H"))
        self.assertEqual(runtime_window.backend_client.sent_commands[0], (6, [0xE0, 0x3F]))
        runtime_window.packet_received.emit(
            {
                "status": "ok",
                "type": "can_over_uart",
                "sender": 6,
                "cmd": 0xE0,
                "params": [0x3A, 0x00, 0x48, 0xEA, 0x2B, 0x1A],
            }
        )
        self._app.processEvents()

        self.assertTrue(events)
        self.assertTrue(events[-1][0])

    def test_verify_loaded_uuid_fails_for_wrong_node(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        events: list[tuple[bool, str]] = []
        controller.verification_finished.connect(lambda passed, reason: events.append((passed, reason)))

        csv_text = "node_id,node_name,uuid\n6,H,1223306010\n"
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "uuid_valid.csv"
            csv_path.write_text(csv_text, encoding="utf-8")
            self.assertTrue(controller.load_uuid_csv(str(csv_path)))

        self.assertTrue(controller.verify_loaded_uuid(6, "H"))
        runtime_window.packet_received.emit(
            {
                "status": "ok",
                "type": "can_over_uart",
                "sender": 5,
                "cmd": 0xE0,
                "params": [0x3A, 0x00, 0x48, 0xEA, 0x2B, 0x1A],
            }
        )
        self._app.processEvents()

        self.assertTrue(events)
        self.assertFalse(events[-1][0])
        self.assertIn("wrong node", events[-1][1])

    def test_verify_loaded_uuid_fails_when_selected_node_missing_in_csv(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        controller = ProductionParameterController(bridge, timeout_ms=100)
        events: list[tuple[bool, str]] = []
        controller.verification_finished.connect(lambda passed, reason: events.append((passed, reason)))

        csv_text = "node_id,node_name,uuid\n6,H,1223306010\n"
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "uuid_valid.csv"
            csv_path.write_text(csv_text, encoding="utf-8")
            self.assertTrue(controller.load_uuid_csv(str(csv_path)))

        self.assertFalse(controller.verify_loaded_uuid(5, "V"))
        self.assertTrue(events)
        self.assertFalse(events[-1][0])
        self.assertIn("No UUID CSV row found", events[-1][1])

    def test_uuid_section_removes_legacy_uuid_csv_controls_and_keeps_workbook_flow(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        title_labels = [label.text() for label in page.uuid_section.findChildren(QLabel) if label.objectName() == "PanelTitle"]
        self.assertIn("IPQC Workbook Parameter Programming", title_labels)
        self.assertFalse(hasattr(page.uuid_section, "load_button"))
        self.assertFalse(hasattr(page.uuid_section, "_file_label"))
        self.assertFalse(hasattr(page.uuid_section, "_validation_label"))
        self.assertFalse(hasattr(page.uuid_section, "_preview_table"))
        self.assertEqual(page.uuid_section.load_workbook_button.text(), "Load IPQC Workbook")
        self.assertEqual(page.uuid_section.write_button.text(), "Write Parameters to MCU")
        self.assertEqual(page.uuid_section.verify_button.text(), "Read Back / Verify")
        self.assertEqual(page.uuid_section.save_button.text(), "Save / Download Completed Workbook")
        self.assertFalse(page.uuid_section.verify_button.isEnabled())
        self.assertFalse(page.uuid_section.write_button.isEnabled())
        self.assertFalse(page.uuid_section.save_button.isEnabled())
        self.assertEqual(page.uuid_section.last_workbook_action_text, "-")
        self.assertFalse(hasattr(page.uuid_section, "_result_csv_label"))
        button_texts = [button.text() for button in page.findChildren(type(page.uuid_section.write_button))]
        self.assertNotIn("Echo Test", button_texts)
        self.assertNotIn("Safe Movement Test", button_texts)
        self.assertFalse(hasattr(page.test_control_section, "_profile_combo"))
        stage_labels = [label.text() for label in page.stage_section.findChildren(QLabel)]
        self.assertIn("Single Axis Functional Test", stage_labels)
        self.assertIn("Sampling Test", stage_labels)
        self.assertIn("Performance Test", stage_labels)
        stage_button_texts = [button.text() for button in page.stage_section.findChildren(QPushButton)]
        self.assertEqual(stage_button_texts.count("Start Test"), 3)

    def test_single_axis_stage_opens_functional_popup(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        page._handle_single_axis_test_requested()
        self._app.processEvents()

        self.assertIsNotNone(page._single_axis_popup)
        assert page._single_axis_popup is not None
        self.assertIsInstance(page._single_axis_popup, SingleAxisFunctionalPopup)
        self.assertEqual(page._single_axis_popup.windowTitle(), "Functional")

    def test_single_axis_popup_warns_when_run_without_node(self) -> None:
        popup = SingleAxisFunctionalPopup(node_options=[(3, "X")])
        with patch("gui.workspace.pages.single_axis_functional_popup.QMessageBox.warning") as warning_mock:
            popup._handle_run_clicked()
        self.assertTrue(warning_mock.called)
        self.assertTrue(popup.run_button.isEnabled())
        self.assertTrue(popup.node_combo.isEnabled())

    def test_single_axis_popup_placeholder_run_updates_status_and_reenables_controls(self) -> None:
        # Suppress any modal dialogs from the placeholder pass flow
        with patch(
            "gui.workspace.pages.single_axis_functional_popup.QMessageBox.information",
            return_value=None,
        ), patch(
            "gui.workspace.pages.single_axis_functional_popup.QMessageBox.warning",
            return_value=None,
        ), patch(
            "gui.workspace.pages.single_axis_functional_popup.SingleAxisFunctionalPopup.ask_start_sampling",
            return_value=False,
        ):
            popup = SingleAxisFunctionalPopup(node_options=[(3, "X")], allow_safe_tx=True)
            popup.node_combo.setCurrentIndex(1)
            popup._handle_run_clicked()

            # Legacy placeholder flow is no longer step-driven. Simulate completion
            # by invoking the pass handler directly to verify UI re-enables and status updates.
            popup.mark_passed()

            status_text = popup.status_block.toPlainText()
            self.assertIn("Functional test PASSED.", status_text)
            self.assertTrue(popup.run_button.isEnabled())
            self.assertTrue(popup.node_combo.isEnabled())

            # Ensure the dialog is closed to avoid stray timers in CI
            popup.close()

    def test_production_page_blocks_write_and_verify_when_no_workbook_loaded(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        self.assertFalse(page.uuid_section.write_button.isEnabled())
        self.assertFalse(page.uuid_section.verify_button.isEnabled())
        page._handle_write_uuid()
        page._handle_verify_uuid()
        self._app.processEvents()

        self.assertEqual(runtime_window.backend_client.sent_commands, [])
        self.assertEqual(page.result_summary_section._status_label.text(), "FAIL")
        self.assertIn("Load an IPQC workbook first", page.result_summary_section._reason_label.text())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook guard tests.")
    def test_production_page_blocks_uuid_actions_when_no_supported_parameters_are_loaded(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)
            wb = load_workbook(workbook_path)
            for row in range(5, 17):
                wb["3X"][f"A{row}"] = f"Unsupported {row}"
            wb.save(workbook_path)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            self.assertFalse(page.uuid_section.verify_button.isEnabled())
            self.assertFalse(page.uuid_section.write_button.isEnabled())
            page._handle_write_uuid()
            page._handle_verify_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands, [])
            self.assertEqual(page.result_summary_section._status_label.text(), "FAIL")
            self.assertIn("No supported parameter rows found", page.progress_section.to_plain_text())
            self.assertIn("Labels found:", page.progress_section.to_plain_text())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook guard tests.")
    def test_production_page_blocks_uuid_actions_when_workbook_expected_sn_invalid(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)
            wb = load_workbook(workbook_path)
            wb["3X"]["B5"] = "not-a-uuid"
            wb.save(workbook_path)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
                self._app.processEvents()

            self.assertFalse(page.uuid_section.verify_button.isEnabled())
            self.assertFalse(page.uuid_section.write_button.isEnabled())
            page._handle_write_uuid()
            page._handle_verify_uuid()
            self._app.processEvents()
            self.assertEqual(runtime_window.backend_client.sent_commands, [])
            self.assertEqual(page.result_summary_section._status_label.text(), "FAIL")
            self.assertIn("Expected S/N in workbook B5 is invalid", page.result_summary_section._reason_label.text())

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook guard tests.")
    def test_parameter_verification_cache_skips_passed_results_and_finishes_after_retry(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
            self._app.processEvents()

        self._select_node(page, "Node 6 - H")
        requests = [
            page._parameter_controller.build_parameter_request(definitions["UUID"], 6, "H", "1223303010"),
            page._parameter_controller.build_parameter_request(definitions["PWM"], 6, "H", "100"),
        ]
        calls: list[list[str]] = []

        def fake_verify_parameters(requests_to_verify):
            names = [request.definition.name for request in requests_to_verify]
            calls.append(names)
            if len(calls) == 1:
                page._handle_parameter_verification_finished(
                    False,
                    "PWM read-back verification - expected 100, actual timeout",
                    [
                        ParameterVerificationResult(
                            definition=definitions["UUID"],
                            expected_text="1223303010",
                            actual_text="1223303010",
                            passed=True,
                            reason="UUID read-back verification",
                        ),
                        ParameterVerificationResult(
                            definition=definitions["PWM"],
                            expected_text="100",
                            actual_text="",
                            passed=False,
                            reason="PWM read-back verification - expected 100, actual timeout",
                        ),
                    ],
                )
            else:
                page._handle_parameter_verification_finished(
                    True,
                    "Workbook parameter read-back verification",
                    [
                        ParameterVerificationResult(
                            definition=definitions["PWM"],
                            expected_text="100",
                            actual_text="100",
                            passed=True,
                            reason="PWM read-back verification",
                        )
                    ],
                )
            return True

        with patch.object(page, "_build_workbook_parameter_requests", return_value=requests), patch.object(
            page._parameter_controller,
            "verify_parameters",
            side_effect=fake_verify_parameters,
        ):
            page._handle_verify_uuid()
            self._app.processEvents()
            page._handle_verify_uuid()
            self._app.processEvents()

        self.assertEqual(calls[0], ["UUID", "PWM"])
        self.assertEqual(calls[1], ["PWM"])
        self.assertTrue(page._last_parameter_verification_results_by_name["UUID"].passed)
        self.assertEqual(page.uuid_section.workbook_validation_text, "Workbook Validation: PASSED")

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook guard tests.")
    def test_parameter_verification_cache_clears_on_workbook_load_node_change_and_written_parameter_update(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            workbook_path_2 = Path(tmpdir) / "ipqc_2.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)
            self._create_ipqc_workbook(workbook_path_2, with_optional_fields=False)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
            self._app.processEvents()

            requests = [
                page._parameter_controller.build_parameter_request(definitions["UUID"], 6, "H", "1223303010"),
                page._parameter_controller.build_parameter_request(definitions["PWM"], 6, "H", "100"),
            ]
            verify_calls: list[list[str]] = []

            def fake_verify_parameters(requests_to_verify):
                names = [request.definition.name for request in requests_to_verify]
                verify_calls.append(names)
                if names == ["PWM"]:
                    page._handle_parameter_verification_finished(
                        True,
                        "Workbook parameter read-back verification",
                        [
                            ParameterVerificationResult(
                                definition=definitions["PWM"],
                                expected_text="100",
                                actual_text="100",
                                passed=True,
                                reason="PWM read-back verification",
                            )
                        ],
                    )
                elif len(verify_calls) == 1:
                    page._handle_parameter_verification_finished(
                        False,
                        "PWM read-back verification - expected 100, actual timeout",
                        [
                            ParameterVerificationResult(
                                definition=definitions["UUID"],
                                expected_text="1223303010",
                                actual_text="1223303010",
                                passed=True,
                                reason="UUID read-back verification",
                            ),
                            ParameterVerificationResult(
                                definition=definitions["PWM"],
                                expected_text="100",
                                actual_text="",
                                passed=False,
                                reason="PWM read-back verification - expected 100, actual timeout",
                            ),
                        ],
                    )
                else:
                    page._handle_parameter_verification_finished(
                        True,
                        "Workbook parameter read-back verification",
                        [
                            ParameterVerificationResult(
                                definition=definitions["UUID"],
                                expected_text="1223303010",
                                actual_text="1223303010",
                                passed=True,
                                reason="UUID read-back verification",
                            ),
                            ParameterVerificationResult(
                                definition=definitions["PWM"],
                                expected_text="100",
                                actual_text="100",
                                passed=True,
                                reason="PWM read-back verification",
                            ),
                        ],
                    )
                return True

            with patch.object(page, "_build_workbook_parameter_requests", return_value=requests), patch.object(
                page._parameter_controller,
                "verify_parameters",
                side_effect=fake_verify_parameters,
            ):
                page._handle_verify_uuid()
                self._app.processEvents()

                with patch(
                    "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                    return_value=(str(workbook_path_2), "Excel Files (*.xlsx)"),
                ):
                    page._handle_load_ipqc_workbook()
                self._app.processEvents()

                page._handle_verify_uuid()
                self._app.processEvents()

                self._select_node(page, "Node 8 - RZ")
                page._handle_test_control_node_selected()
                page._handle_verify_uuid()
                self._app.processEvents()

            self.assertEqual(verify_calls[0], ["UUID", "PWM"])
            self.assertEqual(verify_calls[1], ["UUID", "PWM"])
            self.assertEqual(verify_calls[2], ["UUID", "PWM"])
            self.assertTrue(page._last_parameter_verification_results_by_name["UUID"].passed)
            self.assertEqual(page.uuid_section.workbook_validation_text, "Workbook Validation: PASSED")

    @unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC workbook guard tests.")
    def test_parameter_verification_cache_is_invalidated_by_written_parameter_update(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        definitions = {definition.name: definition for definition in default_workbook_parameter_definitions()}

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "ipqc.xlsx"
            self._create_ipqc_workbook(workbook_path, with_optional_fields=False)
            with patch(
                "gui.workspace.pages.production_page.QFileDialog.getOpenFileName",
                return_value=(str(workbook_path), "Excel Files (*.xlsx)"),
            ):
                page._handle_load_ipqc_workbook()
            self._app.processEvents()

            requests = [
                page._parameter_controller.build_parameter_request(definitions["UUID"], 6, "H", "1223303010"),
                page._parameter_controller.build_parameter_request(definitions["PWM"], 6, "H", "100"),
            ]
            verify_calls: list[list[str]] = []
            write_calls: list[list[str]] = []

            def fake_verify_parameters(requests_to_verify):
                names = [request.definition.name for request in requests_to_verify]
                verify_calls.append(names)
                if len(verify_calls) == 1:
                    page._handle_parameter_verification_finished(
                        False,
                        "PWM read-back verification - expected 100, actual timeout",
                        [
                            ParameterVerificationResult(
                                definition=definitions["UUID"],
                                expected_text="1223303010",
                                actual_text="1223303010",
                                passed=True,
                                reason="UUID read-back verification",
                            ),
                            ParameterVerificationResult(
                                definition=definitions["PWM"],
                                expected_text="100",
                                actual_text="",
                                passed=False,
                                reason="PWM read-back verification - expected 100, actual timeout",
                            ),
                        ],
                    )
                else:
                    page._handle_parameter_verification_finished(
                        True,
                        "Workbook parameter read-back verification",
                        [
                            ParameterVerificationResult(
                                definition=definitions["PWM"],
                                expected_text="100",
                                actual_text="100",
                                passed=True,
                                reason="PWM read-back verification",
                            )
                        ],
                    )
                return True

            def fake_write_parameters(requests_to_write):
                names = [request.definition.name for request in requests_to_write]
                write_calls.append(names)
                page._handle_parameter_write_finished(True, "PWM write sent to Node 6 H.")
                return True, "PWM write sent to Node 6 H."

            with patch.object(page, "_build_workbook_parameter_requests", return_value=requests), patch.object(
                page._parameter_controller,
                "verify_parameters",
                side_effect=fake_verify_parameters,
            ), patch.object(page._parameter_controller, "write_parameters", side_effect=fake_write_parameters):
                page._handle_verify_uuid()
                self._app.processEvents()

                page._handle_write_uuid()
                self._app.processEvents()
                self.assertEqual(write_calls[0], ["PWM"])

                page._handle_verify_uuid()
                self._app.processEvents()

            self.assertEqual(verify_calls[0], ["UUID", "PWM"])
            self.assertEqual(verify_calls[1], ["PWM"])
            self.assertTrue(page._last_parameter_verification_results_by_name["UUID"].passed)
            self.assertEqual(page.uuid_section.workbook_validation_text, "Workbook Validation: PASSED")

    def test_single_axis_return_leg_range_display_uses_middle_travel_delta(self) -> None:
        controller = SingleAxisFunctionalTestController(FunctionalTestConfig(reference_sensor="L", opposite_sensor="R"))
        polarity = decode_nodeconfig_motion_polarity(0x00)
        controller._node_id = 6
        controller._motion_polarity = polarity
        controller._sensor_profile = NodeSensorProfile.from_node_context(6, polarity)
        popup = SingleAxisFunctionalPopup(node_options=[(3, "X")], controller=controller, allow_safe_tx=True)
        differences: list[int] = []
        controller.difference_changed = lambda value: differences.append(value)

        controller._state = controller.S_READ_RANGE1
        controller._wait_for = "getpos_r1"
        controller._handle_getpos(("G", 2_499_678))
        self.assertEqual(popup.range_field.text(), "2499678")

        controller._state = controller.S_READ_RANGE2
        controller._wait_for = "getpos_r2"
        controller._opposite_pos = 2_499_678
        controller._range_1 = 2_499_678
        controller._handle_getpos(("G", -100))

        self.assertEqual(popup.range_field.text(), "1249939")
        self.assertEqual(differences[-1], 100)
        popup.close()
        self._app.processEvents()

    def test_communication_logs_button_reuses_popup_and_keeps_progress_log_intact(self) -> None:
        runtime_window = _FakeRuntimeWindow()
        bridge = _FakeBridge(runtime_window)
        page = ProductionPage(bridge)
        store = runtime_window.communication_log_store

        store.record_out(
            bytes.fromhex("25 A5 06 0B 31 03 88 FF 42 00 00"),
            decoded_line="[N6:H] RUN 255 66 (-190)",
        )
        page.progress_section.append_step("Progress line")

        page._handle_view_logs_requested()
        dialog = page._communication_logs_dialog
        assert dialog is not None
        self.assertIn("[OUT] 25 A5 06 0B 31 03 88 FF 42 00 00 (11)", dialog.log_output.toPlainText())

        dialog.hide()
        store.record_in(
            bytes.fromhex("C8 24 06 0A 25 A5 06 01 31 02 81 4C"),
            decoded_lines=["                              [N6:H] TPOS 'L'"],
        )
        page._handle_view_logs_requested()
        self.assertIs(page._communication_logs_dialog, dialog)
        self.assertIn("[IN ] C8 24 06 0A 25 A5 06 01 31 02 81 4C (12)", dialog.log_output.toPlainText())

        with tempfile.TemporaryDirectory() as temp_dir:
            save_path = Path(temp_dir) / "20260619_093855_communication.log"
            with patch(
                "gui.workspace.dialogs.communication_log_dialog.QFileDialog.getSaveFileName",
                return_value=(str(save_path), "Log Files (*.log)"),
            ):
                dialog._handle_save_clicked()
            saved_text = save_path.read_text(encoding="utf-8")
            self.assertIn("IPQC Communication Log", saved_text)
            self.assertIn("Current Page: Production", saved_text)
            self.assertIn("Selected Node:", saved_text)
            self.assertIn("[OUT] 25 A5 06 0B 31 03 88 FF 42 00 00 (11)", saved_text)

        dialog.clear_button.click()
        self.assertEqual(store.to_plain_text().strip(), "")
        self.assertIn("Progress line", page.progress_section.to_plain_text())

        page._handle_view_logs_requested()
        self.assertIs(page._communication_logs_dialog, dialog)
        self.assertEqual(dialog.log_output.toPlainText().strip(), "")

if __name__ == "__main__":
    unittest.main()
