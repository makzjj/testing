"""Focused tests for the Sampling controller behavior."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - environment dependent.
    _HAS_OPENPYXL = False

from data.binary_cmd_builders import build_tpos, build_vel
from data.binary_cmd_parser import decode_nodeconfig_motion_polarity
from gui.workspace.controllers.sampling_test_controller import SamplingResumeContext, SamplingTestConfig, SamplingTestController
from gui.workspace.pages.production_parameter_controller import build_run
from services.ipqc_excel_adapter import IpqcExcelAdapter
from services.node_sensor_profile import NodeSensorProfile


def _create_ipqc_workbook(path: Path) -> None:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to create IPQC workbook fixtures.")
    wb = Workbook()
    ws = wb.active
    ws.title = "3X"
    sampling_3x = wb.create_sheet("3X_D")
    wb.create_sheet("3X_A")
    ws["A1"] = "Programming"
    ws["B2"] = "Source"
    ws["C2"] = "Programmed"
    ws["D2"] = "Check"
    ws["A3"] = "Operator"
    ws["A4"] = "UUID"
    ws["A5"] = "PWM"
    ws["A6"] = "Proportionate (P)"
    ws["A7"] = "Integral (I)"
    ws["A8"] = "Derivative (D)"
    ws["A9"] = "PID_SlewRate"
    ws["A10"] = "RampDown_Slope"
    ws["A11"] = "RampDown_Step"
    ws["A12"] = "RampDown_MinVel"
    ws["A13"] = "RampDown_TargetOffset"
    ws["A14"] = "RampDown_Region"
    ws["A15"] = "Acceptable_Error"
    ws["B4"] = "1223303010"
    ws["B5"] = "100"
    _populate_sampling_sheet(sampling_3x)
    wb.save(path)


def _populate_sampling_sheet(sheet) -> None:
    pwm_values = [100, 90, 80, 70, 60]
    section_starts = {"Range": 1, "Speed": 18, "Time": 35}
    for section_name, start_row in section_starts.items():
        sheet[f"A{start_row}"] = section_name
        row = start_row + 1
        for pwm in pwm_values:
            sheet[f"A{row}"] = f"PWM {pwm}"
            sheet[f"A{row + 1}"] = f"+{pwm}"
            sheet[f"A{row + 2}"] = f"-{pwm}"
            row += 3


class _SamplingManualClock:
    def __init__(self, values: list[float]) -> None:
        self._values = list(values)
        self._index = 0

    def __call__(self) -> float:
        if self._index >= len(self._values):
            raise AssertionError("Sampling manual clock was exhausted.")
        value = self._values[self._index]
        self._index += 1
        return value


class _RecordingSamplingController(SamplingTestController):
    def __init__(self, adapter: IpqcExcelAdapter, config: SamplingTestConfig, clock, *, node_id: int = 8) -> None:
        super().__init__(adapter, config, clock=clock)
        self.commands: list[list[int]] = []
        self.logs: list[str] = []
        self.states: list[str] = []
        self.statuses: list[str] = []
        self.pwms: list[int] = []
        self.directions: list[str] = []
        self.sample_indices: list[int] = []
        self.completed_counts: list[tuple[int, int]] = []
        self.measurements: list[object] = []
        self.cells: list[str] = []
        self.failures: list[str] = []
        self.aborts: list[str] = []
        self.completed_called = False
        raw_nodeconfig = 0x00 if config.home_sensor == "L" else 0x03
        polarity = decode_nodeconfig_motion_polarity(raw_nodeconfig, allow_unvalidated=True)
        self.set_motion_polarity(polarity)
        self.set_sensor_profile(NodeSensorProfile.from_node_context(node_id, polarity))

    def command_requested(self, payload: list[int]) -> None:
        self.commands.append(list(payload))

    def log_message(self, text: str) -> None:
        self.logs.append(text)

    def state_changed(self, text: str) -> None:
        self.states.append(text)

    def status_changed(self, text: str) -> None:
        self.statuses.append(text)

    def current_pwm_changed(self, pwm: int) -> None:
        self.pwms.append(int(pwm))

    def current_direction_changed(self, direction: str) -> None:
        self.directions.append(direction)

    def current_sample_changed(self, sample_index: int) -> None:
        self.sample_indices.append(int(sample_index))

    def samples_completed_changed(self, completed: int, total: int) -> None:
        self.completed_counts.append((int(completed), int(total)))

    def measurement_completed(self, result) -> None:
        self.measurements.append(result)

    def latest_workbook_cell_written(self, cell_ref: str) -> None:
        self.cells.append(cell_ref)

    def sampling_completed(self) -> None:
        self.completed_called = True

    def sampling_failed(self, reason: str) -> None:
        self.failures.append(reason)

    def sampling_aborted(self, reason: str) -> None:
        self.aborts.append(reason)


@unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for Sampling controller tests.")
class SamplingControllerTests(unittest.TestCase):
    def _build_adapter(self, tmpdir: str) -> IpqcExcelAdapter:
        workbook_path = Path(tmpdir) / "sampling_ipqc.xlsx"
        _create_ipqc_workbook(workbook_path)
        adapter = IpqcExcelAdapter()
        adapter.load_template(workbook_path)
        return adapter

    def _build_sampling_controller(
        self,
        tmpdir: str,
        clock_values: list[float],
        *,
        pwm_values: tuple[int, ...] = (100,),
        samples_per_direction: int = 1,
        home_sensor: str = "L",
        node_id: int = 8,
    ) -> _RecordingSamplingController:
        adapter = self._build_adapter(tmpdir)
        clock = _SamplingManualClock(clock_values)
        return _RecordingSamplingController(
            adapter,
            SamplingTestConfig(
                home_velocity=-190,
                home_sensor=home_sensor,
                pwm_values=pwm_values,
                samples_per_direction=samples_per_direction,
            ),
            clock,
            node_id=node_id,
        )

    def _drive_home_sequence(
        self,
        controller: _RecordingSamplingController,
        *,
        start_pos: int,
        home_packet: list[int] | None = None,
    ) -> None:
        self.assertEqual(controller.commands[0], [0x84, 0x00, 0x50])
        controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
        self.assertEqual(controller.commands[1], [0x81, 0x00, 0x00, 0x00, 0x00])
        controller.handle_runtime_packet(home_packet or [0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x00])
        controller.handle_runtime_packet([0x82, *list(int(start_pos).to_bytes(4, "big", signed=True))])

    @staticmethod
    def _sensor_packet(sensor: str, *, z_form: bool = False) -> list[int]:
        sensor_map = {"L": 0x4C, "R": 0x52, "I": 0x49}
        if z_form:
            return [0x81, 0x5A, sensor_map[sensor]]
        return [0x81, sensor_map[sensor]]

    @staticmethod
    def _tpos_home_packet(event: str, position: int = 0) -> list[int]:
        if event in {"L", "R", "I"}:
            return [0x81, ord(event)]
        if event == "ZL":
            return [0x81, 0x5A, 0x4C]
        if event == "ZR":
            return [0x81, 0x5A, 0x52]
        if event == "E":
            return [0x81, ord("E"), 0x82, *list(int(position).to_bytes(4, "big", signed=True))]
        if event == "N":
            return [0x81, ord("N"), 0x82, *list(int(position).to_bytes(4, "big", signed=True))]
        raise ValueError(f"Unsupported TPOS home packet event: {event}")

    @staticmethod
    def _getpos_packet(position: int) -> list[int]:
        return [0x82, *list(int(position).to_bytes(4, "big", signed=True))]

    @staticmethod
    def _assert_no_dd(controller: _RecordingSamplingController) -> None:
        assert controller.commands.count([0xDD]) == 0

    @staticmethod
    def _assert_getpos_sent(controller: _RecordingSamplingController) -> None:
        assert controller.commands[-1] == [0x82]

    @staticmethod
    def _build_resume_context(
        *,
        node_id: int,
        node_name: str,
        base_group: str,
        sheet_name: str,
        home_sensor: str = "L",
        pwm: int,
        sample_index: int,
        current_direction: str,
        completed_measurements: int,
        total_measurements: int,
        terminal_state: str,
        reason: str,
        resumable: bool,
        sample_incomplete: bool,
        current_pwm_index: int = 0,
    ) -> SamplingResumeContext:
        polarity = decode_nodeconfig_motion_polarity(0x00 if home_sensor == "L" else 0x03, allow_unvalidated=True)
        profile = NodeSensorProfile.from_node_context(node_id, polarity)
        return SamplingResumeContext(
            node_id=node_id,
            node_name=node_name,
            base_group=base_group,
            sheet_name=sheet_name,
            pwm_values=(pwm,),
            samples_per_direction=1,
            current_pwm_index=current_pwm_index,
            current_pwm=pwm,
            current_sample_index=sample_index,
            current_direction=current_direction,
            completed_measurements=completed_measurements,
            total_measurements=total_measurements,
            terminal_state=terminal_state,
            reason=reason,
            resumable=resumable,
            sample_incomplete=sample_incomplete,
            home_sensor=home_sensor,
            sensor_profile_name=profile.profile_name,
        )

    def test_sampling_run_payloads_cover_signed_velocities(self) -> None:
        cases = [
            (100, [0x88, 0x00, 0x64]),
            (-100, [0x88, 0xFF, 0x9C]),
            (90, [0x88, 0x00, 0x5A]),
            (-90, [0x88, 0xFF, 0xA6]),
            (-190, [0x88, 0xFF, 0x42]),
            (80, [0x84, 0x00, 0x50]),
        ]
        for velocity, expected in cases:
            with self.subTest(velocity=velocity):
                if velocity == 80:
                    self.assertEqual(build_vel(velocity), expected)
                else:
                    self.assertEqual(build_run(velocity), expected)

    def test_sampling_controller_owns_state_sequence_for_home_startup(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            adapter = self._build_adapter(tmpdir)
            clock = _SamplingManualClock([0.0])
            controller = _RecordingSamplingController(
                adapter,
                SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                clock,
            )

            self.assertTrue(controller.start(8, "RZ"))
            self.assertEqual(controller.states[0], "HOME_WAIT_VEL_ACK")
            controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
            self.assertIn("HOME_WAIT_TPOS", controller.states)

    def test_sampling_resume_from_abort_or_timeout_rehomes_to_first_pwm_sample(self) -> None:
        cases = [
            {
                "label": "abort",
                "prepare": "abort",
                "resume_summary": "Resume from PWM 100, sample 1",
            },
            {
                "label": "timeout",
                "prepare": "timeout",
                "resume_summary": "Resume from PWM 100, sample 1",
            },
        ]
        for case in cases:
            with self.subTest(case=case["label"]):
                with tempfile.TemporaryDirectory() as tmpdir:
                    adapter = self._build_adapter(tmpdir)
                    clock_values = [0.0, 0.05] if case["prepare"] == "timeout" else [0.0]
                    controller = _RecordingSamplingController(
                        adapter,
                        SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                        _SamplingManualClock(clock_values),
                    )

                    self.assertTrue(controller.start(8, "RZ"))
                    self._drive_home_sequence(controller, start_pos=10)
                    if case["prepare"] == "abort":
                        controller.abort_by_user()
                    else:
                        controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
                        controller.on_timeout()
                    self.assertTrue(controller.can_resume)
                    self.assertEqual(controller.resume_summary, case["resume_summary"])

                    start_len = len(controller.commands)
                    self.assertTrue(controller.resume(node_id=8, node_name="RZ", base_group="3X"))
                    self.assertEqual(controller.commands[start_len], [0x84, 0x00, 0x50])
                    controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
                    self.assertEqual(controller.commands[start_len + 1], [0x81, 0x00, 0x00, 0x00, 0x00])
                    controller.handle_runtime_packet([0x81, 0x53, 0x82, 0x00, 0x00, 0x00, 0x00])
                    controller.handle_runtime_packet([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x00])
                    controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 10])
                    self.assertEqual(controller.commands[start_len + 3], [0x88, 0x00, 0x64])

    def test_sampling_resume_from_abort_or_timeout_rehomes_to_first_pwm_sample_for_r_home(self) -> None:
        cases = [
            {
                "label": "abort",
                "prepare": "abort",
            },
            {
                "label": "timeout",
                "prepare": "timeout",
            },
        ]
        for case in cases:
            with self.subTest(case=case["label"]):
                with tempfile.TemporaryDirectory() as tmpdir:
                    adapter = self._build_adapter(tmpdir)
                    clock_values = [0.0, 0.05] if case["prepare"] == "timeout" else [0.0]
                    controller = _RecordingSamplingController(
                        adapter,
                        SamplingTestConfig(home_velocity=-190, home_sensor="R", pwm_values=(100,), samples_per_direction=1),
                        _SamplingManualClock(clock_values),
                    )

                    self.assertTrue(controller.start(8, "RZ"))
                    self._drive_home_sequence(controller, start_pos=10, home_packet=self._tpos_home_packet("R"))
                    if case["prepare"] == "abort":
                        controller.abort_by_user()
                    else:
                        controller.handle_runtime_packet([0x88, 0x53, 0xFF, 0x9C])
                        controller.on_timeout()
                    self.assertTrue(controller.can_resume)

                    start_len = len(controller.commands)
                    self.assertTrue(controller.resume(node_id=8, node_name="RZ", base_group="3X"))
                    self.assertEqual(controller.commands[start_len], [0x84, 0x00, 0x50])
                    controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
                    self.assertEqual(controller.commands[start_len + 1], [0x81, 0x00, 0x00, 0x00, 0x00])
                    controller.handle_runtime_packet(self._tpos_home_packet("R"))
                    controller.handle_runtime_packet(self._getpos_packet(10))
                    self.assertEqual(controller.commands[start_len + 3], [0x88, 0xFF, 0x9C])

    def test_sampling_resume_from_positive_failure_rehomes_and_restarts_positive_leg(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            adapter = self._build_adapter(tmpdir)
            clock = _SamplingManualClock([0.0, 0.05, 1.0])
            controller = _RecordingSamplingController(
                adapter,
                SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                clock,
            )

            self.assertTrue(controller.start(8, "RZ"))
            self._drive_home_sequence(controller, start_pos=10)
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.on_timeout()

            self.assertTrue(controller.can_resume)
            self.assertIn("sample 1", controller.resume_summary.lower())

            start_len = len(controller.commands)
            self.assertTrue(controller.resume(node_id=8, node_name="RZ", base_group="3X"))
            self.assertEqual(controller.commands[start_len], [0x84, 0x00, 0x50])
            controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
            controller.handle_runtime_packet([0x81, 0x53, 0x82, 0x00, 0x00, 0x00, 0x00])
            controller.handle_runtime_packet([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x00])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 10])
            self.assertEqual(controller.commands[start_len + 3], [0x88, 0x00, 0x64])

    def test_sampling_resume_from_negative_failure_overwrites_positive_cells(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            adapter = self._build_adapter(tmpdir)
            clock = _SamplingManualClock([0.0, 0.5, 1.0, 1.4, 2.0, 2.4])
            write_calls: list[tuple[str, int, str, int, object]] = []
            original_write = adapter.write_sampling_result

            def recording_write(
                section: str,
                pwm: int,
                direction: str,
                sample_index: int,
                value,
                *,
                base_group: str | None = None,
            ):
                if section == "Range":
                    write_calls.append((section, pwm, direction, sample_index, value))
                return original_write(section, pwm, direction, sample_index, value, base_group=base_group)

            adapter.write_sampling_result = recording_write  # type: ignore[assignment]
            controller = _RecordingSamplingController(
                adapter,
                SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                clock,
            )

            self.assertTrue(controller.start(8, "RZ"))
            self._drive_home_sequence(controller, start_pos=10)
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 70])
            controller.handle_runtime_packet([0x88, 0x53, 0xFF, 0x9C])
            controller.on_timeout()

            self.assertTrue(controller.can_resume)
            self.assertEqual(write_calls.count(("Range", 100, "+", 1, 60)), 1)

            start_len = len(controller.commands)
            self.assertTrue(controller.resume(node_id=8, node_name="RZ", base_group="3X"))
            self.assertEqual(controller.commands[start_len], [0x84, 0x00, 0x50])
            controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
            controller.handle_runtime_packet([0x81, 0x53, 0x82, 0x00, 0x00, 0x00, 0x00])
            controller.handle_runtime_packet([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x00])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 10])
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 70])

            self.assertEqual(write_calls.count(("Range", 100, "+", 1, 60)), 2)
            self.assertEqual(controller.commands[start_len + 3], [0x88, 0x00, 0x64])

    def test_sampling_resume_availability_failure_cases(self) -> None:
        cases = [
            {
                "label": "encoder_reset",
                "setup": "encoder_reset",
                "node_id": 8,
                "node_name": "RZ",
                "base_group": "3X",
                "reason_text": "encoder reset",
            },
            {
                "label": "wrong_node",
                "setup": "abort",
                "node_id": 9,
                "node_name": "PZ",
                "base_group": "3X",
                "reason_text": "original sampling node",
            },
            {
                "label": "wrong_sheet",
                "setup": "abort",
                "node_id": 8,
                "node_name": "RZ",
                "base_group": "3X_A",
                "reason_text": "original sampling sheet",
            },
            {
                "label": "motor_fault",
                "setup": "motor_fault",
                "node_id": 8,
                "node_name": "RZ",
                "base_group": "3X",
                "reason_text": "motor fault",
            },
        ]
        for case in cases:
            with self.subTest(case=case["label"]):
                with tempfile.TemporaryDirectory() as tmpdir:
                    adapter = self._build_adapter(tmpdir)
                    clock = _SamplingManualClock([0.0, 0.05])
                    controller = _RecordingSamplingController(
                        adapter,
                        SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                        clock,
                    )

                    self.assertTrue(controller.start(8, "RZ"))
                    self._drive_home_sequence(controller, start_pos=10)
                    if case["setup"] == "encoder_reset":
                        controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
                        controller.handle_runtime_packet([0x81, 0x49])
                    elif case["setup"] == "abort":
                        controller.abort_by_user()
                    elif case["setup"] == "motor_fault":
                        controller._resume_context = SamplingResumeContext(
                            node_id=8,
                            node_name="RZ",
                            base_group="3X",
                            sheet_name="3X_D",
                            pwm_values=(100,),
                            samples_per_direction=1,
                            current_pwm_index=0,
                            current_pwm=100,
                            current_sample_index=1,
                            current_direction="+",
                            completed_measurements=0,
                            total_measurements=2,
                            terminal_state=SamplingTestController.S_FAILED,
                            reason="Motor fault reported during Sampling.",
                            resumable=False,
                            sample_incomplete=True,
                        )
                        controller._running = False
                        controller._state = SamplingTestController.S_FAILED
                    enabled, reason = controller.resume_availability(
                        node_id=case["node_id"],
                        node_name=case["node_name"],
                        base_group=case["base_group"],
                    )
                    self.assertFalse(enabled)
                    self.assertIn(case["reason_text"], reason.lower())

    def test_sampling_home_tpos_accepts_configured_home_sensor_forms_without_dd(self) -> None:
        cases = [
            ("L/E", "L", [0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x0A], [0x88, 0x00, 0x64]),
            ("L/N", "L", [0x81, 0x4E, 0x82, 0x00, 0x00, 0x00, 0x0A], [0x88, 0x00, 0x64]),
            ("L/direct", "L", [0x81, 0x4C], [0x88, 0x00, 0x64]),
            ("L/Z", "L", [0x81, 0x5A, 0x4C], [0x88, 0x00, 0x64]),
            ("R/E", "R", [0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x0A], [0x88, 0xFF, 0x9C]),
            ("R/N", "R", [0x81, 0x4E, 0x82, 0x00, 0x00, 0x00, 0x0A], [0x88, 0xFF, 0x9C]),
            ("R/direct", "R", [0x81, 0x52], [0x88, 0xFF, 0x9C]),
            ("R/Z", "R", [0x81, 0x5A, 0x52], [0x88, 0xFF, 0x9C]),
        ]
        for label, home_sensor, home_packet, expected_run in cases:
            with self.subTest(packet=label):
                with tempfile.TemporaryDirectory() as tmpdir:
                    controller = self._build_sampling_controller(tmpdir, [0.0], home_sensor=home_sensor)

                    self.assertTrue(controller.start(8, "RZ"))
                    controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
                    controller.handle_runtime_packet([0x81, 0x53, 0x82, 0x00, 0x00, 0x00, 0x00])
                    controller.handle_runtime_packet(home_packet)
                    self._assert_getpos_sent(controller)
                    self._assert_no_dd(controller)

                    controller.handle_runtime_packet(self._getpos_packet(10))
                    self.assertEqual(controller.commands[-1], expected_run)
                    self.assertEqual(controller.failures, [])

    def test_sampling_home_tpos_rejects_opposite_sensor_for_each_configuration(self) -> None:
        cases = [
            ("L", [0x81, 0x52]),
            ("L", [0x81, 0x5A, 0x52]),
            ("R", [0x81, 0x4C]),
            ("R", [0x81, 0x5A, 0x4C]),
        ]
        for home_sensor, opposite_packet in cases:
            with self.subTest(home_sensor=home_sensor, packet=opposite_packet):
                with tempfile.TemporaryDirectory() as tmpdir:
                    controller = self._build_sampling_controller(tmpdir, [0.0], home_sensor=home_sensor)

                    self.assertTrue(controller.start(8, "RZ"))
                    controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
                    controller.handle_runtime_packet([0x81, 0x53, 0x82, 0x00, 0x00, 0x00, 0x00])
                    controller.handle_runtime_packet(opposite_packet)

                    self.assertFalse(controller.is_active())
                    self.assertEqual(controller.commands[-1], [0xDD])
                    self.assertTrue(controller.failures)
                    self.assertIn("Unexpected packet", controller.failures[-1])

    def test_sampling_home_getpos_duplicate_configured_home_sensor_is_ignored(self) -> None:
        cases = [
            ("L", self._tpos_home_packet("L"), self._sensor_packet("L"), 10),
            ("L", self._tpos_home_packet("L",), self._sensor_packet("L", z_form=True), 10),
            ("R", self._tpos_home_packet("R"), self._sensor_packet("R"), 10),
            ("R", self._tpos_home_packet("R",), self._sensor_packet("R", z_form=True), 10),
        ]
        for home_sensor, home_packet, duplicate_packet, start_pos in cases:
            with self.subTest(home_sensor=home_sensor, duplicate=duplicate_packet):
                with tempfile.TemporaryDirectory() as tmpdir:
                    controller = self._build_sampling_controller(tmpdir, [0.0], home_sensor=home_sensor)

                    self.assertTrue(controller.start(8, "RZ"))
                    controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
                    controller.handle_runtime_packet([0x81, 0x53, 0x82, 0x00, 0x00, 0x00, 0x00])
                    controller.handle_runtime_packet(home_packet)
                    self._assert_getpos_sent(controller)
                    controller.handle_runtime_packet(duplicate_packet)
                    self._assert_getpos_sent(controller)
                    self._assert_no_dd(controller)
                    controller.handle_runtime_packet(self._getpos_packet(start_pos))

                    self.assertFalse(controller.failures)
                    self.assertTrue(controller.is_active())

    def test_sampling_home_getpos_rejects_opposite_sensor_and_encoder_reset(self) -> None:
        cases = [
            ("L", [0x81, 0x52]),
            ("L", [0x81, 0x49]),
            ("R", [0x81, 0x4C]),
            ("R", [0x81, 0x49]),
        ]
        for home_sensor, packet in cases:
            with self.subTest(home_sensor=home_sensor, packet=packet):
                with tempfile.TemporaryDirectory() as tmpdir:
                    controller = self._build_sampling_controller(tmpdir, [0.0], home_sensor=home_sensor)

                    self.assertTrue(controller.start(8, "RZ"))
                    controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
                    controller.handle_runtime_packet([0x81, 0x53, 0x82, 0x00, 0x00, 0x00, 0x00])
                    controller.handle_runtime_packet(self._tpos_home_packet(home_sensor))
                    controller.handle_runtime_packet(packet)

                    self.assertFalse(controller.is_active())
                    self.assertEqual(controller.commands[-1], [0xDD])
                    self.assertTrue(controller.failures)
                    if packet == [0x81, 0x49]:
                        self.assertIn("Unexpected encoder reset", controller.failures[-1])
                        self.assertFalse(controller.can_resume)
                    else:
                        self.assertIn("Unexpected packet", controller.failures[-1])

    def test_sampling_run_signs_and_workbook_rows_follow_home_sensor_configuration(self) -> None:
        cases = [
            ("L", [0x88, 0x00, 0x64], [0x88, 0xFF, 0x9C], ["+", "-"]),
            ("R", [0x88, 0xFF, 0x9C], [0x88, 0x00, 0x64], ["-", "+"]),
        ]
        for home_sensor, outward_command, return_command, directions in cases:
            with self.subTest(home_sensor=home_sensor):
                with tempfile.TemporaryDirectory() as tmpdir:
                    adapter = self._build_adapter(tmpdir)
                    write_calls: list[tuple[str, str, int, object]] = []
                    original_write = adapter.write_sampling_result

                    def recording_write(section: str, pwm: int, direction: str, sample_index: int, value, *, base_group: str | None = None):
                        if section == "Range":
                            write_calls.append((section, direction, sample_index, value))
                        return original_write(section, pwm, direction, sample_index, value, base_group=base_group)

                    adapter.write_sampling_result = recording_write  # type: ignore[assignment]
                    controller = _RecordingSamplingController(
                        adapter,
                        SamplingTestConfig(home_velocity=-190, home_sensor=home_sensor, pwm_values=(100,), samples_per_direction=1),
                        _SamplingManualClock([0.0, 0.5, 1.0, 1.4, 2.0, 2.4]),
                    )

                    self.assertTrue(controller.start(8, "RZ"))
                    self._drive_home_sequence(
                        controller,
                        start_pos=10,
                        home_packet=self._tpos_home_packet(home_sensor),
                    )
                    expected_outward_sensor = "R" if home_sensor == "L" else "L"
                    expected_return_sensor = home_sensor
                    self.assertEqual(controller.commands[3], list(outward_command))
                    controller.handle_runtime_packet([outward_command[0], 0x53, outward_command[1], outward_command[2]])
                    controller.handle_runtime_packet(self._sensor_packet(expected_outward_sensor))
                    controller.handle_runtime_packet(self._getpos_packet(70))
                    self.assertEqual(controller.commands[-1], list(return_command))
                    controller.handle_runtime_packet([return_command[0], 0x53, return_command[1], return_command[2]])
                    controller.handle_runtime_packet(self._sensor_packet(expected_return_sensor))
                    controller.handle_runtime_packet(self._getpos_packet(20))
                    self.assertEqual(controller.commands[-1][0], 0x81)
                    controller.handle_runtime_packet([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x19])
                    self.assertFalse(controller.failures)
                    self.assertEqual(controller.measurements[0].direction, directions[0])
                    self.assertEqual(controller.measurements[1].direction, directions[1])
                    self.assertEqual(
                        [entry[1] for entry in write_calls[:2]],
                        list(directions),
                    )

    def test_sampling_middle_target_uses_home_and_opposite_positions_for_both_home_sensors(self) -> None:
        cases = [
            ("L", 10, 70, 40),
            ("R", 10, -70, -30),
        ]
        for home_sensor, home_pos, opposite_pos, expected_middle in cases:
            with self.subTest(home_sensor=home_sensor):
                with tempfile.TemporaryDirectory() as tmpdir:
                    controller = self._build_sampling_controller(
                        tmpdir,
                        [0.0, 0.5, 1.0, 1.4, 2.0, 2.4],
                        home_sensor=home_sensor,
                        pwm_values=(100,),
                        samples_per_direction=1,
                    )

                    self.assertTrue(controller.start(8, "RZ"))
                    self._drive_home_sequence(
                        controller,
                        start_pos=home_pos,
                        home_packet=self._tpos_home_packet(home_sensor),
                    )
                    outward_velocity = 100 if home_sensor == "L" else -100
                    return_velocity = -outward_velocity
                    self.assertEqual(controller.commands[3], build_run(outward_velocity))
                    controller.handle_runtime_packet([0x88, 0x53, *build_run(outward_velocity)[1:]])
                    controller.handle_runtime_packet(self._sensor_packet("R" if home_sensor == "L" else "L"))
                    controller.handle_runtime_packet(self._getpos_packet(opposite_pos))
                    self.assertEqual(controller.commands[5], build_run(return_velocity))
                    controller.handle_runtime_packet([0x88, 0x53, *build_run(return_velocity)[1:]])
                    controller.handle_runtime_packet(self._sensor_packet(home_sensor))
                    controller.handle_runtime_packet(self._getpos_packet(home_pos))

                    self.assertEqual(controller.commands[-1], build_tpos(expected_middle))
                    self.assertEqual(controller.state, controller.S_WAIT_MIDDLE_TPOS)
                    self.assertFalse(controller.failures)
                    self.assertFalse(controller.aborts)

    def test_sampling_duplicate_sensor_while_waiting_for_getpos_is_ignored(self) -> None:
        cases = [
            ("positive direct", self._sensor_packet("R"), self._sensor_packet("R"), 70, [0x88, 0xFF, 0x9C], 1),
            ("positive z", self._sensor_packet("R", z_form=True), self._sensor_packet("R"), 70, [0x88, 0xFF, 0x9C], 1),
            ("negative direct", self._sensor_packet("L"), self._sensor_packet("L"), 20, [0x88, 0x00, 0x64], 2),
            ("negative z", self._sensor_packet("L", z_form=True), self._sensor_packet("L"), 20, [0x88, 0x00, 0x64], 2),
        ]
        for label, duplicate_packet, expected_sensor, getpos_value, final_command, expected_measurements in cases:
            with self.subTest(case=label):
                with tempfile.TemporaryDirectory() as tmpdir:
                    controller = self._build_sampling_controller(tmpdir, [0.0, 0.05, 1.0, 1.05], samples_per_direction=2)

                    self.assertTrue(controller.start(8, "RZ"))
                    self._drive_home_sequence(controller, start_pos=10)
                    controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
                    controller.handle_runtime_packet(self._sensor_packet("R"))
                    self._assert_getpos_sent(controller)

                    if expected_sensor == self._sensor_packet("R"):
                        controller.handle_runtime_packet(duplicate_packet)
                        self._assert_getpos_sent(controller)
                        self._assert_no_dd(controller)
                        controller.handle_runtime_packet(self._getpos_packet(getpos_value))
                        self.assertEqual(controller.commands[-1], final_command)
                        self.assertEqual(len(controller.measurements), expected_measurements)
                        self.assertEqual(controller.failures, [])
                        continue

                    controller.handle_runtime_packet(self._getpos_packet(70))
                    self.assertEqual(controller.commands[-1], [0x88, 0xFF, 0x9C])
                    controller.handle_runtime_packet([0x88, 0x53, 0xFF, 0x9C])
                    controller.handle_runtime_packet(expected_sensor)
                    self._assert_getpos_sent(controller)
                    controller.handle_runtime_packet(duplicate_packet)
                    self._assert_getpos_sent(controller)
                    self._assert_no_dd(controller)
                    controller.handle_runtime_packet(self._getpos_packet(20))
                    self.assertEqual(len(controller.measurements), 2)
                    self.assertEqual(controller.commands[-1], [0x88, 0x00, 0x64])
                    self.assertEqual(controller.failures, [])

    def test_sampling_duplicate_sensor_while_waiting_for_getpos_is_ignored_for_r_home(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            controller = self._build_sampling_controller(
                tmpdir,
                [0.0, 0.05, 1.0, 1.05, 2.0, 2.05],
                samples_per_direction=1,
                home_sensor="R",
            )

            self.assertTrue(controller.start(8, "RZ"))
            self._drive_home_sequence(controller, start_pos=10, home_packet=self._tpos_home_packet("R"))

            controller.handle_runtime_packet([0x88, 0x53, 0xFF, 0x9C])
            controller.handle_runtime_packet(self._sensor_packet("L"))
            self._assert_getpos_sent(controller)
            controller.handle_runtime_packet(self._sensor_packet("L", z_form=True))
            self._assert_getpos_sent(controller)
            self._assert_no_dd(controller)
            controller.handle_runtime_packet(self._getpos_packet(70))

            self.assertEqual(controller.commands[-1], [0x88, 0x00, 0x64])
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet(self._sensor_packet("R"))
            self._assert_getpos_sent(controller)
            controller.handle_runtime_packet(self._sensor_packet("R", z_form=True))
            self._assert_getpos_sent(controller)
            self._assert_no_dd(controller)
            controller.handle_runtime_packet(self._getpos_packet(20))
            self.assertEqual(controller.commands[-1][0], 0x81)
            controller.handle_runtime_packet([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x19])

            self.assertFalse(controller.failures)
            self.assertFalse(controller.is_active())

    def test_sampling_duplicate_departure_sensor_after_run_ack_is_ignored_for_both_home_configurations(self) -> None:
        cases = [
            ("L", 10, 70, 10),
            ("R", 10, -70, 10),
        ]
        for home_sensor, home_pos, opposite_pos, return_pos in cases:
            with self.subTest(home_sensor=home_sensor):
                with tempfile.TemporaryDirectory() as tmpdir:
                    controller = self._build_sampling_controller(
                        tmpdir,
                        [0.0, 0.05, 1.0, 1.05],
                        home_sensor=home_sensor,
                        samples_per_direction=1,
                    )

                    self.assertTrue(controller.start(8, "RZ"))
                    self._drive_home_sequence(
                        controller,
                        start_pos=home_pos,
                        home_packet=self._tpos_home_packet(home_sensor),
                    )

                    outward_velocity = 100 if home_sensor == "L" else -100
                    return_velocity = -outward_velocity
                    departure_sensor = home_sensor
                    arrival_sensor = "R" if home_sensor == "L" else "L"
                    return_departure_sensor = arrival_sensor
                    return_arrival_sensor = departure_sensor

                    self.assertEqual(controller.commands[3], build_run(outward_velocity))
                    controller.handle_runtime_packet([0x88, 0x53, *build_run(outward_velocity)[1:]])
                    controller.handle_runtime_packet(self._sensor_packet(departure_sensor))
                    self.assertTrue(controller.is_active())
                    self.assertEqual(controller.state, controller.S_SAMPLE_WAIT_SENSOR)
                    self._assert_no_dd(controller)
                    self.assertEqual(controller.failures, [])
                    controller.handle_runtime_packet(self._sensor_packet(arrival_sensor))
                    self._assert_getpos_sent(controller)
                    controller.handle_runtime_packet(self._getpos_packet(opposite_pos))

                    self.assertEqual(controller.commands[-1], build_run(return_velocity))
                    controller.handle_runtime_packet([0x88, 0x53, *build_run(return_velocity)[1:]])
                    controller.handle_runtime_packet(self._sensor_packet(return_departure_sensor))
                    self.assertTrue(controller.is_active())
                    self.assertEqual(controller.state, controller.S_SAMPLE_WAIT_SENSOR)
                    self._assert_no_dd(controller)
                    self.assertEqual(controller.failures, [])
                    controller.handle_runtime_packet(self._sensor_packet(return_arrival_sensor))
                    self._assert_getpos_sent(controller)
                    controller.handle_runtime_packet(self._getpos_packet(return_pos))

                    self.assertTrue(controller.is_active())
                    self.assertEqual(controller.failures, [])
                    self.assertEqual(controller.commands[-1][0], 0x81)

    def test_sampling_unexpected_packet_after_run_ack_still_fails_when_not_departure_duplicate(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            controller = self._build_sampling_controller(tmpdir, [0.0, 0.05], samples_per_direction=1)

            self.assertTrue(controller.start(8, "RZ"))
            self._drive_home_sequence(controller, start_pos=10)
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x00])

            self.assertFalse(controller.is_active())
            self.assertEqual(controller.commands[-1], [0xDD])
            self.assertTrue(controller.failures)
            self.assertIn("Unexpected packet", controller.failures[-1])

    def test_sampling_opposite_sensor_while_waiting_for_getpos_still_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            controller = self._build_sampling_controller(tmpdir, [0.0, 0.05], samples_per_direction=1)

            self.assertTrue(controller.start(8, "RZ"))
            self._drive_home_sequence(controller, start_pos=10)
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            controller.handle_runtime_packet([0x81, 0x4C])

            self.assertFalse(controller.is_active())
            self.assertEqual(controller.commands[-1], [0xDD])
            self.assertTrue(controller.failures)
            self.assertIn("Unexpected packet", controller.failures[-1])

    def test_sampling_middle_tpos_l_sensor_is_not_treated_as_success(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            controller = self._build_sampling_controller(tmpdir, [0.0, 0.05, 1.0, 1.4, 2.0, 2.4], home_sensor="L")

            self.assertTrue(controller.start(8, "RZ"))
            self._drive_home_sequence(controller, start_pos=10, home_packet=self._tpos_home_packet("L"))
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 70])
            controller.handle_runtime_packet([0x88, 0x53, 0xFF, 0x9C])
            controller.handle_runtime_packet([0x81, 0x52])
            controller.handle_runtime_packet([0x81, 0x4C])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 10])
            self.assertEqual(controller.commands[-1][0], 0x81)
            controller.handle_runtime_packet([0x81, 0x4C])

            self.assertFalse(controller.is_active())
            self.assertEqual(controller.commands[-1], [0xDD])
            self.assertTrue(controller.failures)
            self.assertIn("Unexpected packet while waiting for TPOS middle completion", controller.failures[-1])

    def test_sampling_controller_runs_two_samples_and_writes_workbook_sections(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            adapter = self._build_adapter(tmpdir)
            clock = _SamplingManualClock([0.0, 0.5, 1.0, 1.4, 2.0, 2.8, 3.0, 3.6])
            controller = _RecordingSamplingController(
                adapter,
                SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=2),
                clock,
            )

            self.assertTrue(controller.start(8, "RZ"))
            self._drive_home_sequence(controller, start_pos=10)

            self.assertEqual(controller.commands[3], [0x88, 0x00, 0x64])
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 70])
            self.assertEqual(controller.commands[5], [0x88, 0xFF, 0x9C])
            controller.handle_runtime_packet([0x88, 0x53, 0xFF, 0x9C])
            controller.handle_runtime_packet([0x81, 0x4C])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 20])

            self.assertEqual(controller.commands[7], [0x88, 0x00, 0x64])
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 90])
            self.assertEqual(controller.commands[9], [0x88, 0xFF, 0x9C])
            controller.handle_runtime_packet([0x88, 0x53, 0xFF, 0x9C])
            controller.handle_runtime_packet([0x81, 0x4C])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 30])
            self.assertEqual(controller.commands[11], build_tpos(50))
            controller.handle_runtime_packet([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x32])

            self.assertFalse(controller.is_active())
            self.assertTrue(controller.completed_called)
            self.assertEqual(controller.failures, [])
            self.assertEqual(controller.aborts, [])
            self.assertEqual(controller.completed_counts[-1], (4, 4))
            self.assertEqual(len(controller.measurements), 4)
            self.assertEqual(controller.measurements[0].range_value, 60)
            self.assertAlmostEqual(controller.measurements[0].elapsed_seconds, 0.5)
            self.assertAlmostEqual(controller.measurements[0].speed, 120.0)
            self.assertEqual(controller.measurements[1].range_value, 50)
            self.assertAlmostEqual(controller.measurements[1].elapsed_seconds, 0.4)
            self.assertAlmostEqual(controller.measurements[1].speed, 125.0)
            self.assertEqual(controller.measurements[2].range_value, 80)
            self.assertAlmostEqual(controller.measurements[3].range_value, 60)

            output_path = Path(tmpdir) / "sampling_completed.xlsx"
            adapter.save_completed_workbook(output_path)
            sampling_sheet = load_workbook(output_path)["3X_D"]

        self.assertEqual(sampling_sheet["B3"].value, 60)
        self.assertEqual(sampling_sheet["B4"].value, 50)
        self.assertAlmostEqual(sampling_sheet["B20"].value, 120.0)
        self.assertAlmostEqual(sampling_sheet["B37"].value, 0.5)
        self.assertEqual(sampling_sheet["C3"].value, 80)
        self.assertEqual(sampling_sheet["C4"].value, 60)
        self.assertAlmostEqual(sampling_sheet["C20"].value, 100.0)
        self.assertAlmostEqual(sampling_sheet["C37"].value, 0.8)

    def test_sampling_time_is_rounded_only_for_output_and_writes(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            adapter = self._build_adapter(tmpdir)
            clock = _SamplingManualClock([1.0, 11.114672391])
            controller = _RecordingSamplingController(
                adapter,
                SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                clock,
            )

            write_calls: list[tuple[str, object]] = []
            original_write = adapter.write_sampling_result

            def recording_write(section: str, pwm: int, direction: str, sample_index: int, value, *, base_group: str | None = None):
                if section == "Time":
                    write_calls.append((section, value))
                return original_write(section, pwm, direction, sample_index, value, base_group=base_group)

            adapter.write_sampling_result = recording_write  # type: ignore[assignment]

            self.assertTrue(controller.start(8, "RZ"))
            self.assertEqual(controller.commands[0], [0x84, 0x00, 0x50])
            controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
            self.assertEqual(controller.commands[1], [0x81, 0x00, 0x00, 0x00, 0x00])
            controller.handle_runtime_packet([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x00])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 10])
            self.assertEqual(controller.commands[3], [0x88, 0x00, 0x64])
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x03, 0xE8])

            self.assertTrue(write_calls)
            self.assertEqual(write_calls[0][0], "Time")
            self.assertAlmostEqual(float(write_calls[0][1]), 10.115)
            self.assertAlmostEqual(controller.measurements[0].elapsed_seconds, 10.114672391)
            self.assertAlmostEqual(controller.measurements[0].speed, 990 / 10.114672391, places=9)

    def test_sampling_timeout_sends_dd_and_stops_loop(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            adapter = self._build_adapter(tmpdir)
            clock = _SamplingManualClock([0.0, 0.05, 1.0])
            controller = _RecordingSamplingController(
                adapter,
                SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                clock,
            )

            self.assertTrue(controller.start(8, "RZ"))
            self._drive_home_sequence(controller, start_pos=10)
            self.assertEqual(controller.commands[0], [0x84, 0x00, 0x50])
            self.assertEqual(controller.commands[1], [0x81, 0x00, 0x00, 0x00, 0x00])
            self.assertEqual(controller.commands[2], [0x82])
            self.assertEqual(controller.commands[3], [0x88, 0x00, 0x64])
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.on_timeout()

            self.assertFalse(controller.is_active())
            self.assertEqual(controller.state, controller.S_FAILED)
            self.assertEqual(controller.commands[-1], [0xDD])
            self.assertTrue(controller.failures)
            self.assertIn("Timed out", controller.failures[-1])
            self.assertTrue(controller.can_resume)
            self.assertTrue(
                controller.resume_availability(node_id=8, node_name="RZ", base_group="3X")[0]
            )
            self.assertEqual(controller.commands.count([0xDD]), 1)
            controller.on_timeout()
            self.assertEqual(controller.commands.count([0xDD]), 1)
            controller.handle_runtime_packet([0x81, 0x52])
            self.assertEqual(controller.commands.count([0xDD]), 1)

    def test_sampling_middle_timeout_sends_dd_and_fails_after_final_sample(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            adapter = self._build_adapter(tmpdir)
            clock = _SamplingManualClock([0.0, 0.01, 1.0, 1.4])
            controller = _RecordingSamplingController(
                adapter,
                SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                clock,
            )

            self.assertTrue(controller.start(8, "RZ"))
            self.assertEqual(controller.commands[0], [0x84, 0x00, 0x50])
            controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
            self.assertEqual(controller.commands[1], [0x81, 0x00, 0x00, 0x00, 0x00])
            controller.handle_runtime_packet([0x81, 0x53, 0x82, 0x00, 0x00, 0x00, 0x00])
            controller.handle_runtime_packet([0x81, 0x45, 0x82, 0x00, 0x00, 0x00, 0x00])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 10])
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 20])
            controller.handle_runtime_packet([0x88, 0x53, 0xFF, 0x9C])
            controller.handle_runtime_packet([0x81, 0x4C])
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 10])

            self.assertEqual(controller.commands[-1][0], 0x81)
            controller.on_timeout()

            self.assertFalse(controller.is_active())
            self.assertEqual(controller.state, controller.S_FAILED)
            self.assertEqual(controller.commands[-1], [0xDD])
            self.assertTrue(controller.failures)
            self.assertIn("Timed out", controller.failures[-1])

    def test_sampling_encoder_reset_is_non_resumable_and_does_not_repeat_dd(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            adapter = self._build_adapter(tmpdir)
            controller = _RecordingSamplingController(
                adapter,
                SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                _SamplingManualClock([0.0, 0.05, 1.0]),
            )

            self.assertTrue(controller.start(8, "RZ"))
            self._drive_home_sequence(controller, start_pos=10)
            controller.handle_runtime_packet([0x88, 0x53, 0x00, 0x64])
            controller.handle_runtime_packet([0x81, 0x52])
            self.assertEqual(controller.commands[-1], [0x82])
            controller.handle_runtime_packet([0x81, 0x49])

            self.assertFalse(controller.is_active())
            self.assertEqual(controller.state, controller.S_FAILED)
            self.assertEqual(controller.commands[-1], [0xDD])
            self.assertEqual(controller.commands.count([0xDD]), 1)
            self.assertTrue(controller.failures)
            self.assertIn("Unexpected encoder reset", controller.failures[-1])
            self.assertFalse(controller.can_resume)
            self.assertFalse(controller.resume_availability(node_id=8, node_name="RZ", base_group="3X")[0])
            self.assertIsNotNone(controller.last_terminal_result)
            self.assertEqual(controller.last_terminal_result.final_status, "FAILED")
            self.assertEqual(controller.last_terminal_result.reason, "Unexpected encoder reset during sampling.")
            self.assertEqual(controller.last_terminal_result.failure_context, "PWM 100 | Direction + | Sample 1")
            self.assertEqual(controller.last_terminal_result.resume_text, "Unavailable - encoder reset requires a fresh start.")

            controller.on_timeout()
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 20])
            controller.handle_runtime_packet([0x84, 0x53, 0x00, 0x50])
            self.assertEqual(controller.commands.count([0xDD]), 1)

    def test_sampling_timeout_then_abort_does_not_duplicate_dd(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            adapter = self._build_adapter(tmpdir)
            clock = _SamplingManualClock([0.0])
            controller = _RecordingSamplingController(
                adapter,
                SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                clock,
            )

            self.assertTrue(controller.start(8, "RZ"))
            controller.on_timeout()
            dd_count_after_timeout = controller.commands.count([0xDD])
            self.assertEqual(dd_count_after_timeout, 1)

            self.assertFalse(controller.abort_by_user())
            self.assertEqual(controller.commands.count([0xDD]), 1)

    def test_sampling_rejects_unexpected_packet_while_waiting_for_ack(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            adapter = self._build_adapter(tmpdir)
            clock = _SamplingManualClock([0.0])
            controller = _RecordingSamplingController(
                adapter,
                SamplingTestConfig(home_velocity=-190, pwm_values=(100,), samples_per_direction=1),
                clock,
            )

            self.assertTrue(controller.start(8, "RZ"))
            controller.handle_runtime_packet([0x82, 0x00, 0x00, 0x00, 1])

            self.assertFalse(controller.is_active())
            self.assertEqual(controller.commands[-1], [0xDD])
            self.assertTrue(controller.failures)
            self.assertIn("Unexpected packet", controller.failures[-1])
