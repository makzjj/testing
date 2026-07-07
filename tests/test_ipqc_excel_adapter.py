"""Unit tests for IPQC Excel adapter foundation behavior."""

from __future__ import annotations

import re
import tempfile
import unittest
from pathlib import Path

from services.ipqc_excel_adapter import IpqcExcelAdapter

try:
    from openpyxl import Workbook, load_workbook

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - environment dependent.
    _HAS_OPENPYXL = False


@unittest.skipUnless(_HAS_OPENPYXL, "openpyxl is required for IPQC Excel adapter tests.")
class IpqcExcelAdapterTests(unittest.TestCase):
    def _create_ipqc_template(self, base_dir: str | Path, filename: str = "ipqc_template.xlsx") -> Path:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "3X"
        sampling_3x = workbook.create_sheet("3X_D")
        workbook.create_sheet("3X_A")
        workbook.create_sheet("4Y")
        sampling_4y = workbook.create_sheet("4Y_D")
        workbook.create_sheet("4Y_A")
        summary["A1"] = "Programming"
        summary["B2"] = "Source"
        summary["C2"] = "Programmed"
        summary["D2"] = "Check"
        summary["A3"] = "Operator"
        summary["A4"] = "UUID"
        summary["A5"] = "PWM"
        summary["A6"] = "Proportionate (P)"
        summary["A7"] = "Integral (I)"
        summary["A8"] = "Derivative (D)"
        summary["A9"] = "PID_SlewRate"
        summary["A10"] = "RampDown_Slope"
        summary["A11"] = "RampDown_Step"
        summary["A12"] = "RampDown_MinVel"
        summary["A13"] = "RampDown_TargetOffset"
        summary["A14"] = "RampDown_Region"
        summary["A15"] = "Acceptable_Error"
        summary["B3"] = "operator-a"
        summary["B4"] = "1223303010"
        summary["B5"] = "100"
        summary["B6"] = "N/A"
        self._populate_sampling_sheet(sampling_3x)
        self._populate_sampling_sheet(sampling_4y)
        path = Path(base_dir) / filename
        workbook.save(path)
        return path

    def _populate_sampling_sheet(self, sheet) -> None:
        pwm_values = [100, 90, 80, 70, 60]
        section_starts = {"Range": 1, "Speed": 18, "Time": 35}
        for section_name, start_row in section_starts.items():
            sheet[f"A{start_row}"] = section_name
            row = start_row + 1
            for pwm in pwm_values:
                sheet[f"A{row}"] = f"PWM {pwm}"
                sheet[f"A{row + 1}"] = f"+{pwm}"
                sheet[f"A{row + 2}"] = f"-{pwm}"
                row += 3

    def test_load_template_detects_base_sheet_groups_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            groups = adapter.load_template(template_path)

        self.assertEqual(groups, ["3X", "4Y"])
        self.assertEqual(adapter.active_sheet_group, "3X")

    def test_select_missing_sheet_group_raises_clear_error(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            with self.assertRaisesRegex(ValueError, "not available"):
                adapter.select_sheet_group("9PZ")

    def test_read_expected_summary_reads_sn_pwm_and_optional_values(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            expected = adapter.read_expected_summary(strict=True)

        self.assertEqual(expected.serial_number, "1223303010")
        self.assertEqual(expected.pwm, "100")
        self.assertEqual(expected.operator, "operator-a")
        self.assertEqual(expected.assembler, "")
        self.assertEqual(expected.other_parameters, "N/A")

    def test_read_expected_uuid_serial_reads_b4_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            workbook = load_workbook(template_path)
            workbook["3X"]["B5"] = None
            workbook["3X"]["B6"] = None
            workbook.save(template_path)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            serial = adapter.read_expected_uuid_serial()

        self.assertEqual(serial, "1223303010")

    def test_read_expected_pwm_value_reads_b5(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            pwm_value = adapter.read_expected_pwm_value()

        self.assertEqual(pwm_value, "100")

    def test_write_production_metadata_inserts_assembler_row_and_preserves_parameter_lookup(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            adapter.write_production_metadata("  operator-b  ", " assembler-b ")
            discovered_rows, _raw_labels = adapter.discover_programming_parameter_rows()
            output_path = Path(tmpdir) / "ipqc_completed.xlsx"
            adapter.save_completed_workbook(output_path)
            summary = load_workbook(output_path)["3X"]

        self.assertEqual(summary["A4"].value, "Assembler")
        self.assertEqual(summary["B3"].value, "operator-b")
        self.assertEqual(summary["B4"].value, "assembler-b")
        self.assertEqual(summary["A5"].value, "UUID")
        self.assertEqual(summary["B5"].value, "1223303010")
        self.assertEqual(summary["A6"].value, "PWM")
        self.assertEqual(summary["B6"].value, "100")
        self.assertEqual(discovered_rows["uuid"], 5)
        self.assertEqual(discovered_rows["pwm"], 6)

    def test_read_production_metadata_reads_inserted_assembler_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            workbook = load_workbook(template_path)
            summary = workbook["3X"]
            summary.insert_rows(4, 1)
            summary["A4"] = "Assembler"
            summary["B3"] = "operator-a"
            summary["B4"] = "assembler-a"
            workbook.save(template_path)

            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            metadata = adapter.read_production_metadata()

        self.assertEqual(metadata.operator_name, "operator-a")
        self.assertEqual(metadata.assembler_name, "assembler-a")

    def test_missing_required_expected_cell_is_reported_clearly(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            workbook = load_workbook(template_path)
            workbook["3X"]["B4"] = None
            workbook.save(template_path)

            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            with self.assertRaisesRegex(ValueError, "serial number/UUID"):
                adapter.read_expected_summary(strict=True)

    def test_uuid_pwm_summary_wrapper_helpers_are_removed(self) -> None:
        adapter = IpqcExcelAdapter()
        self.assertFalse(hasattr(adapter, "write_uuid_actual_and_check"))
        self.assertFalse(hasattr(adapter, "write_pwm_actual_and_check"))

    def test_write_summary_result_maps_uuid_and_pwm_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            adapter.write_summary_result("S/N", "1223303011", "PASS")
            adapter.write_summary_result("PWM", "98", "FAIL")
            output_path = Path(tmpdir) / "ipqc_completed.xlsx"
            adapter.save_completed_workbook(output_path)

            completed = load_workbook(output_path)
            summary = completed["3X"]

        self.assertEqual(summary["C4"].value, "1223303011")
        self.assertEqual(summary["D4"].value, "PASS")
        self.assertEqual(summary["C5"].value, "98")
        self.assertEqual(summary["D5"].value, "FAIL")

    def test_write_summary_result_maps_sn_pwm_and_other_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            adapter.write_summary_result("serial", "1223303012", "PASS")
            adapter.write_summary_result("PWM", "101", "FAIL")
            adapter.write_summary_result("Other parameters", "N/A", "PASS")
            output_path = Path(tmpdir) / "ipqc_completed.xlsx"
            adapter.save_completed_workbook(output_path)
            summary = load_workbook(output_path)["3X"]

        self.assertEqual(summary["C4"].value, "1223303012")
        self.assertEqual(summary["D4"].value, "PASS")
        self.assertEqual(summary["C5"].value, "101")
        self.assertEqual(summary["D5"].value, "FAIL")
        self.assertEqual(summary["C6"].value, "N/A")
        self.assertEqual(summary["D6"].value, "PASS")

    def test_write_summary_result_rejects_unknown_parameter(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            with self.assertRaisesRegex(ValueError, "Unsupported summary parameter"):
                adapter.write_summary_result("temperature", "40", "PASS")

    def test_resolve_programming_row_maps_new_layout_labels(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)

            self.assertEqual(adapter.resolve_programming_row("UUID"), 4)
            self.assertEqual(adapter.resolve_programming_row("Proportionate (P)"), 6)
            self.assertEqual(adapter.resolve_programming_row("PID_SlewRate"), 9)
            self.assertEqual(adapter.resolve_programming_row("RampDown_TargetOffset"), 13)

    def test_discover_programming_parameter_rows_scans_column_a_labels(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)

            discovered_rows, raw_labels = adapter.discover_programming_parameter_rows()

        self.assertEqual(raw_labels, ["Operator", "UUID", "PWM", "Proportionate (P)", "Integral (I)", "Derivative (D)", "PID_SlewRate", "RampDown_Slope", "RampDown_Step", "RampDown_MinVel", "RampDown_TargetOffset", "RampDown_Region", "Acceptable_Error"])
        self.assertEqual(discovered_rows["uuid"], 4)
        self.assertEqual(discovered_rows["pwm"], 5)
        self.assertEqual(discovered_rows["proportionate (p)"], 6)
        self.assertEqual(discovered_rows["integral (i)"], 7)
        self.assertEqual(discovered_rows["derivative (d)"], 8)
        self.assertEqual(discovered_rows["pid_slewrate"], 9)
        self.assertEqual(discovered_rows["rampdown_slope"], 10)
        self.assertEqual(discovered_rows["rampdown_step"], 11)
        self.assertEqual(discovered_rows["rampdown_minvel"], 12)
        self.assertEqual(discovered_rows["rampdown_targetoffset"], 13)
        self.assertEqual(discovered_rows["rampdown_region"], 14)
        self.assertEqual(discovered_rows["acceptable_error"], 15)

    def test_discover_sampling_layout_maps_d_sheet_rows_and_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)

            self.assertEqual(adapter.resolve_sampling_sheet_name(), "3X_D")
            self.assertEqual(adapter.sample_index_to_column(1), "B")
            self.assertEqual(adapter.sample_index_to_column(32), "AG")

            layout = adapter.discover_sampling_layout()

        self.assertEqual(layout.sheet_name, "3X_D")
        self.assertEqual(layout.section_headers["range"], 1)
        self.assertEqual(layout.section_headers["speed"], 18)
        self.assertEqual(layout.section_headers["time"], 35)
        self.assertEqual(layout.resolve_row("Range", 100, "+"), 3)
        self.assertEqual(layout.resolve_row("Range", 100, "-"), 4)
        self.assertEqual(layout.resolve_row("Speed", 90, "+"), 23)
        self.assertEqual(layout.resolve_row("Time", 60, "-"), 50)
        self.assertIn("Range", layout.raw_labels)
        self.assertIn("+100", layout.raw_labels)
        self.assertIn("-60", layout.raw_labels)

    def test_discover_sampling_layout_rejects_missing_sampling_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            workbook = load_workbook(template_path)
            del workbook["3X_D"]
            workbook.save(template_path)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)

            with self.assertRaisesRegex(ValueError, "Sampling sheet '3X_D' is missing from workbook"):
                adapter.discover_sampling_layout()

    def test_discover_sampling_layout_rejects_missing_range_section(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            workbook = load_workbook(template_path)
            workbook["3X_D"]["A1"] = ""
            workbook.save(template_path)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)

            with self.assertRaisesRegex(ValueError, "missing required section 'Range'"):
                adapter.discover_sampling_layout()

    def test_discover_sampling_layout_rejects_missing_speed_section(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            workbook = load_workbook(template_path)
            workbook["3X_D"]["A18"] = ""
            workbook.save(template_path)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)

            with self.assertRaisesRegex(ValueError, "missing required section 'Speed'"):
                adapter.discover_sampling_layout()

    def test_discover_sampling_layout_rejects_missing_time_section(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            workbook = load_workbook(template_path)
            workbook["3X_D"]["A35"] = ""
            workbook.save(template_path)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)

            with self.assertRaisesRegex(ValueError, "missing required section 'Time'"):
                adapter.discover_sampling_layout()

    def test_discover_sampling_layout_rejects_missing_pwm_and_sample_labels(self) -> None:
        cases = [
            ("A2", "missing label 'PWM 100'"),
            ("A3", "missing label '+100'"),
            ("A4", "missing label '-100'"),
        ]
        for cell_ref, error_text in cases:
            with self.subTest(cell_ref=cell_ref):
                with tempfile.TemporaryDirectory() as tmpdir:
                    template_path = self._create_ipqc_template(tmpdir)
                    workbook = load_workbook(template_path)
                    workbook["3X_D"][cell_ref] = ""
                    workbook.save(template_path)
                    adapter = IpqcExcelAdapter()
                    adapter.load_template(template_path)

                    with self.assertRaisesRegex(ValueError, re.escape(error_text)):
                        adapter.discover_sampling_layout()

    def test_discover_sampling_layout_rejects_formula_in_raw_sample_cell(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            workbook = load_workbook(template_path)
            workbook["3X_D"]["B3"] = "=SUM(B4:B5)"
            workbook.save(template_path)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)

            with self.assertRaisesRegex(ValueError, "raw sample cell 'B3'.*contains a formula"):
                adapter.discover_sampling_layout()

    def test_write_sampling_result_writes_to_sampling_sheet_cells(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            cell_range = adapter.write_sampling_result("Range", 100, "+", 1, 512)
            cell_speed = adapter.write_sampling_result("Speed", 100, "+", 32, 123.5)
            cell_time = adapter.write_sampling_result("Time", 60, "-", 4, 0.456)
            output_path = Path(tmpdir) / "ipqc_completed.xlsx"
            adapter.save_completed_workbook(output_path)
            sampling = load_workbook(output_path)["3X_D"]

        self.assertEqual(cell_range, "B3")
        self.assertEqual(cell_speed, "AG20")
        self.assertEqual(cell_time, "E50")
        self.assertEqual(sampling["B3"].value, 512)
        self.assertEqual(sampling["AG20"].value, 123.5)
        self.assertEqual(sampling["E50"].value, 0.456)

    def test_clear_sampling_results_removes_stale_values_for_supported_sample_counts_and_preserves_formulas(self) -> None:
        for sample_count in (2, 4, 8, 16, 32):
            with self.subTest(sample_count=sample_count):
                with tempfile.TemporaryDirectory() as tmpdir:
                    template_path = self._create_ipqc_template(tmpdir)
                    adapter = IpqcExcelAdapter()
                    adapter.load_template(template_path)
                    sampling = adapter._workbook["3X_D"]
                    for sample_index in range(1, sample_count + 1):
                        column = adapter.sample_index_to_column(sample_index)
                        sampling[f"{column}3"] = sample_index
                        sampling[f"{column}20"] = sample_index * 10
                        sampling[f"{column}37"] = float(sample_index) / 10.0
                    if sample_count < 32:
                        stale_column = adapter.sample_index_to_column(sample_count + 1)
                        sampling[f"{stale_column}3"] = 999
                        sampling[f"{stale_column}20"] = 999
                        sampling[f"{stale_column}37"] = 999.0
                    sampling["AH3"] = "=SUM(B3:AG3)"
                    sampling["AH20"] = "=SUM(B20:AG20)"
                    sampling["AH37"] = "=SUM(B37:AG37)"

                    cleared_cells = adapter.clear_sampling_results()
                    output_path = Path(tmpdir) / f"cleared_{sample_count}.xlsx"
                    adapter.save_completed_workbook(output_path)
                    reloaded = load_workbook(output_path, data_only=False)["3X_D"]

                self.assertEqual(cleared_cells, 3 * 5 * 2 * 32)
                for sample_index in range(1, 33):
                    column = adapter.sample_index_to_column(sample_index)
                    self.assertIsNone(reloaded[f"{column}3"].value)
                    self.assertIsNone(reloaded[f"{column}20"].value)
                    self.assertIsNone(reloaded[f"{column}37"].value)
                self.assertEqual(reloaded["AH3"].value, "=SUM(B3:AG3)")
                self.assertEqual(reloaded["AH20"].value, "=SUM(B20:AG20)")
                self.assertEqual(reloaded["AH37"].value, "=SUM(B37:AG37)")

    def test_write_programming_parameter_result_writes_row_cells(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            adapter.write_programming_parameter_result("RampDown_Region", "75", "PASS")
            output_path = Path(tmpdir) / "ipqc_completed.xlsx"
            adapter.save_completed_workbook(output_path)
            summary = load_workbook(output_path)["3X"]

        self.assertEqual(summary["C14"].value, "75")
        self.assertEqual(summary["D14"].value, "PASS")

    def test_save_completed_workbook_uses_new_path_and_preserves_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            adapter.write_summary_result("S/N", "1223303011", "PASS")
            output_path = Path(tmpdir) / "ipqc_completed.xlsx"
            saved_path = adapter.save_completed_workbook(output_path)

            original = load_workbook(template_path)["3X"]
            completed = load_workbook(saved_path)["3X"]

        self.assertEqual(saved_path, output_path.resolve())
        self.assertIsNone(original["C4"].value)
        self.assertEqual(completed["C4"].value, "1223303011")
        self.assertEqual(completed["D4"].value, "PASS")

    def test_save_rejects_overwriting_original_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir)
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            with self.assertRaisesRegex(ValueError, "overwrite"):
                adapter.save_completed_workbook(template_path)

    def test_suggest_completed_output_path_is_in_runtime_exports_directory(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = self._create_ipqc_template(tmpdir, filename="line_a.xlsx")
            adapter = IpqcExcelAdapter()
            adapter.load_template(template_path)
            suggested = adapter.suggest_completed_output_path()

        self.assertEqual(suggested.parent, Path.cwd() / "data" / "exports")
        self.assertTrue(suggested.name.startswith("line_a_3X_completed_"))
        self.assertEqual(suggested.suffix, ".xlsx")


if __name__ == "__main__":
    unittest.main()
